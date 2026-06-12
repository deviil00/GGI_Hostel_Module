import io
import csv
import json
import uuid
import hmac
import hashlib
import base64
import openpyxl
from collections import OrderedDict
from django.core import signing
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date as dt_date
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db import models
from django.db.models import Count, Q, Sum, Avg, Max, Case, When, IntegerField, F, Prefetch
from django.urls import reverse
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST

from apps.hostel.models import (
    Hostel, Room, Student, RoomAllocation, Attendance,
    RoomAmenity, RoomAmenityRecord, Announcement,
    RegistrationForm, FormCategory, FormField, RegistrationSubmission, SubmissionResponse, RoomPreference,
    GatePass, Bed, Notification, Visitor, MessMenu, MessFeedback,
    MessDailyMenu, MessWastageRecord, MessItemWastage, DisciplineRecord, MaintenanceLog,
    Semester, RoomAsset, LeaveApplication,
    AnnouncementCategory, AnnouncementAudience, Department,
    RoleModulePermission, CustomStaffRole, StaffProfile, STAFF_MODULES, STUDENT_MODULES, DEFAULT_ROLE_MODULES,
    ROOM_CATEGORIES_GROUPED, ROOM_CATEGORY_CAPACITY, ROOM_CATEGORY_LABELS,
    GatePassCategory, AcademicCalendarEvent,
    LeaveCategory, LeaveExtensionRequest,
    AssetType, Asset, AssetTransfer, AssetLog,
    StudentRestriction,
    DisciplineCategory, DisciplineRecord, DisciplineEvidence, DisciplineAction,
    QuotaPolicy,
)
from apps.hostel.forms import (
    HostelForm, RoomForm, AllocateRoomForm,
    StudentImportForm, CheckoutForm
)
from apps.fees.models import FeeStructure, FeeRecord
from apps.complaints.models import Complaint, ComplaintTimeline
from apps.accounts.models import User


# ─── helpers ─────────────────────────────────────────────────────────────────

def get_hostel_scope(user):
    """
    Returns the Hostel the warden manages, or None if the user is an admin
    (meaning they have access to all hostels).
    """
    if user.role == User.Role.WARDEN:
        return user.managed_hostels.first()
    return None


def _staff_targets(student=None):
    """Return all active admins + the warden of the student's current hostel."""
    targets = list(User.objects.filter(role=User.Role.ADMIN, is_active=True))
    if student:
        alloc = student.allocations.filter(status='active').select_related('room__hostel__warden').first()
        if alloc and alloc.room.hostel.warden:
            w = alloc.room.hostel.warden
            if w not in targets:
                targets.append(w)
    return targets


def admin_required(view_fn):
    """Decorator: only admin or warden can access."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_admin_or_warden:
            messages.error(request, 'Access denied.')
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    return wrapper


def admin_only(view_fn):
    """Decorator: only admin or superadmin can access."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role not in (User.Role.ADMIN, User.Role.SUPER_ADMIN):
            messages.error(request, 'Only admins can access this page.')
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


def superadmin_only(view_fn):
    """Decorator: only superadmin can access."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_super_admin:
            messages.error(request, 'Admin 1 access only.')
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


def reports_download_access(view_fn):
    """Decorator: superadmin always has access; other roles need 'reports_download' module permission."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.is_super_admin:
            return view_fn(request, *args, **kwargs)
        from apps.hostel.models import RoleModulePermission
        allowed = RoleModulePermission.get_for_role(request.user.role)
        if 'reports_download' in allowed:
            return view_fn(request, *args, **kwargs)
        messages.error(request, 'You do not have permission to access Download Reports.')
        return redirect('dashboard')
    wrapper.__name__ = view_fn.__name__
    return wrapper


def mess_only(view_fn):
    """Decorator: mess incharge role (plus superadmin who can preview)."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role not in (User.Role.MESS, User.Role.SUPER_ADMIN, User.Role.ADMIN):
            messages.error(request, 'Mess Incharge access only.')
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


def maintenance_only(view_fn):
    """Decorator: maintenance incharge + warden + admin + superadmin."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role not in (
            User.Role.MAINTENANCE, User.Role.SUPER_ADMIN,
            User.Role.ADMIN, User.Role.WARDEN,
        ):
            messages.error(request, 'Maintenance Incharge access only.')
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


def _get_student_or_redirect(request):
    """
    Return the Student profile for the logged-in user.
    Returns (student, None) on success, (None, redirect_response) if no profile found.
    """
    try:
        return Student.objects.get(user=request.user), None
    except Student.DoesNotExist:
        messages.error(
            request,
            'No student profile is linked to your account. '
            'Please contact the hostel office.'
        )
        return None, redirect('login')


# ─── dashboard ───────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    if request.user.role == User.Role.STUDENT:
        return redirect('student_dashboard')
    if request.user.role == User.Role.SECURITY:
        return redirect('security_dashboard')
    if request.user.role == User.Role.SUPER_ADMIN:
        return redirect('superadmin_dashboard')
    if request.user.role in (User.Role.ADMIN, User.Role.WARDEN):
        return redirect('admin_dashboard')
    # maintenance, mess, and any custom roles all go to staff dashboard
    return redirect('staff_dashboard')


@admin_required
def admin_dashboard(request):
    today = dt_date.today()
    total_residents = Student.objects.filter(is_resident=True).count()
    present_today   = Attendance.objects.filter(date=today, status='present').count()
    absent_today    = Attendance.objects.filter(date=today, status='absent').count()
    leave_today     = Attendance.objects.filter(date=today, status='leave').count()
    marked_today    = Attendance.objects.filter(date=today).count()

    stats = {
        'total_students':  total_residents,
        'total_rooms':     Room.objects.count(),
        'vacant_rooms':    Room.objects.filter(status='vacant').count(),
        'occupied_rooms':  Room.objects.filter(status='occupied').count(),
        'open_complaints': Complaint.objects.filter(status='open').count(),
        'present_today':   present_today,
        'absent_today':    absent_today,
        'leave_today':     leave_today,
        'unmarked_today':  total_residents - marked_today,
        'attendance_done': marked_today > 0,
    }
    recent_allocations = RoomAllocation.objects.filter(
        status='active'
    ).select_related('student', 'room__hostel').order_by('-created_at')[:5]

    recent_complaints = Complaint.objects.filter(
        status='open'
    ).select_related('student').order_by('-created_at')[:5]

    announcements = Announcement.objects.filter(
        is_active=True
    ).filter(Q(expires_at__isnull=True) | Q(expires_at__gte=today)).order_by('-created_at')[:5]

    return render(request, 'admin/dashboard.html', {
        'stats':              stats,
        'today':              today,
        'recent_allocations': recent_allocations,
        'recent_complaints':  recent_complaints,
        'announcements':      announcements,
    })


@login_required
def student_dashboard(request):
    student, _redir = _get_student_or_redirect(request)
    if _redir: return _redir
    allocation = RoomAllocation.objects.filter(
        student=student, status='active'
    ).select_related('room__hostel').first()

    complaints = Complaint.objects.filter(student=student).order_by('-created_at')

    today         = dt_date.today()
    announcements = Announcement.objects.filter(
        is_active=True
    ).filter(Q(expires_at__isnull=True) | Q(expires_at__gte=today)).order_by('-created_at')

    roommates = []
    if allocation:
        roommates = list(
            RoomAllocation.objects.filter(room=allocation.room, status='active')
            .exclude(student=student)
            .select_related('student')
        )

    return render(request, 'student/dashboard.html', {
        'student':       student,
        'allocation':    allocation,
        'complaints':    complaints,
        'announcements': announcements,
        'roommates':     roommates,
    })


# ─── student management ───────────────────────────────────────────────────────

def _get_dept_objects_with_counts():
    depts = list(Department.objects.order_by('name'))
    counts = dict(
        Student.objects.values_list('department')
        .annotate(cnt=Count('id')).values_list('department', 'cnt')
    )
    for d in depts:
        d.student_count = counts.get(d.name, 0)
    return depts


@admin_required
def student_list(request):
    user          = request.user
    is_admin      = user.role in (User.Role.ADMIN, User.Role.SUPER_ADMIN)
    warden_hostel = get_hostel_scope(user)

    # ── Course / Department management (admin only, POST) ──
    if request.method == 'POST' and is_admin:
        action = request.POST.get('action', '')
        if action == 'add_dept':
            dname = request.POST.get('dept_name', '').strip()
            dcode = request.POST.get('dept_code', '').strip().upper()
            if dname:
                existing = Department.objects.filter(name__iexact=dname).first()
                if not existing:
                    Department.objects.create(name=dname, code=dcode, created_by=user)
                    messages.success(request, f'Course "{dname}" added.')
                else:
                    messages.warning(request, f'Course "{dname}" already exists.')
            return redirect('student_list')

        if action == 'toggle_dept':
            dpk = request.POST.get('dept_pk', '').strip()
            if dpk:
                dept_obj = get_object_or_404(Department, pk=dpk)
                dept_obj.is_active = not dept_obj.is_active
                dept_obj.save(update_fields=['is_active'])
                messages.success(request, f'Course "{dept_obj.name}" {"activated" if dept_obj.is_active else "deactivated"}.')
            return redirect('student_list')

        if action == 'delete_dept':
            dpk = request.POST.get('dept_pk', '').strip()
            if dpk:
                dept_obj = get_object_or_404(Department, pk=dpk)
                dept_obj.delete()
                messages.success(request, 'Course deleted.')
            return redirect('student_list')

    qs = Student.objects.order_by('roll_number')
    if warden_hostel:
        qs = qs.filter(allocations__room__hostel=warden_hostel, allocations__status='active').distinct()

    search        = request.GET.get('q', '').strip()
    hostel_filter = request.GET.get('hostel', '').strip()
    dept          = request.GET.get('dept', '')
    year          = request.GET.get('year', '')
    status        = request.GET.get('status', '')
    state_filter      = request.GET.get('state', '').strip()
    country_filter    = request.GET.get('country', '').strip()
    is_active_filter  = request.GET.get('is_active', '').strip()
    entry_filter      = request.GET.get('entry_type', '').strip()
    reporting_filter  = request.GET.get('reporting_date', '').strip()

    if search:
        qs = qs.filter(
            Q(roll_number__icontains=search) |
            Q(name__icontains=search) |
            Q(department__icontains=search)
        )
    if dept:   qs = qs.filter(department__icontains=dept)
    if year:   qs = qs.filter(year=year)
    if status == 'resident':    qs = qs.filter(is_resident=True)
    if status == 'nonresident': qs = qs.filter(is_resident=False)
    if hostel_filter:
        if hostel_filter == 'unallocated':
            qs = qs.exclude(allocations__status='active')
        else:
            qs = qs.filter(allocations__room__hostel__pk=hostel_filter, allocations__status='active').distinct()
    if state_filter:      qs = qs.filter(state__iexact=state_filter)
    if country_filter:    qs = qs.filter(country__iexact=country_filter)
    if is_active_filter == '1':  qs = qs.filter(is_active=True)
    if is_active_filter == '0':  qs = qs.filter(is_active=False)
    if entry_filter:      qs = qs.filter(type_of_entry__iexact=entry_filter)
    if reporting_filter:
        import datetime as _dt
        try:
            _rd = _dt.date.fromisoformat(reporting_filter)
            qs = qs.filter(reporting_date=_rd)
        except ValueError:
            pass

    # Master list (from Department model) + any extra from existing student profiles
    master_depts  = list(Department.objects.values_list('name', flat=True))
    student_depts = list(
        Student.objects.values_list('department', flat=True)
        .exclude(department__isnull=True).exclude(department__exact='').distinct()
    )
    all_departments = sorted(set(master_depts) | set(student_depts))
    filter_departments = qs.values_list('department', flat=True).distinct().order_by('department')

    # Build full list — no paginator, grouped hostel-wise
    all_students = list(qs)
    alloc_map = {
        a.student_id: a
        for a in RoomAllocation.objects.filter(
            student_id__in=[s.pk for s in all_students], status='active'
        ).select_related('room__hostel')
    }
    for s in all_students:
        a = s.alloc = alloc_map.get(s.pk)
        s.alloc_room    = a.room if a else None
        s.alloc_hostel  = a.room.hostel if a else None
        s.hostel_group  = s.alloc_hostel.name if s.alloc_hostel else 'Unallocated'

    # Sort: allocated hostels alphabetically, unallocated last
    all_students.sort(key=lambda s: (
        'zzz' if not s.alloc_hostel else s.hostel_group,
        s.roll_number
    ))

    all_hostels = Hostel.objects.order_by('type', 'name')

    states       = Student.objects.exclude(state='').values_list('state', flat=True).distinct().order_by('state')
    countries    = Student.objects.exclude(country='').exclude(country__isnull=True).values_list('country', flat=True).distinct().order_by('country')
    entry_types  = Student.objects.exclude(type_of_entry='').exclude(type_of_entry__isnull=True).values_list('type_of_entry', flat=True).distinct().order_by('type_of_entry')

    return render(request, 'admin/students.html', {
        'page_students':     all_students,
        'student_count':     len(all_students),
        'search':            search,
        'departments':       filter_departments,
        'all_departments':   all_departments,
        'dept_objects':      _get_dept_objects_with_counts(),
        'filters':           {
            'dept': dept, 'year': year, 'status': status, 'hostel': hostel_filter,
            'state': state_filter, 'country': country_filter, 'is_active': is_active_filter,
            'entry_type': entry_filter, 'reporting_date': reporting_filter,
        },
        'warden_hostel':     warden_hostel,
        'is_admin':          is_admin,
        'all_hostels':       all_hostels,
        'states':            states,
        'countries':         countries,
        'entry_types':       entry_types,
    })


@admin_only
def add_student(request):
    """Add a single student manually."""
    import datetime as _dt
    if request.method == 'POST':
        roll   = request.POST.get('roll_number', '').strip()
        name   = request.POST.get('name', '').strip()
        if not roll or not name:
            messages.error(request, 'Roll Number and Name are required.')
            return redirect('add_student')

        gender_val = request.POST.get('gender', '').strip().lower()
        year_val   = request.POST.get('year', '').strip()
        sem_val    = request.POST.get('semester', '').strip()

        dob_val = None
        raw_dob = request.POST.get('date_of_birth', '').strip()
        if raw_dob:
            try: dob_val = _dt.date.fromisoformat(raw_dob)
            except ValueError: pass

        rep_val = None
        raw_rep = request.POST.get('reporting_date', '').strip()
        if raw_rep:
            try: rep_val = _dt.date.fromisoformat(raw_rep)
            except ValueError: pass

        defaults = {
            'name':              name,
            'phone':             request.POST.get('phone', '').strip(),
            'email':             request.POST.get('email', '').strip(),
            'department':        request.POST.get('department', '').strip(),
            'year':              int(year_val) if year_val.isdigit() else None,
            'semester':          int(sem_val)  if sem_val.isdigit()  else None,
            'address':           request.POST.get('address', '').strip(),
            'state':             request.POST.get('state', '').strip().title(),
            'country':           request.POST.get('country', '').strip().title(),
            'type_of_entry':     request.POST.get('type_of_entry', '').strip(),
            'guardian_name':     request.POST.get('guardian_name', '').strip(),
            'guardian_phone':    request.POST.get('guardian_phone', '').strip(),
            'guardian_relation': request.POST.get('guardian_relation', '').strip(),
            'guardian_email':    request.POST.get('guardian_email', '').strip(),
            'college_id_number': request.POST.get('college_id_number', '').strip(),
        }
        if gender_val in ('male', 'female', 'other'):
            defaults['gender'] = gender_val
        if dob_val:
            defaults['date_of_birth'] = dob_val
        if rep_val:
            defaults['reporting_date'] = rep_val

        defaults = {k: v for k, v in defaults.items() if v not in (None, '')}

        student, created = Student.objects.update_or_create(
            roll_number=roll, defaults=defaults
        )
        if created:
            messages.success(request, f'Student "{name}" ({roll}) added successfully.')
        else:
            messages.info(request, f'Student "{name}" ({roll}) already exists — profile updated.')
        return redirect('student_detail', pk=student.pk)

    all_departments = list(Department.objects.filter(is_active=True).values_list('name', flat=True).order_by('name'))
    return render(request, 'admin/add_student.html', {
        'all_departments': all_departments,
    })


@admin_only
def import_students(request):  # noqa: C901
    """Bulk-import students from the styled Excel template."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        result = _process_student_excel(request.FILES['excel_file'], request.user)
        return render(request, 'admin/import_students.html', {'done': True, **result})
    return render(request, 'admin/import_students.html', {'done': False})


def _process_student_excel(excel_file, uploaded_by):  # noqa: C901
    """Parse the student bulk-upload Excel and create/update students + allocations."""
    import datetime as _dt
    from apps.accounts.models import User as _User

    created_students = updated_students = allocated_rooms = created_users = 0
    skipped_rows = []
    parse_error  = None

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb['Students'] if 'Students' in wb.sheetnames else wb.active
    except Exception as e:
        return {
            'parse_error': str(e), 'created_students': 0, 'updated_students': 0,
            'allocated_rooms': 0, 'created_users': 0, 'skipped_rows': [],
        }

    # Detect whether this is our styled template (title at row 1, headers at row 3)
    # or a plain file (headers at row 1).
    first_val = ws.cell(row=1, column=1).value
    if first_val and 'GGI Hostel ERP' in str(first_val):
        header_row_num = 3
        data_start_row = 4
    else:
        header_row_num = 1
        data_start_row = 2

    raw_headers = [ws.cell(row=header_row_num, column=c).value for c in range(1, 26)]

    # Build column-key → 0-based index map from header keywords
    col = {}
    for i, h in enumerate(raw_headers):
        if not h:
            continue
        h = str(h).strip().lower()
        if 'roll' in h:
            col['roll'] = i
        elif 'name' in h and 'guardian' not in h:
            col['name'] = i
        elif ('unique' in h or 'college id' in h or 'student id' in h) and 'room' not in h:
            col['unique_id'] = i
        elif 'email' in h and 'guardian' not in h:
            col['email'] = i
        elif 'dept' in h or 'branch' in h:
            col['dept'] = i
        elif 'year' in h and 'guardian' not in h:
            col['year'] = i
        elif 'sem' in h:
            col['sem'] = i
        elif 'gender' in h:
            col['gender'] = i
        elif ('phone' in h or 'mobile' in h) and 'guardian' not in h and 'parent' not in h:
            col['phone'] = i
        elif 'dob' in h or 'birth' in h:
            col['dob'] = i
        elif 'guardian' in h and 'name' in h:
            col['guardian_name'] = i
        elif 'guardian' in h and 'relat' in h:
            col['guardian_relation'] = i
        elif 'guardian' in h and 'phone' in h:
            col['guardian_phone'] = i
        elif 'guardian' in h and 'email' in h:
            col['guardian_email'] = i
        elif 'address' in h:
            col['address'] = i
        elif 'hostel' in h:
            col['hostel'] = i
        elif 'room' in h:
            col['room'] = i
        elif 'reporting' in h and 'date' in h:
            col['reporting_date'] = i
        elif 'type' in h and 'entry' in h:
            col['type_of_entry'] = i
        elif h == 'state':
            col['state'] = i
        elif 'country' in h:
            col['country'] = i
        elif 'parent' in h and 'phone' in h:
            col['parent_phone'] = i

    missing = [k for k in ('roll', 'name') if k not in col]
    if missing:
        return {
            'parse_error': f'Required columns not found: {missing}. Use the sample Excel template.',
            'created_students': 0, 'updated_students': 0,
            'allocated_rooms': 0, 'created_users': 0, 'skipped_rows': [],
        }

    for row_num, row in enumerate(
        ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        vals = list(row) + [None] * 20

        if not any(v is not None for v in vals[:25]):
            continue  # blank / separator row

        def _cell(key):
            if key not in col:
                return ''
            v = vals[col[key]]
            return str(v).strip() if v is not None else ''

        roll = _cell('roll')
        name = _cell('name')

        # Skip header-like or empty rows
        if not roll or not name:
            continue
        if roll.lower() in ('roll number', 'roll no', 'roll no*', 'sample'):
            continue

        try:
            # ── DOB parsing ──
            dob_val = None
            if 'dob' in col and vals[col['dob']]:
                raw = vals[col['dob']]
                if isinstance(raw, (_dt.date, _dt.datetime)):
                    dob_val = raw.date() if isinstance(raw, _dt.datetime) else raw
                else:
                    try:
                        dob_val = _dt.date.fromisoformat(str(raw).strip())
                    except ValueError:
                        pass

            # ── Year / Semester sanitise ──
            year_raw = _cell('year')
            try:
                year_val = str(int(float(year_raw))) if year_raw else '1'
            except ValueError:
                year_val = '1'
            if year_val not in ('1', '2', '3', '4'):
                year_val = '1'

            sem_raw = _cell('sem')
            try:
                sem_val = max(1, min(8, int(float(sem_raw)))) if sem_raw else 1
            except ValueError:
                sem_val = 1

            gender_val = _cell('gender').lower()
            if gender_val not in ('male', 'female', 'other'):
                gender_val = ''

            email_val = _cell('email')

            # ── Reporting date parsing ──
            reporting_date_val = None
            if 'reporting_date' in col and vals[col['reporting_date']]:
                raw_rd = vals[col['reporting_date']]
                if isinstance(raw_rd, (_dt.date, _dt.datetime)):
                    reporting_date_val = raw_rd.date() if isinstance(raw_rd, _dt.datetime) else raw_rd
                else:
                    try:
                        reporting_date_val = _dt.date.fromisoformat(str(raw_rd).strip())
                    except ValueError:
                        pass

            type_of_entry_val = _cell('type_of_entry').strip()

            state_val   = _cell('state').strip().title()
            country_val = _cell('country').strip().title()

            student_defaults = {
                'name':              name,
                'phone':             _cell('phone'),
                'department':        _cell('dept'),
                'year':              year_val,
                'semester':          sem_val,
                'address':           _cell('address'),
                'guardian_name':     _cell('guardian_name'),
                'guardian_phone':    _cell('guardian_phone'),
                'guardian_relation': _cell('guardian_relation'),
                'guardian_email':    _cell('guardian_email'),
                'college_id_number': _cell('unique_id'),
            }
            if email_val:
                student_defaults['email'] = email_val
            if gender_val:
                student_defaults['gender'] = gender_val
            if dob_val:
                student_defaults['date_of_birth'] = dob_val
            if state_val:
                student_defaults['state'] = state_val
            if country_val:
                student_defaults['country'] = country_val
            if type_of_entry_val:
                student_defaults['type_of_entry'] = type_of_entry_val
            if reporting_date_val:
                student_defaults['reporting_date'] = reporting_date_val

            student, was_created = Student.objects.update_or_create(
                roll_number=roll, defaults=student_defaults,
            )
            if was_created:
                created_students += 1
            else:
                updated_students += 1

            # ── Create / link User account ──
            if email_val:
                try:
                    user_obj = _User.objects.filter(email=email_val).first()
                    if not user_obj:
                        user_obj = _User(
                            email=email_val,
                            name=name,
                            role=_User.Role.STUDENT,
                            phone=_cell('phone'),
                        )
                        user_obj.set_password(roll)   # temp password = roll number
                        user_obj.save()
                        created_users += 1
                    if student.user_id != user_obj.pk:
                        student.user = user_obj
                        student.save(update_fields=['user'])
                except Exception as ue:
                    skipped_rows.append({
                        'row': row_num, 'data': f'{roll} — user account',
                        'reason': f'Could not create login: {ue}',
                    })

            # ── Room allocation ──
            hostel_name = _cell('hostel')
            room_no     = _cell('room')
            if hostel_name and room_no:
                try:
                    hostel_obj = Hostel.objects.get(name__iexact=hostel_name)
                    room_obj   = Room.objects.get(
                        hostel=hostel_obj, room_number__iexact=room_no
                    )
                    if student.allocations.filter(status='active').exists():
                        skipped_rows.append({
                            'row': row_num,
                            'data': f'{roll} → {hostel_name} / {room_no}',
                            'reason': 'Student already has an active room allocation',
                        })
                    elif room_obj.beds_available <= 0:
                        skipped_rows.append({
                            'row': row_num,
                            'data': f'{roll} → {hostel_name} / {room_no}',
                            'reason': f'Room {room_no} is full (capacity {room_obj.capacity})',
                        })
                    else:
                        avail_bed = room_obj.beds.filter(status='available').first()
                        from django.utils import timezone as _tz
                        RoomAllocation.objects.create(
                            student=student, room=room_obj, bed=avail_bed,
                            allocated_by=uploaded_by,
                            check_in=_tz.now().date(), status='active',
                        )
                        if avail_bed:
                            avail_bed.status = 'occupied'
                            avail_bed.save(update_fields=['status'])
                        room_obj.refresh_from_db()
                        if room_obj.current_occupants >= room_obj.capacity:
                            room_obj.status = 'occupied'
                            room_obj.save(update_fields=['status'])
                        student.is_resident = True
                        student.save(update_fields=['is_resident'])
                        allocated_rooms += 1
                except Hostel.DoesNotExist:
                    skipped_rows.append({
                        'row': row_num, 'data': f'{roll} → {hostel_name}',
                        'reason': f'Hostel "{hostel_name}" not found — check spelling',
                    })
                except Room.DoesNotExist:
                    skipped_rows.append({
                        'row': row_num, 'data': f'{roll} → {hostel_name} / {room_no}',
                        'reason': f'Room "{room_no}" not found in {hostel_name}',
                    })
                except Exception as ae:
                    skipped_rows.append({
                        'row': row_num, 'data': f'{roll} → {hostel_name} / {room_no}',
                        'reason': f'Allocation failed: {ae}',
                    })

        except Exception as e:
            skipped_rows.append({
                'row': row_num, 'data': f'Row {row_num}',
                'reason': str(e),
            })

    return {
        'parse_error':     parse_error,
        'created_students': created_students,
        'updated_students': updated_students,
        'allocated_rooms':  allocated_rooms,
        'created_users':    created_users,
        'skipped_rows':     skipped_rows,
    }


@admin_only
def student_import_sample(request):  # noqa: C901
    """Download a styled 17-column sample Excel for student bulk import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

    def hfill(h): return PatternFill('solid', fgColor=h)
    cc = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lc = Alignment(horizontal='left',   vertical='center', indent=1)

    # A-W  (23 columns)
    # A: Roll No   B: Name    C: Unique ID   D: Email
    # E: Dept      F: Year    G: Semester    H: Gender
    # I: Phone     J: DOB
    # K: Guardian Name  L: Guardian Relation  M: Guardian Phone  N: Guardian Email
    # O: Address   P: Hostel Name   Q: Room No
    # R: Reporting Date  S: Type of Entry  T: State  U: Country
    # V: Parent Phone  W: Fee Status
    for i, w in enumerate(
        [14, 22, 14, 28, 22, 7, 9, 9, 14, 14, 20, 18, 16, 26, 30, 18, 10, 14, 14, 14, 14, 14, 14], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1 — title banner
    tc = ws.cell(row=1, column=1,
                 value='  GGI Hostel ERP  —  Student Bulk Import Template')
    tc.font      = Font(bold=True, color='FFFFFF', size=13)
    tc.fill      = hfill('0D47A1')
    tc.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:W1')
    ws.row_dimensions[1].height = 26

    # Row 2 — section group labels
    for sc, ec, lbl, clr in [
        (1,  4,  'STUDENT IDENTITY',   '1565C0'),
        (5,  8,  'ACADEMIC DETAILS',   '1B5E20'),
        (9,  10, 'CONTACT DETAILS',    '00695C'),
        (11, 14, 'GUARDIAN INFO',      '6A1B9A'),
        (15, 15, 'ADDRESS',            'BF360C'),
        (16, 17, 'ADMIN ASSIGNED',     '455A64'),
        (18, 21, 'EXTRA DETAILS',      '00695C'),
        (22, 23, 'REFERENCE',          '78350F'),
    ]:
        c = ws.cell(row=2, column=sc, value=lbl)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = hfill(clr); c.alignment = cc; c.border = thin
        if sc != ec:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
    ws.row_dimensions[2].height = 18

    # Row 3 — column headers
    hdrs = [
        ('Roll Number*',            '1565C0'),
        ('Student Name*',           '1565C0'),
        ('Unique ID\n(College ID)', '1565C0'),
        ('Email',                   '1565C0'),
        ('Department\n(Branch)',    '1B5E20'),
        ('Year*\n(1–4)',            '1B5E20'),
        ('Semester*\n(1–8)',        '1B5E20'),
        ('Gender',                  '1B5E20'),
        ('Phone',                   '00695C'),
        ('Date of Birth\n(DD-MM-YYYY)', '00695C'),
        ('Guardian Name',           '6A1B9A'),
        ('Guardian\nRelation',      '6A1B9A'),
        ('Guardian\nPhone',         '6A1B9A'),
        ('Guardian\nEmail',         '6A1B9A'),
        ('Address',                 'BF360C'),
        ('Hostel Name\n(Admin fills)', '455A64'),
        ('Room No.\n(Admin fills)',    '455A64'),
        ('Reporting Date\n(DD-MM-YYYY)', '00695C'),
        ('Type of Entry\n(InCampus/OutCampus)', '00695C'),
        ('State',                   '00695C'),
        ('Country',                 '00695C'),
        ('Parent Phone',            '78350F'),
        ('Fee Status\n(Reference only)', '78350F'),
    ]
    for col_idx, (hdr, clr) in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font      = Font(bold=True, color='FFFFFF', size=9)
        cell.fill      = hfill(clr)
        cell.alignment = cc
        cell.border    = thin
    ws.row_dimensions[3].height = 32

    # Rows 4-5 — sample data
    samples = [
        ('2024001', 'Rahul Sharma',   'GGI2024001', 'rahul.sharma@ggi.edu',
         'Computer Science & Engineering', '2', '3', 'Male',
         '9876543210', '15-03-2005',
         'Suresh Sharma', 'Father', '9876543211', 'suresh.sharma@gmail.com',
         'Village Rampur, Tehsil ABC, District XYZ - 110001', '', '',
         '01-08-2024', 'InCampus', 'Haryana', 'India', '9876543211', 'Paid'),
        ('2024002', 'Priya Singh',    'GGI2024002', 'priya.singh@ggi.edu',
         'Electrical & Electronics Engg.', '2', '3', 'Female',
         '9876543212', '20-07-2005',
         'Rajesh Singh', 'Father', '9876543213', '',
         '12 MG Road, Near Bus Stand, City - 110002', '', '',
         '01-08-2024', 'OutCampus', 'Punjab', 'India', '9876543213', 'Pending'),
    ]
    for r_idx, row_data in enumerate(samples, 4):
        row_fill = hfill('EBF3FF') if r_idx % 2 == 0 else hfill('F8FAFF')
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = row_fill
            cell.border    = thin
            cell.alignment = lc if c_idx in (2, 5, 15) else cc
        ws.row_dimensions[r_idx].height = 18

    # Row 6 — separator
    for c in range(1, 24):
        cell = ws.cell(row=6, column=c, value='')
        cell.fill = hfill('F1F5F9'); cell.border = thin

    # Rows 7-17 — notes
    for row_num, text, bold, clr, bg in [
        (7,  'NOTES — Please read before filling:',
              True,  '1565C0', 'EBF3FF'),
        (8,  '1. Roll Number and Student Name are REQUIRED. All other columns are optional but recommended.',
              False, '374151', 'FFFBEB'),
        (9,  '2. Leave Hostel Name (P) and Room No. (Q) blank — Admin assigns them. Filling both triggers auto-allocation.',
              False, '374151', 'FFFBEB'),
        (10, '3. Once Admin assigns a room, the allocation appears automatically in the student\'s profile.',
              False, '374151', 'FFFBEB'),
        (11, '4. If a student with the same Roll Number already exists their record is UPDATED — not duplicated.',
              False, '374151', 'FFFBEB'),
        (12, '5. If Email is provided, a login account is auto-created with temporary password = Roll Number.',
              False, '374151', 'FFFBEB'),
        (13, '6. Date of Birth / Reporting Date format: DD-MM-YYYY  e.g. 15-03-2005',
              False, '374151', 'FFFBEB'),
        (14, '7. Gender: Male / Female / Other  (case-insensitive)',
              False, '374151', 'FFFBEB'),
        (15, '8. Guardian Relation: Father / Mother / Guardian / Relative / Other',
              False, '374151', 'FFFBEB'),
        (16, '9. Type of Entry: InCampus / OutCampus (case-insensitive)',
              False, '374151', 'FFFBEB'),
        (17, '10. Fee Status column (W) is for reference only — it is NOT imported into the system.',
              False, '374151', 'FFFBEB'),
    ]:
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font      = Font(bold=bold, size=10, color=clr)
        cell.fill      = hfill(bg)
        cell.alignment = lc
        cell.border    = thin
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=23)
        ws.row_dimensions[row_num].height = 16

    ws.freeze_panes = 'A4'

    # ── Instructions sheet ──
    wi = wb.create_sheet('Instructions')
    wi.column_dimensions['A'].width = 24
    wi.column_dimensions['B'].width = 62

    def ic(r, c, v, bold=False, clr='1F2937', bg=None, sz=10):
        cell = wi.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, color=clr, size=sz)
        cell.alignment = Alignment(vertical='center', indent=1, wrap_text=True)
        cell.border    = thin
        if bg:
            cell.fill = hfill(bg)
        wi.row_dimensions[r].height = 18
        return cell

    ic(1, 1, 'GGI Hostel ERP — Student Import: Column Guide',
       bold=True, clr='FFFFFF', bg='0D47A1', sz=12)
    wi.merge_cells('A1:B1')
    wi.row_dimensions[1].height = 24

    ic(2, 1, 'Column',      bold=True, clr='FFFFFF', bg='1565C0')
    ic(2, 2, 'Description', bold=True, clr='FFFFFF', bg='1565C0')

    col_info = [
        ('A  Roll Number*',
         'REQUIRED. Unique roll/admission number. Used as the student key — duplicates are updated.'),
        ('B  Student Name*',
         'REQUIRED. Full legal name of the student.'),
        ('C  Unique ID (College ID)',
         'College-issued ID number (e.g. GGI2024001). Used for biometric or ID card linking.'),
        ('D  Email',
         'Official email address. A login account is auto-created with this email; temporary password = Roll Number.'),
        ('E  Department (Branch)',
         'Engineering branch or discipline. e.g. Computer Science & Engineering, Civil Engg.'),
        ('F  Year (1–4)',
         'Current academic year: 1, 2, 3, or 4. Defaults to 1 if left blank.'),
        ('G  Semester (1–8)',
         'Current semester 1 through 8. Defaults to 1 if left blank.'),
        ('H  Gender',
         'Male / Female / Other (case-insensitive). Leave blank if not applicable.'),
        ('I  Phone',
         "Student's personal mobile number."),
        ('J  Date of Birth',
         'Format: DD-MM-YYYY  e.g. 15-03-2005. Used for identity verification.'),
        ('K  Guardian Name',
         'Full name of parent / guardian. Used in gate pass and leave approval workflows.'),
        ('L  Guardian Relation',
         'Relationship to student: Father / Mother / Guardian / Relative / Other.'),
        ('M  Guardian Phone',
         'RECOMMENDED. Used for parent contact on gate pass / leave approval.'),
        ('N  Guardian Email',
         "Guardian's email for notifications (optional)."),
        ('O  Address',
         'Permanent home address. Include village/city, district, state, and PIN code.'),
        ('P  Hostel Name',
         'LEAVE BLANK. Admin/SuperAdmin assigns this after upload. Filling it triggers auto-allocation.'),
        ('Q  Room No.',
         'LEAVE BLANK. Admin/SuperAdmin assigns this after upload. Must be filled together with Hostel Name.'),
        ('R  Reporting Date',
         'Date the student reported to hostel. Format: DD-MM-YYYY.'),
        ('S  Type of Entry',
         'InCampus (student lives in hostel campus) or OutCampus (student lives outside campus). Case-insensitive.'),
        ('T  State',
         "Student's home state. e.g. Haryana, Punjab, Uttar Pradesh."),
        ('U  Country',
         "Student's home country. Defaults to India if left blank."),
        ('V  Parent Phone',
         'Alternative parent/guardian contact number for reference.'),
        ('W  Fee Status',
         'For reference only — NOT imported. Use this column to note fee payment status in your records.'),
    ]
    for i, (col_name, desc) in enumerate(col_info, 3):
        bg = 'EBF3FF' if i % 2 == 0 else 'F8FAFF'
        ic(i, 1, col_name, bold=False, bg=bg)
        ic(i, 2, desc,     bold=False, bg=bg)
        wi.row_dimensions[i].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="student_import_sample.xlsx"'
    return resp


@superadmin_only
def toggle_student_status(request, pk):
    """Superadmin: mark a student active or inactive."""
    from django.utils import timezone as _tz
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        if student.is_active:
            # Mark inactive
            reason   = request.POST.get('inactive_reason', '').strip()
            remarks  = request.POST.get('inactive_remarks', '').strip()
            exit_date_str = request.POST.get('exit_date', '').strip()
            exit_time_str = request.POST.get('exit_time', '').strip()
            import datetime as _dt2
            student.is_active          = False
            student.inactive_reason    = reason
            student.inactive_remarks   = remarks
            student.marked_inactive_by = request.user
            student.marked_inactive_at = _tz.now()
            if exit_date_str:
                try:
                    student.exit_date = _dt2.date.fromisoformat(exit_date_str)
                except ValueError:
                    pass
            if exit_time_str:
                try:
                    student.exit_time = _dt2.time.fromisoformat(exit_time_str)
                except ValueError:
                    pass
            student.save(update_fields=[
                'is_active', 'inactive_reason', 'inactive_remarks',
                'exit_date', 'exit_time', 'marked_inactive_by', 'marked_inactive_at'
            ])
            messages.success(request, f'{student.name} marked as Inactive.')
        else:
            # Re-activate
            student.is_active        = True
            student.inactive_reason  = ''
            student.inactive_remarks = ''
            student.exit_date        = None
            student.exit_time        = None
            student.save(update_fields=[
                'is_active', 'inactive_reason', 'inactive_remarks',
                'exit_date', 'exit_time'
            ])
            messages.success(request, f'{student.name} re-activated.')
        return redirect('student_detail', pk=pk)
    # GET — render modal form (JSON response for AJAX)
    return render(request, 'admin/toggle_student_status_modal.html', {
        'student': student,
    })


@admin_required
def student_detail(request, pk):
    student = get_object_or_404(Student, pk=pk)
    allocations  = RoomAllocation.objects.filter(student=student).select_related('room__hostel').order_by('-created_at')
    fee_records  = FeeRecord.objects.filter(student=student).select_related('fee_structure').order_by('-created_at')
    complaints   = Complaint.objects.filter(student=student).order_by('-created_at')
    gate_passes  = GatePass.objects.filter(student=student).select_related(
        'approved_by', 'exit_allowed_by', 'entry_allowed_by'
    ).order_by('-created_at')
    leave_apps   = LeaveApplication.objects.filter(student=student).select_related('reviewed_by').order_by('-from_date')
    discipline   = DisciplineRecord.objects.filter(student=student).select_related('recorded_by').order_by('-created_at')

    return render(request, 'admin/student_detail.html', {
        'student':     student,
        'allocations': allocations,
        'fee_records': fee_records,
        'complaints':  complaints,
        'gate_passes': gate_passes,
        'leave_apps':  leave_apps,
        'discipline':  discipline,
        'flag_choices': Student.FlagColor.choices,
    })


@admin_required
@require_POST
def student_set_flag(request, pk):
    student = get_object_or_404(Student, pk=pk)
    student.flag_color  = request.POST.get('flag_color', 'green')
    student.flag_note   = request.POST.get('flag_note', '').strip()
    student.flag_set_by = request.user
    student.flag_set_at = timezone.now()
    student.save(update_fields=['flag_color', 'flag_note', 'flag_set_by', 'flag_set_at'])
    messages.success(request, f'Flag updated to {student.flag_color}.')
    return redirect('student_detail', pk=pk)


@require_POST
def student_toggle_edit_permission(request, pk):
    if not (request.user.is_authenticated and request.user.role == User.Role.SUPER_ADMIN):
        messages.error(request, 'Only super admin can change this setting.')
        return redirect('student_detail', pk=pk)
    student = get_object_or_404(Student, pk=pk)
    student.can_edit_profile = not student.can_edit_profile
    student.save(update_fields=['can_edit_profile'])
    status = 'enabled' if student.can_edit_profile else 'disabled'
    messages.success(request, f'Profile editing {status} for {student.name}.')
    return redirect('student_detail', pk=pk)


@login_required
@require_POST
def student_bulk_edit_permission(request):
    """Superadmin: bulk enable/disable profile editing for all (or hostel-scoped) students."""
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Only super admin can do this.')
        return redirect('student_list')
    action   = request.POST.get('bulk_action', 'enable')
    hostel_pk = request.POST.get('hostel', '').strip()
    qs = Student.objects.all()
    if hostel_pk:
        qs = qs.filter(allocations__room__hostel_id=hostel_pk, allocations__status='active').distinct()
    new_val = (action == 'enable')
    count   = qs.update(can_edit_profile=new_val)
    verb    = 'enabled' if new_val else 'disabled'
    messages.success(request, f'Profile editing {verb} for {count} student(s).')
    return redirect('student_list')


@login_required
def my_profile(request):
    student, redir = _get_student_or_redirect(request)
    if redir: return redir
    allocations = RoomAllocation.objects.filter(student=student).select_related('room__hostel').order_by('-created_at')
    fee_records = FeeRecord.objects.filter(student=student).select_related('fee_structure').order_by('-created_at')
    complaints  = Complaint.objects.filter(student=student).order_by('-created_at')
    gate_passes = GatePass.objects.filter(student=student).select_related('approved_by').order_by('-created_at')
    leave_apps  = LeaveApplication.objects.filter(student=student).select_related('reviewed_by').order_by('-from_date')
    discipline  = DisciplineRecord.objects.filter(student=student).select_related('recorded_by').order_by('-created_at')
    return render(request, 'student/profile.html', {
        'student':    student,
        'allocations': allocations,
        'fee_records': fee_records,
        'complaints':  complaints,
        'gate_passes': gate_passes,
        'leave_apps':  leave_apps,
        'discipline':  discipline,
    })


@login_required
def student_edit_profile(request):
    student, redir = _get_student_or_redirect(request)
    if redir: return redir
    if not student.can_edit_profile:
        messages.error(request, 'Profile editing is not currently allowed. Contact admin.')
        return redirect('my_profile')
    if request.method == 'POST':
        student.phone             = request.POST.get('phone', '').strip()
        student.email             = request.POST.get('email', '').strip()
        student.address           = request.POST.get('address', '').strip()
        student.guardian_name     = request.POST.get('guardian_name', '').strip()
        student.guardian_phone    = request.POST.get('guardian_phone', '').strip()
        student.guardian_relation = request.POST.get('guardian_relation', '').strip()
        student.guardian_email    = request.POST.get('guardian_email', '').strip()
        student.aadhar_number     = request.POST.get('aadhar_number', '').strip()
        student.college_id_number = request.POST.get('college_id_number', '').strip()
        if request.FILES.get('photo'):
            student.photo = request.FILES['photo']
        if request.FILES.get('aadhar_doc'):
            student.aadhar_doc = request.FILES['aadhar_doc']
        if request.FILES.get('college_id_doc'):
            student.college_id_doc = request.FILES['college_id_doc']
        student.save()
        messages.success(request, 'Profile updated successfully.')
        return redirect('my_profile')
    return render(request, 'student/edit_profile.html', {'student': student})


# ─── room management ─────────────────────────────────────────────────────────

@admin_required
def room_list(request):
    warden_hostel = get_hostel_scope(request.user)
    hostel_qs = Hostel.objects.all()
    if warden_hostel:
        hostel_qs = hostel_qs.filter(pk=warden_hostel.pk)

    hostels_data = []
    for hostel in hostel_qs.order_by('name'):
        rooms = list(
            Room.objects.filter(hostel=hostel)
            .annotate(active_count=Count('allocations', filter=Q(allocations__status='active')))
            .order_by('floor', 'room_number')
        )
        total       = len(rooms)
        occupied    = sum(1 for r in rooms if r.status == 'occupied')
        vacant      = sum(1 for r in rooms if r.status == 'vacant')
        partial     = sum(1 for r in rooms if r.status == 'partial')
        maintenance = sum(1 for r in rooms if r.status == 'maintenance')

        floors = {}
        for room in rooms:
            f = room.floor
            if f not in floors:
                floors[f] = {'floor': f, 'total': 0, 'occupied': 0, 'vacant': 0, 'partial': 0}
            floors[f]['total'] += 1
            if room.status == 'occupied':
                floors[f]['occupied'] += 1
            elif room.status == 'vacant':
                floors[f]['vacant'] += 1
            elif room.status == 'partial':
                floors[f]['partial'] += 1

        hostels_data.append({
            'hostel':      hostel,
            'total':       total,
            'occupied':    occupied,
            'vacant':      vacant,
            'partial':     partial,
            'maintenance': maintenance,
            'floors':      sorted(floors.values(), key=lambda x: x['floor']),
        })

    return render(request, 'admin/rooms.html', {
        'hostels_data': hostels_data,
        'warden_hostel': warden_hostel,
    })


@admin_only
def room_occupancy_drill(request):
    """Drill-down view for dashboard vacant/partial/occupied cards with filters."""
    warden_hostel  = get_hostel_scope(request.user)
    status_filter  = request.GET.get('status', '')   # vacant | occupied | partial
    hostel_filter  = request.GET.get('hostel', '')
    floor_filter   = request.GET.get('floor', '')
    type_filter    = request.GET.get('room_type', '').strip()
    desig_filter   = request.GET.get('designation', '').strip()

    room_qs = Room.objects.select_related('hostel').annotate(
        active_count=Count('allocations', filter=Q(allocations__status='active'))
    ).order_by('hostel__name', 'floor', 'room_number')

    if warden_hostel:
        room_qs = room_qs.filter(hostel=warden_hostel)
    elif hostel_filter:
        room_qs = room_qs.filter(hostel_id=hostel_filter)

    if status_filter:
        room_qs = room_qs.filter(status=status_filter)
    if floor_filter:
        room_qs = room_qs.filter(floor=floor_filter)
    if type_filter:
        room_qs = room_qs.filter(room_type__icontains=type_filter)
    if desig_filter:
        room_qs = room_qs.filter(designation=desig_filter)

    rooms = list(room_qs)

    # Attach active residents to each room
    from django.db.models import Prefetch as _Prefetch
    alloc_map = {}
    for alloc in RoomAllocation.objects.filter(
        room__in=[r.pk for r in rooms], status='active'
    ).select_related('student'):
        alloc_map.setdefault(alloc.room_id, []).append(alloc.student)
    for room in rooms:
        room.residents = alloc_map.get(room.pk, [])

    # Build filter option lists
    all_hostels  = Hostel.objects.order_by('name') if not warden_hostel else []
    all_floors   = Room.objects.filter(
        **(({'hostel': warden_hostel} if warden_hostel else {}))
    ).values_list('floor', flat=True).distinct().order_by('floor')
    all_types    = Room.objects.values_list('room_type', flat=True).exclude(
        room_type=''
    ).distinct().order_by('room_type')

    # Status counts (scoped to hostel if warden)
    base = Room.objects.all()
    if warden_hostel:
        base = base.filter(hostel=warden_hostel)
    counts = {
        'all':      base.count(),
        'vacant':   base.filter(status='vacant').count(),
        'occupied': base.filter(status='occupied').count(),
        'partial':  base.filter(status='partial').count(),
    }

    paginator = Paginator(rooms, 30)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin/room_drill.html', {
        'page':           page,
        'status_filter':  status_filter,
        'hostel_filter':  hostel_filter,
        'floor_filter':   floor_filter,
        'type_filter':    type_filter,
        'desig_filter':   desig_filter,
        'all_hostels':    all_hostels,
        'all_floors':     all_floors,
        'all_types':      all_types,
        'counts':         counts,
        'warden_hostel':  warden_hostel,
    })


@admin_only
def allocate_room(request, student_pk=None):
    student = get_object_or_404(Student, pk=student_pk) if student_pk else None

    # Guard: already allocated
    if student and RoomAllocation.objects.filter(student=student, status='active').exists():
        messages.error(request, f'{student.name} already has an active room allocation.')
        return redirect('student_detail', pk=student_pk)

    if request.method == 'POST':
        form = AllocateRoomForm(request.POST)
        # When student is pre-set via URL, inject it into the queryset so validation passes
        if student:
            form.fields['student'].queryset = Student.objects.filter(pk=student.pk)

        if form.is_valid():
            alloc_student = form.cleaned_data['student']
            room          = form.cleaned_data['room']

            if RoomAllocation.objects.filter(student=alloc_student, status='active').exists():
                messages.error(request, f'{alloc_student.name} already has an active allocation.')
            elif room.current_occupants >= room.capacity:
                messages.error(request, f'Room {room.room_number} is at full capacity.')
            else:
                RoomAllocation.objects.create(
                    student=alloc_student, room=room,
                    allocated_by=request.user,
                    semester=Semester.current(),
                    check_in=form.cleaned_data['check_in'],
                    notes=form.cleaned_data.get('notes', ''),
                )
                if room.current_occupants >= room.capacity:
                    room.status = 'occupied'
                    room.save(update_fields=['status'])
                alloc_student.is_resident = True
                alloc_student.save(update_fields=['is_resident'])
                messages.success(request, f'Room {room.room_number} allocated to {alloc_student.name}.')
                return redirect('student_detail', pk=alloc_student.pk)
        # fall through to re-render form with errors
    else:
        initial = {'student': student} if student else {}
        form = AllocateRoomForm(initial=initial)

    return render(request, 'admin/allocate_room.html', {'form': form, 'student': student})


@admin_only
def checkout_student(request, allocation_pk):
    allocation = get_object_or_404(RoomAllocation, pk=allocation_pk, status='active')
    if request.method == 'POST':
        form = CheckoutForm(request.POST)
        if form.is_valid():
            allocation.status    = 'checkout'
            allocation.check_out = form.cleaned_data['check_out']
            allocation.notes     += f'\nCheckout: {form.cleaned_data.get("notes","")}'
            allocation.save()
            # Update room status
            room = allocation.room
            if room.current_occupants == 0:
                room.status = 'vacant'
                room.save(update_fields=['status'])
            # Mark student as no longer a resident
            allocation.student.is_resident = False
            allocation.student.save(update_fields=['is_resident'])
            messages.success(request, f'{allocation.student.name} checked out successfully.')
            return redirect('student_detail', pk=allocation.student.pk)
    else:
        form = CheckoutForm(initial={'check_out': timezone.now().date()})
    return render(request, 'admin/checkout.html', {'form': form, 'allocation': allocation})


# ─── complaints ───────────────────────────────────────────────────────────────

@admin_required
def complaint_list(request):  # noqa: C901
    """Admin / Warden view of all complaints with category + status filters."""
    qs = Complaint.objects.select_related(
        'student', 'room__hostel', 'assigned_to'
    ).order_by('-created_at')

    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel:
        qs = qs.filter(room__hostel=warden_hostel)

    f_status   = request.GET.get('status', '')
    f_category = request.GET.get('category', '')
    f_priority = request.GET.get('priority', '')
    f_q        = request.GET.get('q', '').strip()
    if f_status:   qs = qs.filter(status=f_status)
    if f_category: qs = qs.filter(category=f_category)
    if f_priority: qs = qs.filter(priority=f_priority)
    if f_q:
        qs = qs.filter(
            Q(title__icontains=f_q) |
            Q(student__name__icontains=f_q) |
            Q(student__roll_number__icontains=f_q) |
            Q(raised_by_staff__name__icontains=f_q)
        )

    # Scoped base for counts (same hostel filter, no status/cat/priority filters)
    scoped_base = Complaint.objects.all()
    if warden_hostel:
        scoped_base = scoped_base.filter(room__hostel=warden_hostel)
    counts = {
        'all':         scoped_base.count(),
        'submitted':   scoped_base.filter(status='submitted').count(),
        'verified':    scoped_base.filter(status='verified').count(),
        'forwarded':   scoped_base.filter(status='forwarded').count(),
        'in_progress': scoped_base.filter(status='in_progress').count(),
        'resolved':    scoped_base.filter(status='resolved').count(),
        'closed':      scoped_base.filter(status='closed').count(),
        'on_hold':     scoped_base.filter(status='on_hold').count(),
    }

    maintenance_users = User.objects.filter(role=User.Role.MAINTENANCE, is_active=True)
    paginator = Paginator(qs, 20)
    page      = paginator.get_page(request.GET.get('page'))
    return render(request, 'admin/complaints.html', {
        'page_obj':          page,
        'counts':            counts,
        'maintenance_users': maintenance_users,
        'statuses':          Complaint.Status.choices,
        'categories':        Complaint.Category.choices,
        'priorities':        Complaint.Priority.choices,
        'filters': {
            'status': f_status, 'category': f_category,
            'priority': f_priority, 'q': f_q,
        },
    })


@login_required
def complaint_detail(request, pk):
    """Full complaint detail + timeline. Accessible by student (own), staff, admin."""
    complaint = get_object_or_404(
        Complaint.objects.select_related(
            'student', 'raised_by_staff', 'room__hostel', 'assigned_to'
        ),
        pk=pk
    )
    role = request.user.role
    is_student = (role == User.Role.STUDENT)
    if is_student:
        student, redir = _get_student_or_redirect(request)
        if redir: return redir
        if complaint.student != student:
            messages.error(request, 'Access denied.')
            return redirect('student_dashboard')

    timeline = complaint.timeline.select_related('updated_by').order_by('created_at')
    maintenance_users = User.objects.filter(role=User.Role.MAINTENANCE, is_active=True)

    return render(request, 'complaints/complaint_detail.html', {
        'complaint':         complaint,
        'timeline':          timeline,
        'maintenance_users': maintenance_users,
        'priorities':        Complaint.Priority.choices,
        'is_warden':         role == User.Role.WARDEN,
        'is_admin':          role in (User.Role.ADMIN, User.Role.SUPER_ADMIN),
        'is_maintenance':    role == User.Role.MAINTENANCE,
        'is_student':        is_student,
    })


@login_required
@require_POST
def complaint_action(request, pk):  # noqa: C901
    """Handles all role-specific complaint actions with timeline logging."""
    complaint = get_object_or_404(Complaint, pk=pk)
    role      = request.user.role
    action    = request.POST.get('action', '')
    comment   = request.POST.get('comment', '').strip()
    new_status = None

    _staff_roles = (User.Role.WARDEN, User.Role.ADMIN, User.Role.SUPER_ADMIN)

    # Warden / Admin / SuperAdmin: verify → set priority, add remarks
    if action == 'verify' and role in _staff_roles:
        complaint.status         = Complaint.Status.VERIFIED
        complaint.priority       = request.POST.get('priority', Complaint.Priority.MEDIUM)
        complaint.warden_remarks = comment
        new_status = Complaint.Status.VERIFIED

    # Warden / Admin / SuperAdmin: forward to maintenance incharge
    elif action == 'forward' and role in _staff_roles:
        mi_id = request.POST.get('maintenance_user')
        mi    = User.objects.filter(pk=mi_id, role=User.Role.MAINTENANCE).first()
        complaint.status       = Complaint.Status.FORWARDED
        complaint.assigned_to  = mi
        complaint.forwarded_at = timezone.now()
        complaint.warden_remarks = comment or complaint.warden_remarks
        new_status = Complaint.Status.FORWARDED
        if mi:
            send_notification(
                [mi],
                f'Complaint Assigned — {complaint.title}',
                f'A {complaint.get_category_display()} complaint has been assigned to you by {request.user.name}.',
                notif_type='complaint', link=f'/complaints/{complaint.pk}/',
            )

    # Maintenance: mark in_progress
    elif action == 'in_progress' and role == User.Role.MAINTENANCE:
        complaint.status = Complaint.Status.IN_PROGRESS
        complaint.maintenance_remarks = comment
        new_status = Complaint.Status.IN_PROGRESS

    # Maintenance: mark resolved
    elif action == 'resolve' and role == User.Role.MAINTENANCE:
        complaint.status              = Complaint.Status.RESOLVED
        complaint.maintenance_remarks = comment
        complaint.resolved_at         = timezone.now()
        new_status = Complaint.Status.RESOLVED
        targets = list(filter(None, [
            complaint.student.user if complaint.student_id else complaint.raised_by_staff,
            complaint.room.hostel.warden if (complaint.room and complaint.room.hostel) else None,
        ]))
        send_notification(
            targets,
            f'Complaint Resolved — {complaint.title}',
            f'Your complaint has been resolved. ' + (f'Note: {comment}' if comment else ''),
            notif_type='complaint', link=f'/complaints/{complaint.pk}/',
        )

    # Warden / Admin / SuperAdmin: close
    elif action == 'close' and role in (User.Role.WARDEN, User.Role.ADMIN, User.Role.SUPER_ADMIN):
        complaint.status    = Complaint.Status.CLOSED
        complaint.closed_at = timezone.now()
        new_status = Complaint.Status.CLOSED

    # Add comment only (any authorised role)
    elif action == 'comment' and role in (
        User.Role.WARDEN, User.Role.ADMIN, User.Role.SUPER_ADMIN,
        User.Role.MAINTENANCE,
    ):
        if comment:
            ComplaintTimeline.objects.create(
                complaint=complaint, status=complaint.status,
                comment=comment, updated_by=request.user,
            )
            messages.success(request, 'Comment added.')
        return redirect('complaint_detail', pk=pk)

    if new_status:
        complaint.save()
        ComplaintTimeline.objects.create(
            complaint=complaint, status=new_status,
            comment=comment, updated_by=request.user,
        )
        messages.success(request, f'Complaint updated to: {complaint.get_status_display()}.')
    else:
        messages.warning(request, 'No valid action taken.')

    # Redirect: maintenance goes to their dashboard, others to complaint detail
    if role == User.Role.MAINTENANCE:
        return redirect('maintenance_incharge_dashboard')
    return redirect('complaint_detail', pk=pk)


@login_required
def raise_complaint(request):
    """Student, Warden, or SuperAdmin raises a complaint."""
    role = request.user.role
    is_staff_reporter = role in (
        User.Role.SUPER_ADMIN, User.Role.ADMIN, User.Role.WARDEN
    )

    student = None
    if not is_staff_reporter:
        student, _redir = _get_student_or_redirect(request)
        if _redir: return _redir

    if request.method == 'POST':
        title_val    = request.POST.get('title', '').strip()
        category_val = request.POST.get('category', '')
        desc_val     = request.POST.get('description', '').strip()
        if not title_val or not category_val or not desc_val:
            messages.error(request, 'Please fill all required fields.')
            return redirect('raise_complaint')

        if is_staff_reporter:
            location_val = request.POST.get('location', '').strip()
            priority_val = request.POST.get('priority', '') or None
            forward_now  = request.POST.get('forward_now') == '1'
            maint_user_id = request.POST.get('maintenance_user', '').strip()

            status = Complaint.Status.VERIFIED if forward_now else Complaint.Status.SUBMITTED
            complaint = Complaint.objects.create(
                student=None,
                raised_by_staff=request.user,
                location=location_val,
                category=category_val,
                title=title_val,
                description=desc_val,
                priority=priority_val,
                status=status,
            )
            comment_text = (
                f'Complaint reported by {role.replace("_"," ").title()} — {request.user.name}.'
            )
            ComplaintTimeline.objects.create(
                complaint=complaint, status=status,
                comment=comment_text, updated_by=request.user,
            )
            if forward_now and maint_user_id:
                maint_user = User.objects.filter(
                    pk=maint_user_id, role=User.Role.MAINTENANCE
                ).first()
                if maint_user:
                    complaint.assigned_to  = maint_user
                    complaint.forwarded_at = timezone.now()
                    complaint.status       = Complaint.Status.FORWARDED
                    complaint.save(update_fields=['assigned_to', 'forwarded_at', 'status'])
                    ComplaintTimeline.objects.create(
                        complaint=complaint, status=Complaint.Status.FORWARDED,
                        comment=f'Directly forwarded to {maint_user.name}.',
                        updated_by=request.user,
                    )
            messages.success(request, 'Complaint submitted and visible to maintenance team.')
            return redirect('complaint_list')
        else:
            allocation = RoomAllocation.objects.filter(student=student, status='active').first()
            complaint = Complaint.objects.create(
                student=student,
                room=allocation.room if allocation else None,
                category=category_val,
                title=title_val,
                description=desc_val,
                status=Complaint.Status.SUBMITTED,
            )
            ComplaintTimeline.objects.create(
                complaint=complaint, status=Complaint.Status.SUBMITTED,
                comment='Complaint submitted by student.', updated_by=request.user,
            )
            send_notification(
                _staff_targets(student),
                f'New Complaint — {title_val}',
                f'{student.name} raised a {complaint.get_category_display()} complaint.',
                notif_type='complaint', link=f'/complaints/{complaint.pk}/',
            )
            messages.success(request, 'Complaint submitted. The warden will review it shortly.')
            return redirect('my_complaints')

    maintenance_users = User.objects.filter(role=User.Role.MAINTENANCE, is_active=True)
    return render(request, 'complaints/raise_complaint.html', {
        'categories':        Complaint.Category.choices,
        'priorities':        Complaint.Priority.choices,
        'is_staff_reporter': is_staff_reporter,
        'maintenance_users': maintenance_users,
    })


@login_required
def my_complaints(request):
    """Student: list their own complaints with timeline status."""
    student, redir = _get_student_or_redirect(request)
    if redir: return redir
    base_qs    = Complaint.objects.filter(student=student).prefetch_related('timeline')
    f_status   = request.GET.get('status', '')
    f_category = request.GET.get('category', '')
    qs = base_qs.order_by('-created_at')
    if f_status:   qs = qs.filter(status=f_status)
    if f_category: qs = qs.filter(category=f_category)
    counts = {
        'all':         base_qs.count(),
        'submitted':   base_qs.filter(status='submitted').count(),
        'verified':    base_qs.filter(status='verified').count(),
        'in_progress': base_qs.filter(status='in_progress').count(),
        'resolved':    base_qs.filter(status='resolved').count(),
        'closed':      base_qs.filter(status='closed').count(),
    }
    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'student/my_complaints.html', {
        'page_obj':  page_obj,
        'counts':    counts,
        'statuses':  Complaint.Status.choices,
        'categories': Complaint.Category.choices,
        'filters':   {'status': f_status, 'category': f_category},
    })


# ─── fees ────────────────────────────────────────────────────────────────────

@admin_only
def fee_list(request):
    qs = FeeRecord.objects.select_related(
        'student', 'fee_structure'
    ).order_by('-created_at')
    status = request.GET.get('status', '')
    search = request.GET.get('q', '').strip()
    if status: qs = qs.filter(status=status)
    if search: qs = qs.filter(
        Q(student__roll_number__icontains=search) |
        Q(student__name__icontains=search)
    )
    paginator = Paginator(qs, 25)
    page      = paginator.get_page(request.GET.get('page'))
    return render(request, 'admin/fees.html', {
        'page_obj': page,
        'filters': {'status': status, 'q': search},
        'summary': qs.aggregate(
            total=Sum('total_amount'),
            collected=Sum('amount_paid'),
        ),
    })


@admin_only
@require_POST
def update_fee(request, pk):
    record = get_object_or_404(FeeRecord, pk=pk)
    amount = request.POST.get('amount_paid')
    mode   = request.POST.get('payment_mode', '')
    ref    = request.POST.get('transaction_ref', '')
    if amount:
        record.amount_paid   = Decimal(amount)
        record.payment_mode  = mode
        record.transaction_ref = ref
        record.paid_on       = timezone.now().date()
        if record.amount_paid >= record.total_amount:
            record.status = 'paid'
        elif record.amount_paid > 0:
            record.status = 'partial'
        record.save()
        if record.student.user:
            send_notification(
                [record.student.user],
                'Fee Payment Recorded',
                f'₹{record.amount_paid} payment recorded. Status: {record.get_status_display()}. Balance: ₹{record.total_amount - record.amount_paid}.',
                notif_type='general',
            )
        messages.success(request, 'Fee record updated.')
    return redirect('fee_list')


@admin_only
def generate_fees(request):
    """Generate fee records for all active residents for a fee structure."""
    if request.method == 'POST':
        structure_id = request.POST.get('fee_structure')
        structure = get_object_or_404(FeeStructure, pk=structure_id)
        students  = Student.objects.filter(is_resident=True)
        created   = 0
        for student in students:
            _, was_created = FeeRecord.objects.get_or_create(
                student=student,
                fee_structure=structure,
                defaults={'total_amount': structure.total}
            )
            if was_created:
                created += 1
        messages.success(request, f'Generated {created} fee records.')
        return redirect('fee_list')

    structures = FeeStructure.objects.filter(is_active=True).order_by('-created_at')
    return render(request, 'admin/generate_fees.html', {'structures': structures})


# ─── attendance ───────────────────────────────────────────────────────────────

@admin_required
def attendance_list(request):
    """Warden → redirect to their hostel detail. Admin/SuperAdmin → hostel overview cards."""
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel:
        url = reverse('hostel_attendance_detail', kwargs={'hostel_pk': warden_hostel.pk})
        qs = request.GET.urlencode()
        return redirect(url + (f'?{qs}' if qs else ''))

    date_str = request.GET.get('date', '')
    try:
        selected_date = dt_date.fromisoformat(date_str) if date_str else dt_date.today()
    except ValueError:
        selected_date = dt_date.today()

    hostels = Hostel.objects.all().order_by('name')
    hostel_stats = []
    for h in hostels:
        stu_ids = list(
            Student.objects.filter(
                is_resident=True,
                allocations__room__hostel=h,
                allocations__status='active',
            ).distinct().values_list('pk', flat=True)
        )
        total   = len(stu_ids)
        att_qs  = Attendance.objects.filter(date=selected_date, student_id__in=stu_ids)
        present = att_qs.filter(status='present').count()
        absent  = att_qs.filter(status='absent').count()
        leave   = att_qs.filter(status='leave').count()
        marked  = att_qs.count()
        hostel_stats.append({
            'hostel':       h,
            'total':        total,
            'present':      present,
            'absent':       absent,
            'leave':        leave,
            'unmarked':     total - marked,
            'present_pct':  round(present / total * 100) if total else 0,
            'absent_pct':   round(absent  / total * 100) if total else 0,
            'leave_pct':    round(leave   / total * 100) if total else 0,
            'marked_pct':   round(marked  / total * 100) if total else 0,
        })

    all_ids   = list(Student.objects.filter(is_resident=True).values_list('pk', flat=True))
    total_all = len(all_ids)
    att_all   = Attendance.objects.filter(date=selected_date, student_id__in=all_ids)
    overall   = {
        'total':    total_all,
        'present':  att_all.filter(status='present').count(),
        'absent':   att_all.filter(status='absent').count(),
        'leave':    att_all.filter(status='leave').count(),
        'unmarked': total_all - att_all.count(),
    }

    return render(request, 'admin/attendance.html', {
        'selected_date': selected_date,
        'hostel_stats':  hostel_stats,
        'overall':       overall,
    })


@admin_required
def mark_attendance(request):
    """Mark or update attendance for a given date."""
    if request.method == 'POST':
        date_str = request.POST.get('date', '')
        try:
            selected_date = dt_date.fromisoformat(date_str)
        except ValueError:
            selected_date = dt_date.today()

        students = Student.objects.filter(is_resident=True)
        for student in students:
            status = request.POST.get(f'status_{student.pk}')
            if status in dict(Attendance.Status.choices):
                Attendance.objects.update_or_create(
                    student=student,
                    date=selected_date,
                    defaults={'status': status, 'marked_by': request.user},
                )

        messages.success(request, f'Attendance saved for {selected_date}.')
        return HttpResponseRedirect(reverse('attendance_list') + f'?date={selected_date}')

    date_str = request.GET.get('date', '')
    try:
        selected_date = dt_date.fromisoformat(date_str) if date_str else dt_date.today()
    except ValueError:
        selected_date = dt_date.today()

    students    = Student.objects.filter(is_resident=True).order_by('roll_number')
    records_map = {r.student_id: r for r in Attendance.objects.filter(date=selected_date)}
    student_data = [{'student': s, 'record': records_map.get(s.id)} for s in students]

    return render(request, 'admin/mark_attendance.html', {
        'selected_date': selected_date,
        'student_data':  student_data,
        'statuses':      Attendance.Status.choices,
    })


@admin_required
def attendance_report(request):
    """Per-student attendance percentage summary."""
    students = Student.objects.filter(is_resident=True).annotate(
        total_days=Count('attendance_records'),
        present_days=Count(Case(When(attendance_records__status='present', then=1), output_field=IntegerField())),
        absent_days=Count(Case(When(attendance_records__status='absent',  then=1), output_field=IntegerField())),
        leave_days=Count(Case(When(attendance_records__status='leave',   then=1), output_field=IntegerField())),
    ).order_by('roll_number')

    return render(request, 'admin/attendance_report.html', {'students': students})


@admin_required
def hostel_attendance_detail(request, hostel_pk):
    """Floor/room/student card attendance view for a single hostel."""
    hostel = get_object_or_404(Hostel, pk=hostel_pk)

    # Wardens may only view their own hostel
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel and warden_hostel.pk != hostel.pk:
        messages.error(request, 'Access denied.')
        return redirect('attendance_list')

    date_str = request.GET.get('date', '')
    try:
        selected_date = dt_date.fromisoformat(date_str) if date_str else dt_date.today()
    except ValueError:
        selected_date = dt_date.today()

    filter_year = request.GET.get('year', '')
    filter_dept = request.GET.get('dept', '')
    filter_sem  = request.GET.get('sem',  '')

    student_qs = Student.objects.filter(
        is_resident=True,
        allocations__room__hostel=hostel,
        allocations__status='active',
    ).distinct()

    if filter_year:
        student_qs = student_qs.filter(year=filter_year)
    if filter_dept:
        student_qs = student_qs.filter(department=filter_dept)
    if filter_sem:
        student_qs = student_qs.filter(semester=filter_sem)

    attend_map = {
        a.student_id: a
        for a in Attendance.objects.filter(date=selected_date, student__in=student_qs)
    }

    # Group: floor → room → [student records]
    floors = OrderedDict()
    allocs = (
        RoomAllocation.objects
        .filter(status='active', room__hostel=hostel, student__in=student_qs)
        .select_related('student', 'room')
        .order_by('room__floor', 'room__room_number', 'student__roll_number')
    )
    for alloc in allocs:
        floor_num = alloc.room.floor
        room_key  = str(alloc.room.pk)
        floors.setdefault(floor_num, OrderedDict())
        if room_key not in floors[floor_num]:
            floors[floor_num][room_key] = {'room': alloc.room, 'students': []}
        record = attend_map.get(alloc.student.pk)
        floors[floor_num][room_key]['students'].append({
            'student': alloc.student,
            'record':  record,
            'status':  record.status if record else 'unmarked',
        })

    total   = student_qs.count()
    att_qs  = Attendance.objects.filter(date=selected_date, student__in=student_qs)
    present = att_qs.filter(status='present').count()
    absent  = att_qs.filter(status='absent').count()
    leave   = att_qs.filter(status='leave').count()

    departments = (
        Student.objects
        .filter(is_resident=True, allocations__room__hostel=hostel, allocations__status='active')
        .values_list('department', flat=True).distinct().order_by('department')
    )

    return render(request, 'admin/hostel_attendance_detail.html', {
        'hostel':        hostel,
        'selected_date': selected_date,
        'floors':        floors,
        'stats': {
            'total':       total,
            'present':     present,
            'absent':      absent,
            'leave':       leave,
            'unmarked':    total - present - absent - leave,
            'present_pct': round(present / total * 100) if total else 0,
            'absent_pct':  round(absent  / total * 100) if total else 0,
            'leave_pct':   round(leave   / total * 100) if total else 0,
        },
        'filter_year': filter_year,
        'filter_dept': filter_dept,
        'filter_sem':  filter_sem,
        'departments': departments,
        'years':       Student.Year.choices,
        'semesters':   range(1, 9),
    })


@admin_required
def toggle_attendance(request):
    """AJAX: cycle one student's attendance status for a date."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    student_id = data.get('student_id')
    date_str   = data.get('date')

    try:
        student = Student.objects.get(pk=student_id, is_resident=True)
    except (Student.DoesNotExist, Exception):
        return JsonResponse({'error': 'Student not found'}, status=404)

    # Wardens may only mark their hostel's students
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel:
        in_hostel = student.allocations.filter(
            status='active', room__hostel=warden_hostel
        ).exists()
        if not in_hostel:
            return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        selected_date = dt_date.fromisoformat(date_str)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date'}, status=400)

    # Cycle: unmarked → present → absent → leave → unmarked
    cycle = [None, 'present', 'absent', 'leave']
    record  = Attendance.objects.filter(student=student, date=selected_date).first()
    current = record.status if record else None
    try:
        idx = cycle.index(current)
    except ValueError:
        idx = 0
    next_status = cycle[(idx + 1) % len(cycle)]

    if next_status is None:
        if record:
            record.delete()
        new_status = 'unmarked'
    else:
        Attendance.objects.update_or_create(
            student=student,
            date=selected_date,
            defaults={'status': next_status, 'marked_by': request.user},
        )
        new_status = next_status

    return JsonResponse({'status': new_status, 'student_id': str(student.pk)})


# ─── student self-registration ───────────────────────────────────────────────

def student_register(request):
    """Public signup page for students."""
    if request.user.is_authenticated:
        return redirect('dashboard')

    # Use master Department list; fallback to existing student profile depts if list is empty
    master_depts  = list(Department.objects.filter(is_active=True).values_list('name', flat=True).order_by('name'))
    if not master_depts:
        master_depts = list(
            Student.objects.values_list('department', flat=True)
            .exclude(department__isnull=True).exclude(department__exact='').distinct().order_by('department')
        )
    departments = master_depts or [
        'Computer Science', 'Information Technology', 'Electronics & Communication',
        'Mechanical Engineering', 'Civil Engineering', 'Electrical Engineering',
        'MBA', 'MCA', 'BCA', 'B.Com', 'B.Sc', 'Other',
    ]

    if request.method == 'POST':
        name        = request.POST.get('name', '').strip()
        roll_number = request.POST.get('roll_number', '').strip()
        email       = request.POST.get('email', '').strip()
        phone       = request.POST.get('phone', '').strip()
        department  = request.POST.get('department', '').strip()
        year        = request.POST.get('year', '1')
        semester    = request.POST.get('semester', '1')
        password    = request.POST.get('password', '')
        confirm     = request.POST.get('confirm_password', '')

        errors = []
        if not all([name, roll_number, email, department, password]):
            errors.append('All required fields must be filled.')
        if password != confirm:
            errors.append('Passwords do not match.')
        if len(password) < 6:
            errors.append('Password must be at least 6 characters.')
        if User.objects.filter(email=email).exists():
            errors.append('An account with this email already exists.')
        if Student.objects.filter(roll_number=roll_number).exists():
            errors.append('A student with this roll number already exists.')

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'base/register.html', {
                'departments': departments,
                'post': request.POST,  
            })

        user = User.objects.create_user(
            email=email,
            password=password,
            name=name,
            role=User.Role.STUDENT,
        )
        Student.objects.create(
            user=user,
            roll_number=roll_number,
            name=name,
            email=email,
            phone=phone,
            department=department,
            year=year,
            semester=int(semester),
            is_resident=False,
        )
        messages.success(request, 'Account created successfully! Please sign in.')
        return redirect('login')

    return render(request, 'base/register.html', {'departments': departments})


# ─── registration form system (admin) ────────────────────────────────────────

@admin_only
def registration_form_list(request):
    is_superadmin = (request.user.role == User.Role.SUPER_ADMIN)
    form_categories = FormCategory.objects.all()

    if request.method == 'POST' and is_superadmin:
        action = request.POST.get('action', '')
        if action == 'add_form_category':
            name  = request.POST.get('cat_name', '').strip()
            icon  = request.POST.get('cat_icon', 'bi-file-earmark-text').strip()
            color = request.POST.get('cat_color', '#1565C0').strip()
            if name:
                t_years   = [int(y) for y in request.POST.getlist('cat_years') if y.isdigit()]
                t_hostels = request.POST.getlist('cat_hostels')
                t_depts   = request.POST.getlist('cat_depts')
                cat_obj, created = FormCategory.objects.get_or_create(
                    name=name,
                    defaults={
                        'icon': icon, 'color': color, 'created_by': request.user,
                        'target_years': t_years, 'target_hostels': t_hostels, 'target_depts': t_depts,
                    }
                )
                if not created:
                    cat_obj.icon = icon; cat_obj.color = color
                    cat_obj.target_years = t_years
                    cat_obj.target_hostels = t_hostels
                    cat_obj.target_depts = t_depts
                    cat_obj.save()
                messages.success(request, f'Category "{name}" {"added" if created else "updated"}.')
            else:
                messages.error(request, 'Category name is required.')
        elif action == 'delete_form_category':
            cat_pk = request.POST.get('cat_pk', '')
            cat = FormCategory.objects.filter(pk=cat_pk).first()
            if cat:
                cat.delete()
                messages.success(request, f'Category "{cat.name}" deleted.')
        return redirect('registration_form_list')

    qs = RegistrationForm.objects.annotate(
        submission_count=Count('submissions'),
        pending_count=Count('submissions', filter=Q(submissions__status='submitted')),
        approved_count=Count('submissions', filter=Q(submissions__status='approved')),
    ).select_related('form_category').order_by('-created_at')

    # Group forms by category
    categorized = {}
    for cat in form_categories:
        cat_forms = [f for f in qs if f.form_category_id == cat.pk]
        if cat_forms:
            categorized[cat] = cat_forms
    uncategorized = [f for f in qs if f.form_category_id is None]

    # For category-manager targeting UI
    hostels_qs   = Hostel.objects.order_by('type', 'name')
    hostels_by_type = {}
    for h in hostels_qs:
        hostels_by_type.setdefault(h.type, []).append(h)
    all_depts = sorted(set(
        list(Department.objects.filter(is_active=True).values_list('name', flat=True)) +
        list(Student.objects.values_list('department', flat=True).distinct())
    ))

    return render(request, 'admin/registration_forms.html', {
        'forms':           qs,
        'categorized':     categorized,
        'uncategorized':   uncategorized,
        'form_categories': form_categories,
        'is_superadmin':   is_superadmin,
        'hostels_by_type': hostels_by_type,
        'all_depts':       all_depts,
    })


@admin_only
def create_registration_form(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Only superadmin can create forms.')
        return redirect('registration_form_list')

    hostels         = Hostel.objects.all()
    field_types     = FormField.FieldType.choices
    form_categories = FormCategory.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        if not title:
            messages.error(request, 'Form title is required.')
            return render(request, 'admin/create_registration_form.html', {
                'hostels': hostels, 'field_types': field_types,
                'pref_range': range(1, 7), 'form_categories': form_categories,
            })

        cat_pk   = request.POST.get('form_category', '').strip()
        cat_obj  = FormCategory.objects.filter(pk=cat_pk).first() if cat_pk else None
        form_obj = RegistrationForm.objects.create(
            title=title,
            description=request.POST.get('description', '').strip(),
            form_category=cat_obj,
            deadline=request.POST.get('deadline') or None,
            max_preferences=int(request.POST.get('max_preferences', 3)),
            is_active=request.POST.get('is_active') == 'on',
            created_by=request.user,
        )
        _save_form_fields(request.POST, form_obj)
        messages.success(request, f'Form "{form_obj.title}" created successfully.')
        return redirect('registration_form_list')

    return render(request, 'admin/create_registration_form.html', {
        'hostels': hostels, 'field_types': field_types,
        'pref_range': range(1, 7), 'form_categories': form_categories,
    })


@admin_only
def edit_registration_form(request, pk):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Only superadmin can edit forms.')
        return redirect('registration_form_list')

    form_obj        = get_object_or_404(RegistrationForm, pk=pk)
    hostels         = Hostel.objects.all()
    field_types     = FormField.FieldType.choices
    form_categories = FormCategory.objects.all()

    if request.method == 'POST':
        cat_pk                   = request.POST.get('form_category', '').strip()
        form_obj.title           = request.POST.get('title', '').strip()
        form_obj.description     = request.POST.get('description', '').strip()
        form_obj.form_category   = FormCategory.objects.filter(pk=cat_pk).first() if cat_pk else None
        form_obj.deadline        = request.POST.get('deadline') or None
        form_obj.max_preferences = int(request.POST.get('max_preferences', 3))
        form_obj.is_active       = request.POST.get('is_active') == 'on'
        form_obj.save()
        form_obj.fields.all().delete()
        _save_form_fields(request.POST, form_obj)
        messages.success(request, 'Form updated.')
        return redirect('registration_form_list')

    return render(request, 'admin/create_registration_form.html', {
        'form_obj':        form_obj,
        'existing_fields': list(form_obj.fields.all()),
        'hostels':         hostels,
        'field_types':     field_types,
        'pref_range':      range(1, 7),
        'form_categories': form_categories,
        'edit_mode':       True,
    })


def _save_form_fields(post, form_obj):
    """Helper: read indexed field_* POST params and create FormField objects."""
    count = int(post.get('field_count', 0))
    for i in range(count):
        label = post.get(f'field_label_{i}', '').strip()
        if not label:
            continue
        opts_raw  = post.get(f'field_options_{i}', '')
        opts_list = [o.strip() for o in opts_raw.splitlines() if o.strip()]
        FormField.objects.create(
            form=form_obj,
            label=label,
            field_type=post.get(f'field_type_{i}', 'text'),
            options=opts_list,
            is_required=f'field_required_{i}' in post,
            order=i,
            help_text=post.get(f'field_help_{i}', '').strip(),
        )


@admin_only
def form_submissions(request, pk):
    form_obj = get_object_or_404(RegistrationForm, pk=pk)
    qs = RegistrationSubmission.objects.filter(form=form_obj).select_related('student', 'allocated_room').order_by('-submitted_at')

    status = request.GET.get('status', '')
    if status:
        qs = qs.filter(status=status)

    paginator = Paginator(qs, 20) 
    page      = paginator.get_page(request.GET.get('page'))

    counts = form_obj.submissions.values('status').annotate(n=Count('status'))
    status_counts = {c['status']: c['n'] for c in counts}

    return render(request, 'admin/form_submissions.html', {  
        'form_obj':       form_obj,
        'page_obj':       page,
        'filters':        {'status': status},
        'status_choices': RegistrationSubmission.Status.choices,
        'status_counts':  status_counts,
    })


@admin_only
def submission_detail(request, pk):
    submission = get_object_or_404(
        RegistrationSubmission.objects.select_related(
            'form', 'student', 'allocated_room__hostel', 'reviewed_by'
        ),
        pk=pk
    )
    responses   = submission.responses.select_related('field').order_by('field__order')
    preferences = submission.preferences.select_related('hostel')

    # For each preference find matching vacant rooms
    pref_matches = []
    for pref in preferences:
        qs = Room.objects.filter(status='vacant').annotate(
            active_count=Count('allocations', filter=Q(allocations__status='active'))
        ).select_related('hostel')
        if pref.hostel:    qs = qs.filter(hostel=pref.hostel)
        if pref.room_type: qs = qs.filter(room_type=pref.room_type)
        if pref.floor:     qs = qs.filter(floor=pref.floor)
        pref_matches.append({
            'pref':    pref,
            'rooms':   qs[:6],
            'count':   qs.count(),
        })

    if request.method == 'POST':
        action = request.POST.get('action', '')
        submission.admin_notes = request.POST.get('admin_notes', '').strip()
        submission.reviewed_by = request.user
        submission.reviewed_at = timezone.now()

        if action == 'approve':
            submission.status = 'approved'
            room_id = request.POST.get('allocate_room_id', '').strip()
            if room_id:
                room = get_object_or_404(Room, pk=room_id)
                if room.beds_available <= 0:
                    messages.error(request, f'Room {room.room_number} is full.')
                elif RoomAllocation.objects.filter(student=submission.student, status='active').exists():
                    messages.error(request, 'Student already has an active room allocation.')
                else:
                    RoomAllocation.objects.create(
                        student=submission.student,
                        room=room,         
                        allocated_by=request.user,   
                        check_in=dt_date.today(),
                        notes=f'Allocated via registration form: {submission.form.title}',
                    )
                    room.status = 'occupied'
                    room.save(update_fields=['status'])
                    submission.allocated_room = room
                    submission.student.is_resident = True
                    submission.student.save(update_fields=['is_resident'])
                    messages.success(request, f'Approved and room {room.room_number} allocated to {submission.student.name}.')
            else:
                messages.success(request, f'Application approved for {submission.student.name}.')
        elif action == 'reject':
            submission.status = 'rejected'
            messages.success(request, f'Application rejected for {submission.student.name}.')
        elif action == 'waitlist':
            submission.status = 'waitlisted'
            messages.success(request, f'{submission.student.name} waitlisted.')
        elif action == 'review':
            submission.status = 'under_review'
            messages.info(request, 'Marked as under review.')

        submission.save()
        return redirect('submission_detail', pk=pk)

    return render(request, 'admin/submission_detail.html', {
        'submission':   submission,
        'responses':    responses,
        'pref_matches': pref_matches,
    })


@admin_only
def form_analytics(request, pk):
    """Per-form analytics: response breakdown per field + submission status chart."""
    form_obj    = get_object_or_404(RegistrationForm.objects.select_related('form_category'), pk=pk)
    all_subs    = RegistrationSubmission.objects.filter(form=form_obj)
    total       = all_subs.count()

    # Status breakdown
    status_counts = {s: 0 for s, _ in RegistrationSubmission.Status.choices}
    for row in all_subs.values('status').annotate(n=Count('status')):
        status_counts[row['status']] = row['n']

    # Per-field analytics
    fields_data = []
    for field in form_obj.fields.order_by('order'):
        responses_qs = SubmissionResponse.objects.filter(
            field=field, submission__form=form_obj
        ).values_list('value', flat=True)
        values = [v for v in responses_qs if v.strip()]

        if field.field_type in ('select', 'radio', 'checkbox'):
            # Count each option
            count_map = {}
            for v in values:
                # checkbox may have comma-separated values
                for part in v.split(','):
                    part = part.strip()
                    if part:
                        count_map[part] = count_map.get(part, 0) + 1
            sorted_options = sorted(count_map.items(), key=lambda x: -x[1])
            fields_data.append({
                'field':    field,
                'type':     'chart',
                'options':  [{'label': k, 'count': v} for k, v in sorted_options],
                'total_responses': len(values),
            })
        else:
            # text-type: show list of unique answers
            from collections import Counter
            cnt = Counter(values)
            fields_data.append({
                'field':         field,
                'type':          'text',
                'values':        values[:200],
                'total_responses': len(values),
                'common':        cnt.most_common(10),
            })

    # Submission timeline: count per day (last 30 days)
    from django.db.models.functions import TruncDate
    timeline = (
        all_subs
        .annotate(day=TruncDate('submitted_at'))
        .values('day')
        .annotate(n=Count('id'))
        .order_by('day')
    )
    timeline_labels = [str(r['day']) for r in timeline]
    timeline_counts = [r['n'] for r in timeline]

    return render(request, 'admin/form_analytics.html', {
        'form_obj':      form_obj,
        'total':         total,
        'status_counts': status_counts,
        'status_choices': RegistrationSubmission.Status.choices,
        'fields_data':   fields_data,
        'timeline_labels': json.dumps(timeline_labels),
        'timeline_counts': json.dumps(timeline_counts),
    })


@admin_only
def export_form_excel(request, pk):
    """Export all form submissions + responses to Excel."""
    form_obj = get_object_or_404(RegistrationForm, pk=pk)
    subs = (
        RegistrationSubmission.objects
        .filter(form=form_obj)
        .select_related('student', 'allocated_room', 'reviewed_by')
        .prefetch_related('responses__field', 'preferences__hostel')
        .order_by('submitted_at')
    )
    fields = list(form_obj.fields.order_by('order'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Submissions'

    hdr_fill   = PatternFill('solid', fgColor='1565C0')
    hdr_font   = Font(color='FFFFFF', bold=True, size=11)
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_align = Alignment(vertical='top', wrap_text=True)
    thin       = Side(style='thin', color='CBD5E1')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    base_cols = ['#', 'Name', 'Roll No.', 'Dept', 'Year', 'Status', 'Submitted At', 'Reviewed By', 'Admin Notes', 'Allocated Room']
    all_cols  = base_cols + [f.label for f in fields]
    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    for ri, sub in enumerate(subs, 2):
        resp_map = {str(r.field_id): r.value for r in sub.responses.all()}
        row_data = [
            ri - 1,
            sub.student.name,
            sub.student.roll_number,
            sub.student.department,
            sub.student.year,
            sub.get_status_display(),
            sub.submitted_at.strftime('%d %b %Y %H:%M'),
            sub.reviewed_by.name if sub.reviewed_by else '',
            sub.admin_notes,
            sub.allocated_room.room_number if sub.allocated_room else '',
        ] + [resp_map.get(str(f.pk), '') for f in fields]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = body_align
            cell.border    = border
            if ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='EFF6FF')

    # Auto-width
    for ci in range(1, len(all_cols) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or '')) for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2.append(['Form', form_obj.title])
    ws2.append(['Category', form_obj.form_category.name if form_obj.form_category else '—'])
    ws2.append(['Total Submissions', subs.count()])
    ws2.append([])
    ws2.append(['Status', 'Count'])
    for val, label in RegistrationSubmission.Status.choices:
        ws2.append([label, subs.filter(status=val).count()])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_title = form_obj.title[:30].replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="form_{safe_title}.xlsx"'
    wb.save(response)
    return response


# ─── registration form system (student) ──────────────────────────────────────

@login_required
def student_announcements(request):
    today   = dt_date.today()
    student = getattr(request.user, 'student_profile', None)
    qs = (
        Announcement.objects
        .filter(is_active=True, audience__in=['all', 'students', 'custom'])
        .filter(Q(expires_at__isnull=True) | Q(expires_at__gte=today))
        .select_related('category', 'created_by', 'custom_audience')
        .order_by('-created_at')
    )
    # Resolve student's active hostel info once (avoids N+1 queries)
    student_hostel_id   = None
    student_hostel_type = None
    if student:
        active_alloc = (
            RoomAllocation.objects
            .filter(student=student, status='active')
            .select_related('room__hostel')
            .first()
        )
        if active_alloc:
            student_hostel_id   = str(active_alloc.room.hostel_id)
            student_hostel_type = active_alloc.room.hostel.type

    # Python-level sub-filtering for year / dept / hostel
    result = []
    for ann in qs:
        if student:
            if ann.target_years and student.year not in ann.target_years:
                continue
            if ann.target_depts and student.department not in ann.target_depts:
                continue
            # Hostel filtering — specific hostels take priority over type-level gender
            if ann.target_hostels:
                if student_hostel_id not in [str(h) for h in ann.target_hostels]:
                    continue
            elif ann.target_gender != 'all':
                # Legacy type-level filter
                if student_hostel_type != ann.target_gender:
                    continue
        result.append(ann)
    return render(request, 'student/announcements.html', {'announcements': result})


@login_required
def available_forms(request):
    student, _redir = _get_student_or_redirect(request)
    if _redir: return _redir
    today = dt_date.today()

    # Resolve student's hostel once
    student_hostel_id = None
    active_alloc = (
        RoomAllocation.objects
        .filter(student=student, status='active')
        .select_related('room__hostel')
        .first()
    )
    if active_alloc:
        student_hostel_id = str(active_alloc.room.hostel_id)

    all_forms = (
        RegistrationForm.objects
        .filter(is_active=True)
        .filter(Q(deadline__isnull=True) | Q(deadline__gte=today))
        .select_related('form_category')
        .annotate(already_applied=Count('submissions', filter=Q(submissions__student=student)))
        .order_by('-created_at')
    )

    # Filter by category audience targeting
    visible = []
    for f in all_forms:
        cat = f.form_category
        if cat:
            if cat.target_years and student.year not in cat.target_years:
                continue
            if cat.target_depts and student.department not in cat.target_depts:
                continue
            if cat.target_hostels and student_hostel_id not in [str(h) for h in cat.target_hostels]:
                continue
        visible.append(f)

    return render(request, 'student/available_forms.html', {
        'student': student,
        'forms':   visible,
        'today':   today,
    })


@login_required
def apply_form(request, form_pk):
    student, _redir = _get_student_or_redirect(request)
    if _redir: return _redir
    form_obj = get_object_or_404(RegistrationForm, pk=form_pk, is_active=True)

    if RegistrationSubmission.objects.filter(form=form_obj, student=student).exists():
        messages.warning(request, 'You have already applied to this form.')
        return redirect('my_applications')

    if form_obj.deadline and form_obj.deadline < dt_date.today():
        messages.error(request, 'The deadline for this form has passed.')
        return redirect('available_forms')

    if request.method == 'POST':
        submission = RegistrationSubmission.objects.create(   
            form=form_obj,
            student=student,
            status='submitted',         
        )
        for field in form_obj.fields.all():
            if field.field_type == 'checkbox':
                value = ', '.join(request.POST.getlist(f'field_{field.pk}'))
            else:
                value = request.POST.get(f'field_{field.pk}', '')
            SubmissionResponse.objects.create(
                submission=submission,
                field=field,
                value=value,
            )
        messages.success(request, f'Application for "{form_obj.title}" submitted successfully!')
        return redirect('my_applications')

    return render(request, 'student/apply_form.html', {
        'student':  student,
        'form_obj': form_obj,
        'fields':   form_obj.fields.all(),
    })


@login_required
def my_applications(request):
    student, _redir = _get_student_or_redirect(request)
    if _redir: return _redir
    submissions = RegistrationSubmission.objects.filter(
        student=student
    ).select_related('form', 'allocated_room__hostel').prefetch_related('preferences__hostel').order_by('-submitted_at')

    return render(request, 'student/my_applications.html', {
        'student':     student,
        'submissions': submissions,
    })
 

# ─── apply_form updated to save amenity preferences ───────────────────────────
# (replaced inline above — amenity IDs saved as JSON list per preference)


# ─── room amenity management ──────────────────────────────────────────────────

@admin_required
def amenity_list(request):
    warden_hostel = get_hostel_scope(request.user)

    if request.method == 'POST':
        action   = request.POST.get('action', 'add')

        # ── Amenity type management ──
        if action == 'add_type':
            type_name = request.POST.get('type_name', '').strip()
            type_icon = request.POST.get('type_icon', 'bi-star').strip() or 'bi-star'
            if type_name:
                RoomAmenity.objects.get_or_create(
                    name=type_name,
                    defaults={'icon': type_icon, 'created_by': request.user},
                )
                messages.success(request, f'Amenity type "{type_name}" added.')
            return redirect('amenity_list')

        if action == 'delete_type':
            type_pk = request.POST.get('type_pk', '').strip()
            if type_pk:
                at = get_object_or_404(RoomAmenity, pk=type_pk)
                at.delete()
                messages.success(request, f'Amenity type "{at.name}" deleted.')
            return redirect('amenity_list')

        # ── Record management ──
        rec_pk   = request.POST.get('rec_pk', '').strip()
        room_pk  = request.POST.get('room',  '').strip()
        amen_pk  = request.POST.get('amenity', '').strip()
        cond     = request.POST.get('condition', RoomAmenityRecord.Condition.GOOD)
        notes    = request.POST.get('notes', '').strip()

        if action == 'delete' and rec_pk:
            rec = get_object_or_404(RoomAmenityRecord, pk=rec_pk)
            if warden_hostel and rec.room.hostel_id != warden_hostel.pk:
                messages.error(request, 'Access denied.')
            else:
                rec.delete()
                messages.success(request, 'Amenity record removed.')
            return redirect('amenity_list')

        amen_name = request.POST.get('amenity_name', '').strip()
        if room_pk and amen_name:
            room = get_object_or_404(Room, pk=room_pk)
            if warden_hostel and room.hostel_id != warden_hostel.pk:
                messages.error(request, 'Access denied.')
            else:
                # Get or create the amenity type by name (case-insensitive lookup)
                amenity = RoomAmenity.objects.filter(name__iexact=amen_name).first()
                if not amenity:
                    amenity = RoomAmenity.objects.create(
                        name=amen_name, created_by=request.user
                    )
                qty = max(1, int(request.POST.get('quantity', 1) or 1))
                if rec_pk:
                    rec = get_object_or_404(RoomAmenityRecord, pk=rec_pk)
                    rec.room = room; rec.amenity = amenity
                    rec.quantity = qty; rec.condition = cond; rec.notes = notes; rec.save()
                    messages.success(request, 'Record updated.')
                else:
                    RoomAmenityRecord.objects.create(
                        room=room, amenity=amenity, quantity=qty, condition=cond,
                        notes=notes, added_by=request.user,
                    )
                    messages.success(request, f'"{amenity.name}" (×{qty}) added to {room}.')
        return redirect('amenity_list')

    # Scope records to warden's hostel if applicable
    rec_qs = (
        RoomAmenityRecord.objects
        .select_related('room__hostel', 'amenity', 'added_by')
        .filter(room__hostel=warden_hostel) if warden_hostel
        else RoomAmenityRecord.objects.select_related('room__hostel', 'amenity', 'added_by')
    ).order_by('-created_at')

    # Rooms available to add (scoped)
    room_qs = (
        Room.objects.filter(hostel=warden_hostel) if warden_hostel
        else Room.objects.all()
    ).select_related('hostel').order_by('hostel__name', 'floor', 'room_number')

    amenity_types = RoomAmenity.objects.all().order_by('name')
    hostels       = Hostel.objects.all().order_by('name')
    conditions    = RoomAmenityRecord.Condition.choices

    return render(request, 'admin/amenities.html', {
        'records':       rec_qs,
        'rooms':         room_qs,
        'amenity_types': amenity_types,
        'hostels':       hostels,
        'conditions':    conditions,
        'is_warden':     warden_hostel is not None,
        'warden_hostel': warden_hostel,
    })


@admin_only
@require_POST
def delete_amenity(request, pk):
    """Legacy: delete an amenity type definition."""
    a = get_object_or_404(RoomAmenity, pk=pk)
    name = a.name
    a.delete()
    messages.success(request, f'Amenity type "{name}" deleted.')
    return redirect('amenity_list')


@admin_only
def edit_room(request, pk):
    room      = get_object_or_404(Room, pk=pk)
    hostels   = Hostel.objects.all()
    amenities = RoomAmenity.objects.all()

    if request.method == 'POST':
        room.hostel      = get_object_or_404(Hostel, pk=request.POST.get('hostel'))
        room.room_number = request.POST.get('room_number', '').strip()
        room.floor       = int(request.POST.get('floor', 1))
        room.room_type   = request.POST.get('room_type', 'double')
        room.capacity    = int(request.POST.get('capacity', 2))
        room.status      = request.POST.get('status', 'vacant')
        room.save()
        selected = request.POST.getlist('amenities')
        room.amenities.set(RoomAmenity.objects.filter(pk__in=selected))
        messages.success(request, f'Room {room.room_number} updated.')
        return redirect('room_list')

    return render(request, 'admin/edit_room.html', {
        'room':               room,
        'hostels':            hostels,
        'amenities':          amenities,
        'status_choices':     Room.Status.choices,
        'selected_amenities': list(room.amenities.values_list('pk', flat=True)),
        'room_categories':    ROOM_CATEGORIES_GROUPED,
    })


# ─── announcements (admin) ────────────────────────────────────────────────────

@login_required
def announcement_list(request):
    """
    Admin/SuperAdmin  → full management (create, edit, delete, toggle).
    All other staff   → filtered view-only based on their role.
    """
    user     = request.user
    is_admin = user.role in (User.Role.ADMIN, User.Role.SUPER_ADMIN)
    today    = dt_date.today()

    if request.method == 'POST':
        if not is_admin:
            messages.error(request, 'Only Admin 2 and Admin 1 can manage announcements.')
            return redirect('announcement_list')

        action = request.POST.get('action', 'save')

        # ── Category management ──
        if action == 'add_category':
            cname = request.POST.get('cat_name', '').strip()
            cicon = request.POST.get('cat_icon', 'bi-megaphone').strip() or 'bi-megaphone'
            ccolor = request.POST.get('cat_color', '#1565C0').strip() or '#1565C0'
            if cname:
                AnnouncementCategory.objects.get_or_create(
                    name=cname, defaults={'icon': cicon, 'color': ccolor, 'created_by': user}
                )
                messages.success(request, f'Category "{cname}" added.')
            return redirect('announcement_list')

        if action == 'delete_category':
            cat_pk = request.POST.get('cat_pk', '').strip()
            if cat_pk:
                cat = get_object_or_404(AnnouncementCategory, pk=cat_pk)
                cat.delete()
                messages.success(request, 'Category deleted.')
            return redirect('announcement_list')

        # ── Custom audience management ──
        if action == 'add_audience':
            aname = request.POST.get('aud_name', '').strip()
            adesc = request.POST.get('aud_desc', '').strip()
            aicon = request.POST.get('aud_icon', 'bi-people-fill').strip() or 'bi-people-fill'
            acolor = request.POST.get('aud_color', '#059669').strip() or '#059669'
            if aname:
                aud = AnnouncementAudience.objects.filter(name__iexact=aname).first()
                if not aud:
                    AnnouncementAudience.objects.create(
                        name=aname, description=adesc, icon=aicon,
                        color=acolor, created_by=user
                    )
                    messages.success(request, f'Audience group "{aname}" created.')
                else:
                    messages.warning(request, f'Audience group "{aname}" already exists.')
            return redirect('announcement_list')

        if action == 'delete_audience':
            aud_pk = request.POST.get('aud_pk', '').strip()
            if aud_pk:
                aud = get_object_or_404(AnnouncementAudience, pk=aud_pk)
                aud.delete()
                messages.success(request, 'Audience group deleted.')
            return redirect('announcement_list')

        # ── Announcement save ──
        pk         = request.POST.get('pk', '').strip()
        title      = request.POST.get('title', '').strip()
        content    = request.POST.get('content', '').strip()
        priority   = request.POST.get('priority', 'info')
        expires_at = request.POST.get('expires_at') or None
        is_active  = request.POST.get('is_active') == 'on'
        cat_name   = request.POST.get('category_name', '').strip()
        t_years    = request.POST.getlist('target_years')
        t_depts    = request.POST.getlist('target_depts')
        t_gender   = request.POST.get('target_gender', 'all')
        t_hostels  = request.POST.getlist('target_hostels')  # list of hostel UUID strings

        # Resolve audience — custom groups are posted as "custom_<uuid>"
        audience_raw    = request.POST.get('audience', 'all')
        custom_aud_id   = None
        if audience_raw.startswith('custom_'):
            custom_aud_id = audience_raw[7:]   # strip 'custom_' prefix → UUID string
            audience_val  = 'custom'
        else:
            audience_val = audience_raw

        category = None
        if cat_name:
            category, _ = AnnouncementCategory.objects.get_or_create(
                name=cat_name, defaults={'created_by': user}
            )

        if title and content:
            attachment_file = request.FILES.get('attachment')
            if pk:
                ann = get_object_or_404(Announcement, pk=pk)
                ann.title = title; ann.content = content; ann.priority = priority
                ann.audience = audience_val; ann.category = category
                ann.custom_audience_id = custom_aud_id or None
                ann.target_years = t_years; ann.target_depts = t_depts
                ann.target_gender = t_gender; ann.target_hostels = t_hostels
                ann.expires_at = expires_at; ann.is_active = is_active
                if attachment_file:
                    if ann.attachment:
                        ann.attachment.delete(save=False)
                    ann.attachment = attachment_file
                elif request.POST.get('clear_attachment'):
                    if ann.attachment:
                        ann.attachment.delete(save=False)
                    ann.attachment = None
                ann.save()
                messages.success(request, 'Announcement updated.')
            else:
                ann = Announcement(
                    title=title, content=content, priority=priority,
                    audience=audience_val, category=category,
                    custom_audience_id=custom_aud_id or None,
                    target_years=t_years, target_depts=t_depts,
                    target_gender=t_gender, target_hostels=t_hostels,
                    expires_at=expires_at,
                    is_active=is_active, created_by=user,
                )
                if attachment_file:
                    ann.attachment = attachment_file
                ann.save()
                messages.success(request, f'Announcement "{title}" published.')
        return redirect('announcement_list')

    # ── GET — build queryset scoped to role ──
    base_qs = (
        Announcement.objects
        .select_related('category', 'created_by', 'custom_audience')
        .filter(Q(expires_at__isnull=True) | Q(expires_at__gte=today))
        .order_by('-created_at')
    )

    if is_admin:
        ann_qs = base_qs  # admin sees everything
    else:
        role = user.role
        role_audiences = {
            User.Role.WARDEN:      ['all', 'staff_all', 'warden'],
            User.Role.SECURITY:    ['all', 'staff_all', 'security'],
            User.Role.MAINTENANCE: ['all', 'staff_all', 'maintenance'],
            User.Role.MESS:        ['all', 'staff_all', 'mess'],
        }
        allowed = role_audiences.get(role, ['all', 'staff_all'])
        ann_qs  = base_qs.filter(is_active=True, audience__in=allowed)

    categories       = AnnouncementCategory.objects.all().order_by('name')
    custom_audiences = AnnouncementAudience.objects.all().order_by('name')
    master_depts  = list(Department.objects.filter(is_active=True).values_list('name', flat=True))
    student_depts = list(
        Student.objects.values_list('department', flat=True)
        .exclude(department__isnull=True).exclude(department__exact='')
        .distinct()
    )
    departments = sorted(set(master_depts) | set(student_depts))

    # Hostels grouped by type for announcement targeting
    all_hostels = list(Hostel.objects.order_by('type', 'name'))
    hostels_by_type = {}
    hostel_map = {}   # uuid_str → hostel name for badge display
    for h in all_hostels:
        hostels_by_type.setdefault(h.type, []).append(h)
        hostel_map[str(h.pk)] = h.name

    return render(request, 'admin/announcements.html', {
        'announcements':    ann_qs,
        'categories':       categories,
        'custom_audiences': custom_audiences,
        'priorities':       Announcement.Priority.choices,
        'audiences':        Announcement.Audience.choices,
        'year_choices':     Student.Year.choices,
        'departments':      departments,
        'gender_choices':   Announcement.GENDER_CHOICES,
        'hostels_by_type':  hostels_by_type,
        'hostel_map':       hostel_map,
        'hostel_type_labels': {'boys': 'Boys Hostels', 'girls': 'Girls Hostels', 'staff': 'Staff Hostels'},
        'is_admin':         is_admin,
    })


@admin_only
@require_POST
def toggle_announcement(request, pk):
    ann = get_object_or_404(Announcement, pk=pk)
    ann.is_active = not ann.is_active
    ann.save(update_fields=['is_active'])
    messages.success(request, f'Announcement {"activated" if ann.is_active else "deactivated"}.')
    return redirect('announcement_list')


@admin_only
@require_POST
def delete_announcement(request, pk):
    ann = get_object_or_404(Announcement, pk=pk)
    if ann.attachment:
        ann.attachment.delete(save=False)
    ann.delete()
    messages.success(request, 'Announcement deleted.')
    return redirect('announcement_list')


# ─── student profile edit (admin) ────────────────────────────────────────────

@admin_only
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.name           = request.POST.get('name', '').strip()
        student.email          = request.POST.get('email', '').strip()
        student.phone          = request.POST.get('phone', '').strip()
        student.department     = request.POST.get('department', '').strip()
        student.year           = request.POST.get('year', '1')
        student.semester       = int(request.POST.get('semester', 1))
        student.gender         = request.POST.get('gender', '')
        student.guardian_name  = request.POST.get('guardian_name', '').strip()
        student.guardian_phone = request.POST.get('guardian_phone', '').strip()
        student.address        = request.POST.get('address', '').strip()
        dob = request.POST.get('date_of_birth', '').strip()
        student.date_of_birth  = dob if dob else None
        if 'photo' in request.FILES:
            student.photo = request.FILES['photo']
        student.save()
        messages.success(request, f'Profile updated for {student.name}.')
        return redirect('student_detail', pk=pk)

    master_depts  = list(Department.objects.filter(is_active=True).values_list('name', flat=True))
    student_depts = list(
        Student.objects.values_list('department', flat=True)
        .exclude(department__isnull=True).exclude(department__exact='').distinct()
    )
    all_departments = sorted(set(master_depts) | set(student_depts))
    return render(request, 'admin/edit_student.html', {
        'student': student,
        'all_departments': all_departments,
    })


# ─── room transfer (admin) ────────────────────────────────────────────────────

@admin_only
def transfer_room(request, pk):
    student = get_object_or_404(Student, pk=pk)
    current = RoomAllocation.objects.filter(
        student=student, status='active'
    ).select_related('room__hostel').first()

    available_rooms = Room.objects.select_related('hostel').annotate(
        active_count=Count('allocations', filter=Q(allocations__status='active'))
    ).filter(
        active_count__lt=F('capacity'),
        status__in=['vacant', 'occupied'],
    ).order_by('hostel__name', 'floor', 'room_number')
    if current:
        available_rooms = available_rooms.exclude(pk=current.room.pk)

    if request.method == 'POST':
        new_room = get_object_or_404(Room, pk=request.POST.get('new_room'))
        if new_room.beds_available <= 0:
            messages.error(request, f'Room {new_room.room_number} is now full.')
            return redirect('transfer_room', pk=pk)

        if current:
            old_room        = current.room
            current.status    = 'checkout'
            current.check_out = dt_date.today()
            current.notes    += f'\nTransferred to Room {new_room.room_number} on {dt_date.today()}'
            current.save()
            if old_room.current_occupants == 0:
                old_room.status = 'vacant'
                old_room.save(update_fields=['status'])

        RoomAllocation.objects.create(
            student=student,
            room=new_room,
            allocated_by=request.user,
            check_in=dt_date.today(),
            notes=f'Transferred from Room {current.room.room_number if current else "N/A"}',
        )
        new_room.status = 'occupied'
        new_room.save(update_fields=['status'])
        messages.success(request, f'{student.name} transferred to Room {new_room.room_number}.')
        return redirect('student_detail', pk=pk)

    return render(request, 'admin/transfer_room.html', {
        'student':         student,
        'current':         current,
        'available_rooms': available_rooms,
    })


# ─── fee defaulters report (admin) ───────────────────────────────────────────

@admin_only
def fee_defaulters(request):
    qs = FeeRecord.objects.filter(
        status__in=['pending', 'partial']
    ).select_related('student', 'fee_structure').order_by('student__roll_number')

    if request.GET.get('export') == 'csv':     
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fee_defaulters.csv"'
        w = csv.writer(response)
        w.writerow(['Roll No', 'Name', 'Department', 'Room', 'Fee Structure',
                    'Total Due', 'Amount Paid', 'Balance', 'Status'])
        for r in qs:
            room = r.student.current_room
            w.writerow([
                r.student.roll_number, r.student.name, r.student.department,
                f'{room.room_number} — {room.hostel.name}' if room else 'Not Allocated',
                str(r.fee_structure),
                r.total_amount, r.amount_paid,
                r.total_amount - r.amount_paid,
                r.get_status_display(),
            ]) 
        return response

    summary = qs.aggregate(due=Sum('total_amount'), paid=Sum('amount_paid'))
    total_due  = summary['due']  or 0
    total_paid = summary['paid'] or 0 
 
    return render(request, 'admin/fee_defaulters.html', {
        'records':       qs,
        'total_due':     total_due,
        'total_paid':    total_paid,
        'total_balance': total_due - total_paid,
         'count':         qs.count(),
    })


# ─── occupancy report ─────────────────────────────────────────────────────────

def _hostel_occupancy_stats(hostel):
    """Return per-floor list and aggregate stats for one hostel."""
    rooms = (
        Room.objects.filter(hostel=hostel)
        .annotate(active_allocs=Count('allocations', filter=Q(allocations__status='active')))
        .order_by('floor', 'room_number')
    )
    floors = OrderedDict()
    for room in rooms:
        f = room.floor
        if f not in floors:
            floors[f] = {'floor': f, 'total': 0, 'occupied': 0, 'partial': 0, 'vacant': 0, 'maint': 0}
        floors[f]['total'] += 1
        if room.status == 'maintenance':
            floors[f]['maint'] += 1
        elif room.active_allocs >= room.capacity and room.active_allocs > 0:
            floors[f]['occupied'] += 1
        elif room.active_allocs > 0:
            floors[f]['partial'] += 1
        else:
            floors[f]['vacant'] += 1

    fl_list  = list(floors.values())
    total    = sum(f['total']    for f in fl_list)
    occupied = sum(f['occupied'] for f in fl_list)
    partial  = sum(f['partial']  for f in fl_list)
    vacant   = sum(f['vacant']   for f in fl_list)
    maint    = sum(f['maint']    for f in fl_list)
    pct      = round((occupied + partial) / total * 100) if total else 0
    return fl_list, {'total': total, 'occupied': occupied, 'partial': partial,
                     'vacant': vacant, 'maint': maint, 'pct': pct}


@admin_required
def occupancy_report(request):
    """Warden → redirect to their hostel detail. Admin/SuperAdmin → hostel overview cards."""
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel:
        return redirect(reverse('hostel_occupancy_detail', kwargs={'hostel_pk': warden_hostel.pk}))

    hostels = Hostel.objects.all().order_by('name')
    cards   = []
    for h in hostels:
        _, stats = _hostel_occupancy_stats(h)
        total_beds    = Bed.objects.filter(room__hostel=h).count()
        occupied_beds = Bed.objects.filter(room__hostel=h, status='occupied').count()
        cards.append({
            'hostel': h,
            'total_beds': total_beds, 'occupied_beds': occupied_beds,
            **stats,
        })

    overall = {
        'total':    sum(c['total']    for c in cards),
        'occupied': sum(c['occupied'] for c in cards),
        'partial':  sum(c['partial']  for c in cards),
        'vacant':   sum(c['vacant']   for c in cards),
        'maint':    sum(c['maint']    for c in cards),
    }
    ov_denom = overall['total']
    overall['pct'] = round((overall['occupied'] + overall['partial']) / ov_denom * 100) if ov_denom else 0

    return render(request, 'admin/occupancy_report.html', {
        'cards':   cards,
        'overall': overall,
        # bar-chart data for overview
        'chart_labels':   json.dumps([c['hostel'].name for c in cards]),
        'chart_occupied': json.dumps([c['occupied']    for c in cards]),
        'chart_partial':  json.dumps([c['partial']     for c in cards]),
        'chart_vacant':   json.dumps([c['vacant']      for c in cards]),
        'chart_pct':      json.dumps([c['pct']         for c in cards]),
    })


@admin_required
def hostel_occupancy_detail(request, hostel_pk):
    """Per-hostel occupancy breakdown with floor-wise charts."""
    hostel = get_object_or_404(Hostel, pk=hostel_pk)
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel and warden_hostel.pk != hostel.pk:
        messages.error(request, 'Access denied.')
        return redirect('occupancy_report')

    fl_list, stats = _hostel_occupancy_stats(hostel)
    total_beds    = Bed.objects.filter(room__hostel=hostel).count()
    occupied_beds = Bed.objects.filter(room__hostel=hostel, status='occupied').count()
    stats.update({'total_beds': total_beds, 'occupied_beds': occupied_beds})

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Occupancy Report'

        hdr_fill  = PatternFill('solid', fgColor='0F4C81')
        hdr_font  = Font(bold=True, color='FFFFFF', size=11)
        tot_fill  = PatternFill('solid', fgColor='1565C0')
        tot_font  = Font(bold=True, color='FFFFFF', size=11)
        alt_fill  = PatternFill('solid', fgColor='EBF3FF')
        red_fill  = PatternFill('solid', fgColor='FEE2E2')
        yel_fill  = PatternFill('solid', fgColor='FFFBEB')
        grn_fill  = PatternFill('solid', fgColor='ECFDF5')
        thin      = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )
        center    = Alignment(horizontal='center', vertical='center')
        headers   = ['Hostel', 'Floor', 'Total Rooms', 'Fully Occupied',
                     'Partially Occupied', 'Vacant', 'Maintenance', 'Occupancy %']
        col_widths = [28, 12, 14, 16, 20, 10, 14, 14]

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=f'Occupancy Report — {hostel.name}')
        title_cell.font      = Font(bold=True, color='FFFFFF', size=13)
        title_cell.fill      = PatternFill('solid', fgColor='0D47A1')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Header row
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 22

        # Data rows
        for ri, fl in enumerate(fl_list, 3):
            fpct = round((fl['occupied'] + fl['partial']) / fl['total'] * 100) if fl['total'] else 0
            row  = [hostel.name, f"Floor {fl['floor']}", fl['total'],
                    fl['occupied'], fl['partial'], fl['vacant'], fl['maint'], f'{fpct}%']
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = center; cell.border = thin
                # colour-code the count cells
                if ci == 4 and fl['occupied']:   cell.fill = red_fill
                elif ci == 5 and fl['partial']:  cell.fill = yel_fill
                elif ci == 6 and fl['vacant']:   cell.fill = grn_fill
                elif fill:                       cell.fill = fill

        # Totals row
        tr = len(fl_list) + 3
        totals = [hostel.name, 'TOTAL', stats['total'], stats['occupied'],
                  stats['partial'], stats['vacant'], stats['maint'], f"{stats['pct']}%"]
        for ci, val in enumerate(totals, 1):
            cell = ws.cell(row=tr, column=ci, value=val)
            cell.font = tot_font; cell.fill = tot_fill
            cell.alignment = center; cell.border = thin
        ws.row_dimensions[tr].height = 20

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"occupancy_{hostel.name.replace(' ', '_')}.xlsx"
        resp  = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    return render(request, 'admin/hostel_occupancy_detail.html', {
        'hostel': hostel,
        'floors': fl_list,
        'stats':  stats,
        # floor-wise bar chart
        'chart_labels':   json.dumps([f'Floor {f["floor"]}' for f in fl_list]),
        'chart_occupied': json.dumps([f['occupied'] for f in fl_list]),
        'chart_partial':  json.dumps([f['partial']  for f in fl_list]),
        'chart_vacant':   json.dumps([f['vacant']   for f in fl_list]),
        # doughnut chart (overall breakdown)
        'donut_data':  json.dumps([stats['occupied'], stats['partial'], stats['vacant'], stats['maint']]),
        'donut_labels': json.dumps(['Fully Occupied', 'Partially Occupied', 'Vacant', 'Maintenance']),
    })


# ─── Gate Pass ───────────────────────────────────────────────────────────────

@login_required
def apply_gate_pass(request):
    """Student applies for a new gate pass."""
    from datetime import datetime as _dt
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir
    restriction = _get_active_restriction(student, 'gate_pass')
    if restriction:
        return render(request, 'student/restricted.html', {
            'restriction': restriction, 'action': 'apply for a gate pass'
        })

    quota = _get_quota_status(student, 'gate_pass')
    if quota and quota['blocked']:
        return render(request, 'student/quota_exceeded.html', {
            'quota': quota, 'action': 'gate pass', 'back_url': 'my_gate_passes'
        })

    allocation    = RoomAllocation.objects.filter(student=student, status='active').select_related('room__hostel').first()
    gp_categories = GatePassCategory.objects.filter(is_active=True)

    def _ctx(extra=None):
        base = {
            'gp_categories': gp_categories, 'student': student,
            'allocation': allocation, 'quota': quota,
        }
        if extra:
            base.update(extra)
        return base

    if request.method == 'POST':
        departure_date       = request.POST.get('departure_date')
        departure_time       = request.POST.get('departure_time')
        expected_return_date = request.POST.get('expected_return_date')
        expected_return_time = request.POST.get('expected_return_time')
        destination          = request.POST.get('destination', '').strip()
        reason               = request.POST.get('reason', '').strip()
        category_id          = request.POST.get('category') or None

        if not all([departure_date, departure_time, expected_return_date, expected_return_time, destination, reason]):
            messages.error(request, 'All fields are required.')
            return render(request, 'student/apply_gate_pass.html', _ctx())

        # Enforce category application time window
        if category_id:
            cat = GatePassCategory.objects.filter(pk=category_id, is_active=True).first()
            if cat and (cat.start_time or cat.end_time):
                now_time = timezone.localtime(timezone.now()).time()
                if cat.start_time and now_time < cat.start_time:
                    messages.error(request, f'Gate passes for "{cat.name}" can only be applied from {cat.start_time.strftime("%I:%M %p")}.')
                    return render(request, 'student/apply_gate_pass.html', _ctx())
                if cat.end_time and now_time > cat.end_time:
                    messages.error(request, f'Gate passes for "{cat.name}" must be applied before {cat.end_time.strftime("%I:%M %p")}.')
                    return render(request, 'student/apply_gate_pass.html', _ctx())

        GatePass.objects.create(
            student              = student,
            category_id          = category_id,
            reason               = reason,
            destination          = destination,
            departure_date       = departure_date,
            departure_time       = departure_time,
            expected_return_date = expected_return_date,
            expected_return_time = expected_return_time,
        )
        send_notification(
            _staff_targets(student),
            f'New Gate Pass Request — {student.name}',
            f'{student.name} ({student.roll_number}) applied for a gate pass to {destination} on {departure_date}.',
            notif_type='gate_pass',
            link='/gatepasses/',
        )
        messages.success(request, 'Gate pass application submitted successfully.')
        return redirect('my_gate_passes')

    return render(request, 'student/apply_gate_pass.html', _ctx())


@login_required
def my_gate_passes(request):
    """Student views their own gate passes."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    passes = GatePass.objects.filter(student=student).select_related('approved_by')
    return render(request, 'student/my_gate_passes.html', {'passes': passes})


@admin_required
def gate_pass_list(request):
    """Admin/Warden sees gate pass requests (warden scoped to their hostel)."""
    warden_hostel = get_hostel_scope(request.user)

    g = request.GET
    status_filter  = g.get('status', '')
    search         = g.get('q', '').strip()
    hostel_id      = g.get('hostel', '')
    department     = g.get('department', '').strip()
    category_id    = g.get('category', '')
    room_number    = g.get('room', '').strip()
    date_from      = g.get('date_from', '').strip()
    date_to        = g.get('date_to', '').strip()
    year_filter    = g.get('year', '')
    semester_filter = g.get('semester', '')
    gender_filter   = g.get('gender', '')
    state_filter    = g.get('state', '').strip()

    _alloc_prefetch = Prefetch(
        'student__allocations',
        queryset=RoomAllocation.objects.filter(status='active').select_related('room__hostel'),
        to_attr='active_allocs',
    )
    base_qs = GatePass.objects.select_related(
        'student', 'approved_by', 'category',
        'exit_allowed_by', 'entry_allowed_by',
    ).prefetch_related(_alloc_prefetch)
    if warden_hostel:
        base_qs = base_qs.filter(
            student__allocations__room__hostel=warden_hostel,
            student__allocations__status='active',
        ).distinct()

    qs = base_qs.order_by('-created_at')

    if status_filter:
        qs = qs.filter(status=status_filter)
    if search:
        qs = qs.filter(
            Q(student__name__icontains=search) |
            Q(student__roll_number__icontains=search) |
            Q(gp_id__icontains=search) |
            Q(destination__icontains=search)
        )
    if hostel_id:
        qs = qs.filter(
            student__allocations__room__hostel_id=hostel_id,
            student__allocations__status='active',
        ).distinct()
    if department:
        qs = qs.filter(student__department__icontains=department)
    if category_id:
        qs = qs.filter(category_id=category_id)
    if room_number:
        qs = qs.filter(
            student__allocations__room__room_number__icontains=room_number,
            student__allocations__status='active',
        ).distinct()
    if date_from:
        qs = qs.filter(departure_date__gte=date_from)
    if date_to:
        qs = qs.filter(departure_date__lte=date_to)
    if year_filter:
        qs = qs.filter(student__year=year_filter)
    if semester_filter:
        qs = qs.filter(student__semester=semester_filter)
    if gender_filter:
        qs = qs.filter(student__gender=gender_filter)
    if state_filter:
        qs = qs.filter(student__state__iexact=state_filter)

    counts = {
        'all':      base_qs.count(),
        'pending':  base_qs.filter(status='pending').count(),
        'approved': base_qs.filter(status='approved').count(),
        'rejected': base_qs.filter(status='rejected').count(),
        'returned': base_qs.filter(status='returned').count(),
    }

    paginator = Paginator(qs, 20)
    page = paginator.get_page(g.get('page'))

    all_hostels    = Hostel.objects.order_by('name') if not warden_hostel else []
    gp_categories  = GatePassCategory.objects.filter(is_active=True).order_by('name')
    departments    = Student.objects.values_list('department', flat=True).distinct().order_by('department')
    states         = Student.objects.exclude(state='').values_list('state', flat=True).distinct().order_by('state')
    years          = Student.Year.choices

    active_filters = any([search, hostel_id, department, category_id, room_number, date_from, date_to, year_filter, semester_filter, gender_filter, state_filter])

    return render(request, 'admin/gate_passes.html', {
        'page':           page,
        'status_filter':  status_filter,
        'counts':         counts,
        'statuses':       GatePass.Status.choices,
        'warden_hostel':  warden_hostel,
        # filter values (repopulate form)
        'q':              search,
        'hostel_id':      hostel_id,
        'f_department':   department,
        'f_category':     category_id,
        'f_room':         room_number,
        'date_from':      date_from,
        'date_to':        date_to,
        'year_filter':    year_filter,
        'semester_filter': semester_filter,
        'gender_filter':  gender_filter,
        'state_filter':   state_filter,
        # filter options
        'all_hostels':    all_hostels,
        'gp_categories':  gp_categories,
        'departments':    departments,
        'states':         states,
        'years':          years,
        'active_filters': active_filters,
    })


@admin_required
def gate_pass_detail(request, pk):
    """Admin views a single gate pass and takes action."""
    gp = get_object_or_404(GatePass, pk=pk)

    if request.method == 'POST':
        if not request.user.is_admin_or_warden:
            messages.error(request, 'Only admin or warden can take action on gate passes.')
            return redirect('gate_pass_detail', pk=pk)

        action        = request.POST.get('action')
        admin_remarks = request.POST.get('admin_remarks', '').strip()

        if action == 'approve':
            gp.status        = GatePass.Status.APPROVED
            gp.approved_by   = request.user
            gp.approved_at   = timezone.now()
            gp.admin_remarks = admin_remarks
            gp.save()
            messages.success(request, f'Gate pass for {gp.student.name} approved.')
            # Notify student
            send_notification(
                [gp.student.user],
                'Gate Pass Approved',
                f'Your gate pass to {gp.destination} has been approved by {request.user.name}.'
                f'Proceed to the gate — the security guard will permit your exit.',
                notif_type='gate_pass',
            )
            # Notify all active security guards
            guards = list(User.objects.filter(role=User.Role.SECURITY, is_active=True))
            send_notification(
                guards,
                f'Gate Pass Approved — {gp.student.name}',
                f'{gp.student.name} ({gp.student.roll_number}) gate pass approved for '
                f'{gp.destination}. Permit exit from the Security Dashboard when student arrives.',
                notif_type='gate_pass',
            )

        elif action == 'reject':
            gp.status        = GatePass.Status.REJECTED
            gp.approved_by   = request.user
            gp.approved_at   = timezone.now()
            gp.admin_remarks = admin_remarks
            gp.save()
            messages.warning(request, f'Gate pass for {gp.student.name} rejected.')

        elif action == 'mark_returned':
            actual_return_date = request.POST.get('actual_return_date')
            gp.status             = GatePass.Status.RETURNED
            gp.actual_return_date = actual_return_date or dt_date.today()
            gp.save()
            messages.success(request, f'{gp.student.name} marked as returned.')

        return redirect('gate_pass_detail', pk=pk)

    return render(request, 'admin/gate_pass_detail.html', {
        'gp': gp,
        'can_action': request.user.is_admin_or_warden,
    })

# ─── Notification Helper ──────────────────────────────────────────────────────

def send_notification(recipients, title, message, notif_type='general', link=''):
    if not isinstance(recipients, (list, tuple, set)):
        recipients = [recipients]
    Notification.objects.bulk_create([
        Notification(recipient=u, title=title, message=message,
                     notification_type=notif_type, link=link, popup_shown=False)
        for u in recipients if u is not None
    ])


# ─── Notifications ────────────────────────────────────────────────────────────

@login_required
def notification_list(request):
    notifs = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    notifs.filter(is_read=False).update(is_read=True)
    paginator = Paginator(notifs, 25)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'notifications/list.html', {'page': page})


@login_required
@require_POST
def clear_all_notifications(request):
    Notification.objects.filter(recipient=request.user).delete()
    messages.success(request, 'All notifications cleared.')
    return redirect('notification_list')


@login_required
def poll_notifications(request):
    """Return any notifications not yet shown as popup, then mark them shown."""
    pending = list(
        Notification.objects.filter(recipient=request.user, popup_shown=False)
        .order_by('created_at')
        .values('title', 'message', 'notification_type', 'link')
    )
    Notification.objects.filter(recipient=request.user, popup_shown=False).update(popup_shown=True)
    unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    return JsonResponse({'notifications': pending, 'unread_count': unread_count})


# ─── Visitor Management ───────────────────────────────────────────────────────

def visitor_walkin(request, hostel_pk=None):
    """Public QR-scanned walk-in form — no login required."""
    all_hostels = Hostel.objects.order_by('type', 'name')
    today = dt_date.today()

    if request.method == 'POST':
        purpose_type = request.POST.get('purpose_type', 'guardian')

        # Hostel is only collected for guardian visits (from the guardian sub-section)
        if purpose_type == 'guardian':
            hostel_id  = request.POST.get('hostel_guardian', '').strip()
            hostel_obj = Hostel.objects.filter(pk=hostel_id).first() if hostel_id else None
            if not hostel_obj:
                return render(request, 'visitor/walkin.html', {
                    'all_hostels': all_hostels, 'today': today,
                    'relations': Visitor.Relation.choices,
                    'error': 'Please select the hostel the student resides in.',
                })
        else:
            hostel_obj = None  # Admission / Guest — no specific hostel

        import datetime as _dt2
        raw_end_date = request.POST.get('visit_end_date', '').strip()
        visit_end_date = None
        if raw_end_date:
            try: visit_end_date = _dt2.date.fromisoformat(raw_end_date)
            except ValueError: pass

        raw_visit_date = request.POST.get('visit_date_main', '').strip()
        visit_date_val = today
        if raw_visit_date:
            try: visit_date_val = _dt2.date.fromisoformat(raw_visit_date)
            except ValueError: pass

        visitor = Visitor(
            hostel            = hostel_obj,
            visitor_name      = request.POST.get('visitor_name', '').strip(),
            email             = request.POST.get('email', '').strip(),
            phone             = request.POST.get('phone', '').strip(),
            id_proof_type     = request.POST.get('id_proof_type', 'Aadhar').strip(),
            id_number         = request.POST.get('id_number', '').strip(),
            visit_date        = visit_date_val,
            visit_end_date    = visit_end_date,
            expected_in_time  = request.POST.get('time_in') or '09:00',
            expected_out_time = request.POST.get('time_out') or '18:00',
            purpose_type      = purpose_type,
            purpose           = request.POST.get('purpose', '').strip(),
            state             = request.POST.get('state', '').strip().title(),
            country           = request.POST.get('country', 'India').strip().title(),
            # Guardian fields
            student_name_text = request.POST.get('student_name', '').strip(),
            relation          = request.POST.get('relation', ''),
            # Guest fields
            is_college_student = request.POST.get('is_college_student') == 'yes',
        )
        if request.FILES.get('id_upload'):
            visitor.id_upload = request.FILES['id_upload']
        if request.FILES.get('college_id_upload'):
            visitor.college_id_upload = request.FILES['college_id_upload']

        # Try to link a resident student for guardian visits — try NPF ID first, then name
        if purpose_type == 'guardian':
            npf_id = request.POST.get('student_npf_id', '').strip()
            matched = None
            if npf_id:
                matched = Student.objects.filter(college_id_number=npf_id).first()
            if not matched and visitor.student_name_text:
                matched = Student.objects.filter(
                    name__iexact=visitor.student_name_text, is_resident=True
                ).first()
            if matched:
                visitor.student = matched

        visitor.save()

        # Notify warden of the hostel (only for guardian visits with a known hostel)
        if hostel_obj:
            warden = hostel_obj.warden
            if warden:
                purpose_label = dict(Visitor.PurposeType.choices).get(purpose_type, purpose_type)
                send_notification(warden,
                    f'New Visitor — {hostel_obj.name}',
                    f'{visitor.visitor_name} ({purpose_label}) arriving at {visitor.expected_in_time.strftime("%H:%M")} today.',
                    'visitor', '/visitors/')

        # Notify admins
        from apps.accounts.models import User as UserModel
        admins = list(UserModel.objects.filter(role='admin', is_active=True))
        if admins:
            hostel_label = hostel_obj.name if hostel_obj else 'General Walk-in'
            send_notification(admins,
                f'Walk-in Visitor — {hostel_label}',
                f'{visitor.visitor_name} registered via walk-in QR form.',
                'visitor', '/visitors/')

        # Notify student if guardian visit
        if purpose_type == 'guardian' and visitor.student and visitor.student.user:
            rel_label = dict(Visitor.Relation.choices).get(visitor.relation, visitor.relation)
            hostel_label = hostel_obj.name if hostel_obj else 'the hostel'
            send_notification(visitor.student.user,
                'Visitor Arriving',
                f'Your {rel_label or "visitor"} {visitor.visitor_name} is arriving at {hostel_label} today at {visitor.expected_in_time.strftime("%H:%M")}. Please acknowledge.',
                'visitor', '/my/visitors/')
            visitor.student_notified = True
            visitor.save(update_fields=['student_notified'])

        return render(request, 'visitor/walkin_success.html', {
            'visitor': visitor, 'hostel': hostel_obj,
        })

    return render(request, 'visitor/walkin.html', {
        'all_hostels': all_hostels, 'today': today,
        'relations': Visitor.Relation.choices,
    })


@admin_required
def visitor_hostel_qr(request, pk):
    """Generate and display a QR code for a hostel's walk-in form URL."""
    hostel = get_object_or_404(Hostel, pk=pk)
    import qrcode, io, base64
    base_url = request.build_absolute_uri(f'/visitor/walk-in/{hostel.pk}/')
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=12, border=4)
    qr.add_data(base_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()
    return render(request, 'visitor/hostel_qr.html', {
        'hostel': hostel, 'qr_b64': qr_b64, 'walkin_url': base_url,
    })


@admin_required
def visitor_general_qr(request):
    """Single QR code for the general walk-in form (no hostel pre-selected)."""
    import qrcode, io, base64
    # Use the actual host from the request so the URL works on any device on the network
    host = request.get_host()  # e.g. 172.16.12.237:8000
    scheme = 'https' if request.is_secure() else 'http'
    base_url = f'{scheme}://{host}/visitor/walk-in/'
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=12, border=4)
    qr.add_data(base_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()
    return render(request, 'visitor/hostel_qr.html', {
        'hostel': None, 'qr_b64': qr_b64, 'walkin_url': base_url,
    })


@login_required
def visitor_acknowledge(request, pk):
    """Student acknowledges an incoming visitor notification."""
    visitor = get_object_or_404(Visitor, pk=pk)
    student = getattr(request.user, 'student_profile', None)
    if student and visitor.student == student and not visitor.student_acknowledged:
        visitor.student_acknowledged    = True
        visitor.student_acknowledged_at = timezone.now()
        visitor.save(update_fields=['student_acknowledged', 'student_acknowledged_at'])
        messages.success(request, 'You have acknowledged the visitor.')
    return redirect('my_visitors')


@admin_required
def visitor_list(request):
    warden_hostel  = get_hostel_scope(request.user)
    hostel_filter  = request.GET.get('hostel', '')
    status_filter  = request.GET.get('status', '')
    date_filter    = request.GET.get('date', '')
    purpose_filter = request.GET.get('purpose', '')
    state_filter   = request.GET.get('state', '').strip()
    country_filter = request.GET.get('country', '').strip()

    base_qs = Visitor.objects.select_related('student', 'hostel', 'approved_by')
    if warden_hostel:
        base_qs = base_qs.filter(hostel=warden_hostel)
    elif hostel_filter:
        base_qs = base_qs.filter(hostel_id=hostel_filter)

    qs = base_qs.order_by('-created_at')
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_filter:
        qs = qs.filter(visit_date=date_filter)
    if purpose_filter:
        qs = qs.filter(purpose_type=purpose_filter)
    if state_filter:
        qs = qs.filter(state__iexact=state_filter)
    if country_filter:
        qs = qs.filter(country__iexact=country_filter)

    paginator = Paginator(qs, 25)
    page      = paginator.get_page(request.GET.get('page'))
    hostels   = Hostel.objects.all()

    status_counts = [
        (val, label, base_qs.filter(status=val).count())
        for val, label in Visitor.Status.choices
    ]
    purpose_counts = [
        (val, label, base_qs.filter(purpose_type=val).count())
        for val, label in Visitor.PurposeType.choices
    ]
    visitor_states    = Visitor.objects.exclude(state='').values_list('state', flat=True).distinct().order_by('state')
    visitor_countries = Visitor.objects.exclude(country='').values_list('country', flat=True).distinct().order_by('country')

    return render(request, 'admin/visitors.html', {
        'page':            page,
        'status_filter':   status_filter,
        'date_filter':     date_filter,
        'hostel_filter':   hostel_filter,
        'purpose_filter':  purpose_filter,
        'state_filter':    state_filter,
        'country_filter':  country_filter,
        'statuses':        Visitor.Status.choices,
        'purposes':        Visitor.PurposeType.choices,
        'status_counts':   status_counts,
        'purpose_counts':  purpose_counts,
        'hostels':         hostels,
        'today':           dt_date.today(),
        'warden_hostel':   warden_hostel,
        'is_superadmin':   request.user.role == User.Role.SUPER_ADMIN,
        'visitor_states':  visitor_states,
        'visitor_countries': visitor_countries,
    })


@admin_only
@require_POST
def visitor_action(request, pk):
    visitor = get_object_or_404(Visitor, pk=pk)
    action  = request.POST.get('action')
    if action == 'approve':
        visitor.status      = Visitor.Status.APPROVED
        visitor.approved_by = request.user
        visitor.remarks     = request.POST.get('remarks', '')
        visitor.save()
        if visitor.student and visitor.student.user:
            send_notification(visitor.student.user, 'Visitor Approved',
                f'Your visitor {visitor.visitor_name} has been approved for {visitor.visit_date}.',
                'visitor')
        messages.success(request, f'Visitor {visitor.visitor_name} approved.')
    elif action == 'reject':
        visitor.status      = Visitor.Status.REJECTED
        visitor.approved_by = request.user
        visitor.remarks     = request.POST.get('remarks', '')
        visitor.save()
        if visitor.student and visitor.student.user:
            send_notification(visitor.student.user, 'Visitor Request Rejected',
                f'Visitor {visitor.visitor_name} on {visitor.visit_date} was rejected.',
                'visitor')
        messages.warning(request, f'Visitor {visitor.visitor_name} rejected.')
    elif action == 'checkin':
        visitor.status         = Visitor.Status.CHECKED_IN
        visitor.actual_in_time = timezone.now()
        visitor.save()
        messages.success(request, f'{visitor.visitor_name} checked in at {timezone.now().strftime("%H:%M")}.')
    elif action == 'checkout':
        visitor.status          = Visitor.Status.CHECKED_OUT
        visitor.actual_out_time = timezone.now()
        visitor.save()
        messages.success(request, f'{visitor.visitor_name} checked out.')
    return redirect('visitor_list')


@login_required
def register_visitor(request):
    hostels  = Hostel.objects.all()
    students = Student.objects.filter(is_resident=True).order_by('name')
    if request.method == 'POST':
        hostel_id  = request.POST.get('hostel')
        student_id = request.POST.get('student') or None
        visitor = Visitor(
            hostel_id        = hostel_id,
            student_id       = student_id,
            visitor_name     = request.POST.get('visitor_name', '').strip(),
            relation         = request.POST.get('relation'),
            phone            = request.POST.get('phone', '').strip(),
            id_proof_type    = request.POST.get('id_proof_type', 'Aadhar').strip(),
            id_number        = request.POST.get('id_number', '').strip(),
            purpose          = request.POST.get('purpose', '').strip(),
            visit_date       = request.POST.get('visit_date'),
            expected_in_time = request.POST.get('expected_in_time') or '09:00',
            expected_out_time= request.POST.get('expected_out_time') or '18:00',
        )
        visitor.save()
        from apps.accounts.models import User as UserModel
        admins = list(UserModel.objects.filter(role__in=['admin', 'warden'], is_active=True))
        hostel_obj = Hostel.objects.filter(pk=hostel_id).first()
        send_notification(admins, 'New Visitor Request',
            f'{visitor.visitor_name} ({visitor.relation}) wants to visit {hostel_obj.name if hostel_obj else ""} on {visitor.visit_date}.',
            'visitor', '/visitors/')
        messages.success(request, 'Visitor request submitted. Admin will review shortly.')
        return redirect('register_visitor')
    return render(request, 'visitor/register.html', {
        'hostels': hostels, 'students': students,
        'relations': Visitor.Relation.choices,
        'today': dt_date.today(),
    })


@login_required
def my_visitors(request):
    student, redir = _get_student_or_redirect(request)
    if redir: return redir
    visitors = (
        Visitor.objects
        .filter(student=student)
        .select_related('hostel')
        .order_by('-created_at')
    )
    pending_ack = visitors.filter(student_notified=True, student_acknowledged=False)
    return render(request, 'student/my_visitors.html', {
        'visitors':    visitors,
        'pending_ack': pending_ack,
    })


# ─── Mess Management (unified) ───────────────────────────────────────────────

@admin_required
def mess_menu_admin(request):
    """Legacy redirect — keep old URL working."""
    return redirect('mess_management')


@admin_required
def mess_management(request):  # noqa: C901
    """Unified Mess Management: Weekly Menu, Monthly Upload, Feedback, Wastage."""
    import calendar as _cal

    is_warden    = request.user.role == User.Role.WARDEN
    is_superadmin = request.user.is_super_admin
    is_admin     = request.user.role in (User.Role.ADMIN, User.Role.SUPER_ADMIN)
    today        = dt_date.today()
    today_day    = today.strftime('%A').lower()
    month_start  = today.replace(day=1)

    warden_hostel = get_hostel_scope(request.user)
    hostels = (
        Hostel.objects.filter(pk=warden_hostel.pk)
        if warden_hostel else Hostel.objects.all().order_by('name')
    )

    tab = request.GET.get('tab', 'menu')
    selected_hostel = request.GET.get('hostel', str(warden_hostel.pk) if warden_hostel else '')

    # ── POST routing ──────────────────────────────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action', '')

        # Weekly menu: save / delete
        if action in ('save', 'delete') and not is_warden:
            if action == 'save':
                hostel_id = request.POST.get('hostel_id') or None
                day       = request.POST.get('day')
                meal_type = request.POST.get('meal_type')
                items     = request.POST.get('items', '').strip()
                obj, created = MessMenu.objects.update_or_create(
                    hostel_id=hostel_id, day=day, meal_type=meal_type,
                    defaults={'items': items, 'is_active': True, 'created_by': request.user},
                )
                if obj.is_overridden if hasattr(obj, 'is_overridden') else False:
                    obj.is_overridden = True
                    obj.updated_by = request.user
                    obj.save()
                messages.success(request, f'Menu saved for {day} {meal_type}.')
            elif action == 'delete':
                MessMenu.objects.filter(pk=request.POST.get('menu_id')).delete()
                messages.success(request, 'Menu entry deleted.')
            return redirect(f"{request.path}?tab=menu&hostel={request.POST.get('hostel_id','')}")

        # Daily menu override: save / delete
        if action == 'save_daily' and is_admin:
            hostel_id = request.POST.get('hostel_id') or None
            menu_date_str = request.POST.get('menu_date')
            meal_type     = request.POST.get('meal_type')
            items         = request.POST.get('items', '').strip()
            try:
                import datetime as _dt2
                menu_date = _dt2.date.fromisoformat(menu_date_str)
                existing  = MessDailyMenu.objects.filter(
                    hostel_id=hostel_id, menu_date=menu_date, meal_type=meal_type
                ).first()
                if existing:
                    existing.items         = items
                    existing.is_overridden = True
                    existing.updated_by    = request.user
                    existing.save()
                else:
                    MessDailyMenu.objects.create(
                        hostel_id=hostel_id, menu_date=menu_date, meal_type=meal_type,
                        items=items, original_items=items,
                        upload_month=menu_date.strftime('%Y-%m'),
                        created_by=request.user, updated_by=request.user,
                    )
                messages.success(request, f'Daily menu saved for {menu_date}.')
            except (ValueError, Exception) as e:
                messages.error(request, f'Could not save daily menu: {e}')
            return redirect(f"{request.path}?tab=menu&hostel={hostel_id or ''}")

        # Annual / Monthly Excel upload — reads ALL sheets in workbook
        if action == 'upload_excel' and is_admin:
            excel_file  = request.FILES.get('excel_file')
            hostel_id   = request.POST.get('hostel_id') or None
            force_month = request.POST.get('force_month', '').strip()

            if not excel_file:
                messages.error(request, 'No file selected.')
                return redirect(f"{request.path}?tab=upload")

            hostel_obj = Hostel.objects.filter(pk=hostel_id).first() if hostel_id else None

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
            except Exception as e:
                messages.error(request, f'Could not open file: {e}')
                return redirect(f"{request.path}?tab=upload")

            MEALS = ['breakfast', 'lunch', 'snacks', 'dinner']
            MEAL_KEYWORDS = {
                'breakfast': ['breakfast', 'morning'],
                'lunch':     ['lunch', 'noon'],
                'snacks':    ['snack', 'evening'],
                'dinner':    ['dinner', 'night'],
            }
            import datetime as _dt2

            created = updated = overridden_skipped = 0
            skipped_rows_count = 0
            upload_month_key = ''   # tracks first valid month seen across all sheets

            def _parse_sheet(ws):
                nonlocal created, updated, overridden_skipped, skipped_rows_count, upload_month_key

                # Skip non-data sheets (Instructions, etc.)
                if ws.title.strip().lower() in ('instructions', 'guide', 'readme'):
                    return

                # Detect header row (first row containing 'date' or 'day')
                header_row = None
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    joined = ' '.join(str(c).lower() for c in row if c)
                    if 'date' in joined or 'day' in joined:
                        header_row = list(row)
                        break
                if header_row is None:
                    return  # skip sheets with no recognisable header

                meal_col = {}
                for i, cell in enumerate(header_row):
                    if cell is None:
                        continue
                    h = str(cell).strip().lower()
                    for meal, keywords in MEAL_KEYWORDS.items():
                        if any(k in h for k in keywords):
                            meal_col[meal] = i
                            break

                date_col = next((i for i, c in enumerate(header_row)
                                 if c and 'date' in str(c).lower()), 0)

                data_start = 2
                for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if list(row) == header_row:
                        data_start = ridx + 1
                        break

                for row_num, row in enumerate(
                    ws.iter_rows(min_row=data_start, values_only=True), start=data_start
                ):
                    vals = list(row) + [None] * 10
                    raw_date = vals[date_col]
                    if raw_date is None:
                        continue

                    menu_date = None
                    if isinstance(raw_date, (_dt2.date, _dt2.datetime)):
                        menu_date = raw_date.date() if isinstance(raw_date, _dt2.datetime) else raw_date
                    else:
                        for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d %b %Y', '%d %B %Y'):
                            try:
                                menu_date = _dt2.datetime.strptime(str(raw_date).strip(), fmt).date()
                                break
                            except ValueError:
                                pass

                    if menu_date is None:
                        skipped_rows_count += 1
                        continue

                    month_key = menu_date.strftime('%Y-%m')
                    if not upload_month_key:
                        upload_month_key = month_key

                    for meal in MEALS:
                        if meal not in meal_col:
                            continue
                        items_raw = vals[meal_col[meal]]
                        if not items_raw:
                            continue
                        items_str = str(items_raw).strip()
                        if not items_str:
                            continue

                        existing = MessDailyMenu.objects.filter(
                            hostel=hostel_obj, menu_date=menu_date, meal_type=meal
                        ).first()

                        if existing and existing.is_overridden:
                            if force_month in (existing.upload_month, month_key):
                                existing.items          = items_str
                                existing.original_items = items_str
                                existing.is_overridden  = False
                                existing.upload_month   = month_key
                                existing.updated_by     = request.user
                                existing.save()
                                updated += 1
                            else:
                                overridden_skipped += 1
                            continue

                        if existing:
                            existing.items          = items_str
                            existing.original_items = items_str
                            existing.upload_month   = month_key
                            existing.updated_by     = request.user
                            existing.save()
                            updated += 1
                        else:
                            MessDailyMenu.objects.create(
                                hostel=hostel_obj, menu_date=menu_date, meal_type=meal,
                                items=items_str, original_items=items_str,
                                upload_month=month_key,
                                created_by=request.user, updated_by=request.user,
                            )
                            created += 1

            # Process every sheet in the workbook
            for ws in wb.worksheets:
                _parse_sheet(ws)

            redirect_url = (
                f"{request.path}?tab=upload"
                f"&u_done=1&u_created={created}&u_updated={updated}"
                f"&u_prot={overridden_skipped}&u_skip={skipped_rows_count}"
                f"&u_month={upload_month_key}"
                f"&hostel={hostel_id or ''}"
            )
            if force_month:
                redirect_url += f"&force_month={force_month}"
            return redirect(redirect_url)

        # Wastage entry (for admin quick-entry)
        if action == 'save_wastage' and is_admin:
            w_hostel_id = request.POST.get('hostel_id')
            w_hostel    = Hostel.objects.filter(pk=w_hostel_id).first() if w_hostel_id else None
            if w_hostel:
                for meal, _ in MessMenu.MealType.choices:
                    kg_raw = request.POST.get(f'wastage_{meal}', '').strip()
                    if kg_raw:
                        try:
                            kg_val = float(kg_raw)
                            if kg_val >= 0:
                                MessWastageRecord.objects.update_or_create(
                                    hostel=w_hostel, date=today, meal_type=meal,
                                    defaults={
                                        'wastage_kg': kg_val,
                                        'notes':      request.POST.get(f'notes_{meal}', '').strip(),
                                        'filled_by':  request.user,
                                    },
                                )
                        except ValueError:
                            pass
                messages.success(request, f'Wastage saved for {today.strftime("%d %b %Y")}.')
            return redirect(f"{request.path}?tab=wastage&hostel={w_hostel_id or ''}")

        return redirect(f"{request.path}?tab={tab}&hostel={selected_hostel}")

    # ── GET context ───────────────────────────────────────────────────────────

    # Tab: Weekly Menu
    menus_qs = MessMenu.objects.select_related('hostel')
    if selected_hostel:
        menus_qs = menus_qs.filter(hostel_id=selected_hostel)
    menu_grid = []
    for day, day_label in MessMenu.DayOfWeek.choices:
        day_meals = []
        for meal, meal_label in MessMenu.MealType.choices:
            dm = _get_today_daily_menu(
                Hostel.objects.filter(pk=selected_hostel).first() if selected_hostel else None
            ) if day == today_day else {}
            day_meals.append({
                'meal':       meal,
                'meal_label': meal_label,
                'menu':       menus_qs.filter(day=day, meal_type=meal).first(),
                'daily':      dm.get(meal) if day == today_day else None,
            })
        menu_grid.append({'day': day, 'day_label': day_label, 'meals': day_meals})

    # Tab: Upload result params
    upload_done    = request.GET.get('u_done') == '1'
    upload_result  = {}
    if upload_done:
        upload_result = {
            'created':           int(request.GET.get('u_created', 0)),
            'updated':           int(request.GET.get('u_updated', 0)),
            'overridden_skipped': int(request.GET.get('u_prot', 0)),
            'skipped_rows_count': int(request.GET.get('u_skip', 0)),
            'upload_month':      request.GET.get('u_month', ''),
            'force_month':       request.GET.get('force_month', ''),
        }

    # Tab: Feedback
    feedback_qs = MessFeedback.objects.select_related('student', 'hostel').order_by('-date', '-created_at')
    if selected_hostel:
        feedback_qs = feedback_qs.filter(hostel_id=selected_hostel)
    fb_meal = request.GET.get('fb_meal', '')
    if fb_meal:
        feedback_qs = feedback_qs.filter(meal_type=fb_meal)
    fb_paginator = Paginator(feedback_qs, 20)
    fb_page      = fb_paginator.get_page(request.GET.get('fbpage'))

    avg_by_meal = list(
        MessFeedback.objects.values('meal_type')
        .annotate(avg=Avg('rating'), count=Count('id'))
        .order_by('meal_type')
    )
    feedback_summary = list(
        MessFeedback.objects.values('meal_type').annotate(total=Count('id'))
    )

    # Tab: Wastage — item-level records from MessItemWastage (campus-wide, hostel=None)
    item_wastage_qs = (
        MessItemWastage.objects
        .select_related('filled_by')
        .filter(hostel__isnull=True)
        .order_by('-date', 'meal_type', 'item_name')
    )
    w_date_filter = request.GET.get('w_date', '')
    if w_date_filter:
        import datetime as _dt2
        try:
            item_wastage_qs = item_wastage_qs.filter(
                date=_dt2.date.fromisoformat(w_date_filter)
            )
        except ValueError:
            pass
    w_paginator = Paginator(item_wastage_qs, 50)
    w_page      = w_paginator.get_page(request.GET.get('wpage'))

    # Today's wastage report — campus-wide, grouped by meal for the daily view
    _today_iw = list(
        MessItemWastage.objects.filter(hostel__isnull=True, date=today)
        .select_related('filled_by')
        .order_by('meal_type', 'item_name')
    )
    today_wastage_by_meal = {}
    for _iw in _today_iw:
        today_wastage_by_meal.setdefault(_iw.meal_type, []).append(_iw)
    today_wastage_last_updated = max((_iw.updated_at for _iw in _today_iw), default=None)
    today_wastage_filled_by    = _today_iw[0].filled_by if _today_iw else None

    # Last update (campus-wide)
    last_wastage_update = (
        MessItemWastage.objects.filter(hostel__isnull=True)
        .order_by('-updated_at')
        .select_related('filled_by')
        .first()
    )

    # Month totals
    wastage_month_kg = (
        MessItemWastage.objects.filter(hostel__isnull=True, date__gte=month_start)
        .aggregate(total=Sum('wasted_qty'))['total'] or 0
    )

    today_daily = MessDailyMenu.objects.filter(menu_date=today)
    if selected_hostel:
        today_daily = today_daily.filter(
            Q(hostel_id=selected_hostel) | Q(hostel__isnull=True)
        )
    today_daily_map = {dm.meal_type: dm for dm in today_daily}

    fb_this_month = MessFeedback.objects.filter(date__gte=month_start)
    fb_count      = fb_this_month.count()
    fb_avg        = round(fb_this_month.aggregate(avg=Avg('rating'))['avg'] or 0, 1)

    return render(request, 'admin/mess_management.html', {
        'tab':               tab,
        'hostels':           hostels,
        'selected_hostel':   selected_hostel,
        'is_warden':         is_warden,
        'is_admin':          is_admin,
        'is_superadmin':     is_superadmin,
        'today':             today,
        'today_day':         today_day,
        # Weekly menu
        'menu_grid':         menu_grid,
        'meals':             MessMenu.MealType.choices,
        'today_daily_map':   today_daily_map,
        'feedback_summary':  feedback_summary,
        # Upload
        'upload_done':       upload_done,
        'upload_result':     upload_result,
        # Feedback
        'fb_page':           fb_page,
        'fb_meal':           fb_meal,
        'avg_by_meal':       avg_by_meal,
        'fb_count':          fb_count,
        'fb_avg':            fb_avg,
        # Wastage (item-level)
        'w_page':                    w_page,
        'w_date_filter':             w_date_filter,
        'wastage_month_kg':          wastage_month_kg,
        'today_wastage_by_meal':     today_wastage_by_meal,
        'today_wastage_last_updated': today_wastage_last_updated,
        'today_wastage_filled_by':   today_wastage_filled_by,
        'last_wastage_update':       last_wastage_update,
        'today_daily':               today_daily,
    })


@login_required
def mess_menu_student(request):
    student, redir = _get_student_or_redirect(request)
    if redir: return redir
    hostel = None
    alloc  = RoomAllocation.objects.filter(student=student, status='active').select_related('room__hostel').first()
    if alloc:
        hostel = alloc.room.hostel

    menus_qs = MessMenu.objects.filter(is_active=True)
    if hostel:
        menus_qs = menus_qs.filter(Q(hostel=hostel) | Q(hostel__isnull=True))

    # Build as list-of-lists so templates need no custom filter
    menu_grid = []
    for day, day_label in MessMenu.DayOfWeek.choices:
        day_meals = []
        for meal, meal_label in MessMenu.MealType.choices:
            day_meals.append({
                'meal': meal,
                'meal_label': meal_label,
                'menu': menus_qs.filter(day=day, meal_type=meal).first(),
            })
        menu_grid.append({'day': day, 'day_label': day_label, 'meals': day_meals})

    if request.method == 'POST':
        today     = dt_date.today()
        meal_type = request.POST.get('meal_type')
        rating    = request.POST.get('rating')
        comment   = request.POST.get('comment', '').strip()
        if hostel and meal_type and rating:
            MessFeedback.objects.update_or_create(
                student=student, date=today, meal_type=meal_type,
                defaults={'rating': int(rating), 'comment': comment, 'hostel': hostel}
            )
            messages.success(request, 'Feedback submitted, thank you!')
        return redirect('mess_menu_student')

    today_feedback = MessFeedback.objects.filter(student=student, date=dt_date.today())
    already_rated  = list(today_feedback.values_list('meal_type', flat=True))

    return render(request, 'student/mess_menu.html', {
        'menu_grid': menu_grid, 'hostel': hostel,
        'meals': MessMenu.MealType.choices, 'already_rated': already_rated,
    })


# ─── Mess — Bulk Upload (Monthly Menu) ───────────────────────────────────────

def _get_today_daily_menu(hostel=None):
    """Return {meal_type: MessDailyMenu} for today, optionally filtered by hostel."""
    today = dt_date.today()
    qs = MessDailyMenu.objects.filter(menu_date=today)
    if hostel:
        qs = qs.filter(Q(hostel=hostel) | Q(hostel__isnull=True))
    return {m.meal_type: m for m in qs}


@admin_only
def mess_bulk_upload(request):  # noqa: C901
    """Legacy redirect — now handled inside mess_management."""
    return redirect(f"{reverse('mess_management')}?tab=upload")


@admin_only
def _mess_bulk_upload_legacy(request):  # noqa: C901
    """Upload a monthly mess menu Excel. One sheet, rows = days, cols = meals."""
    import calendar, datetime as _dt

    hostels = Hostel.objects.all().order_by('name')
    done = False
    result = {}

    if request.method == 'POST':
        excel_file  = request.FILES.get('excel_file')
        hostel_id   = request.POST.get('hostel_id') or None
        force_month = request.POST.get('force_month', '').strip()   # "YYYY-MM" to force-override

        if not excel_file:
            messages.error(request, 'No file selected.')
            return redirect('mess_bulk_upload')

        hostel_obj = None
        if hostel_id:
            hostel_obj = Hostel.objects.filter(pk=hostel_id).first()

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            messages.error(request, f'Could not open file: {e}')
            return redirect('mess_bulk_upload')

        MEALS = ['breakfast', 'lunch', 'snacks', 'dinner']
        MEAL_KEYWORDS = {
            'breakfast': ['breakfast', 'morning'],
            'lunch':     ['lunch', 'noon'],
            'snacks':    ['snack', 'evening'],
            'dinner':    ['dinner', 'night'],
        }

        # Detect header row (row with "date" keyword)
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            joined = ' '.join(str(c).lower() for c in row if c)
            if 'date' in joined or 'day' in joined:
                header_row = list(row)
                break

        if header_row is None:
            messages.error(request, 'Could not detect header row. Use the sample template.')
            return redirect('mess_bulk_upload')

        # Map column indices to meal types
        meal_col = {}
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            h = str(cell).strip().lower()
            for meal, keywords in MEAL_KEYWORDS.items():
                if any(k in h for k in keywords):
                    meal_col[meal] = i
                    break

        if not meal_col:
            messages.error(request, 'No meal columns found. Expected Breakfast / Lunch / Snacks / Dinner headers.')
            return redirect('mess_bulk_upload')

        # Detect date column (first col with 'date' or index 0)
        date_col = next((i for i, c in enumerate(header_row)
                         if c and 'date' in str(c).lower()), 0)

        created = updated = overridden_skipped = 0
        skipped_rows = []
        upload_month_key = ''

        data_start = None
        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if list(row) == header_row:
                data_start = ridx + 1
                break
        if data_start is None:
            data_start = 2  # fallback

        for row_num, row in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), start=data_start
        ):
            vals = list(row) + [None] * 10
            raw_date = vals[date_col]
            if raw_date is None:
                continue

            # Parse date
            menu_date = None
            import datetime as _dt2
            if isinstance(raw_date, (_dt2.date, _dt2.datetime)):
                menu_date = raw_date.date() if isinstance(raw_date, _dt2.datetime) else raw_date
            else:
                for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d %b %Y', '%d %B %Y'):
                    try:
                        menu_date = _dt2.datetime.strptime(str(raw_date).strip(), fmt).date()
                        break
                    except ValueError:
                        pass

            if menu_date is None:
                skipped_rows.append({'row': row_num, 'data': str(raw_date), 'reason': 'Unrecognised date format'})
                continue

            if not upload_month_key:
                upload_month_key = menu_date.strftime('%Y-%m')

            for meal in MEALS:
                if meal not in meal_col:
                    continue
                items_raw = vals[meal_col[meal]]
                if not items_raw:
                    continue
                items_str = str(items_raw).strip()
                if not items_str:
                    continue

                existing = MessDailyMenu.objects.filter(
                    hostel=hostel_obj, menu_date=menu_date, meal_type=meal
                ).first()

                if existing and existing.is_overridden:
                    if force_month == existing.upload_month or force_month == menu_date.strftime('%Y-%m'):
                        existing.items          = items_str
                        existing.original_items = items_str
                        existing.is_overridden  = False
                        existing.upload_month   = upload_month_key
                        existing.updated_by     = request.user
                        existing.save()
                        updated += 1
                    else:
                        overridden_skipped += 1
                    continue

                if existing:
                    existing.items          = items_str
                    existing.original_items = items_str
                    existing.upload_month   = upload_month_key
                    existing.updated_by     = request.user
                    existing.save()
                    updated += 1
                else:
                    MessDailyMenu.objects.create(
                        hostel=hostel_obj, menu_date=menu_date, meal_type=meal,
                        items=items_str, original_items=items_str,
                        upload_month=upload_month_key,
                        created_by=request.user, updated_by=request.user,
                    )
                    created += 1

        done = True
        result = {
            'created': created, 'updated': updated,
            'overridden_skipped': overridden_skipped,
            'skipped_rows': skipped_rows,
            'upload_month': upload_month_key,
            'hostel': hostel_obj,
        }

    return render(request, 'admin/mess_bulk_upload.html', {
        'done': done, 'hostels': hostels, **result,
    })


@admin_only
def mess_bulk_sample(request):
    """Download annual mess-menu Excel template — 12 monthly sheets + Instructions."""
    import calendar, datetime as _dt

    today = _dt.date.today()
    year  = today.year

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
    def hfill(h): return PatternFill('solid', fgColor=h)
    cc = Alignment(horizontal='center', vertical='center', wrap_text=True)
    DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    sample_menus = [
        'Idli (2 pcs), Sambar, Coconut Chutney, Tea / Coffee',
        'Rice, Dal Tadka, Aloo Gobi, Chapati (2), Salad, Buttermilk',
        'Bread Pakora, Tomato Sauce, Tea',
        'Chapati (3), Paneer Butter Masala, Rice, Dal, Raita',
    ]
    meal_headers = [('Date*', 'F59E0B'), ('Day', 'D97706'),
                    ('Breakfast', '1565C0'), ('Lunch', '1565C0'),
                    ('Evening Snacks', '1976D2'), ('Dinner', '0D47A1')]

    wb = openpyxl.Workbook()
    # Remove default sheet; we'll create one per month
    wb.remove(wb.active)

    for month_num in range(1, 13):
        month_name = MONTHS[month_num - 1]
        sheet_title = f'{month_name[:3]} {year}'
        num_days    = calendar.monthrange(year, month_num)[1]

        ws = wb.create_sheet(title=sheet_title)

        for col_ltr, w in zip('ABCDEF', [13, 11, 32, 32, 32, 32]):
            ws.column_dimensions[col_ltr].width = w

        # Row 1 — title banner
        tc = ws.cell(row=1, column=1,
                     value=f'  GGI Hostel ERP  —  Mess Menu  —  {month_name} {year}')
        tc.font      = Font(bold=True, color='FFFFFF', size=13)
        tc.fill      = hfill('0D47A1')
        tc.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 26

        # Row 2 — group labels
        for sc, ec, lbl, clr in [
            (1, 2, 'DATE INFO', '0F4C81'),
            (3, 6, 'MEAL ITEMS — enter each dish comma-separated or Alt+Enter within the cell', '1B5E20'),
        ]:
            c2 = ws.cell(row=2, column=sc, value=lbl)
            c2.font = Font(bold=True, color='FFFFFF', size=9)
            c2.fill = hfill(clr); c2.alignment = cc; c2.border = thin
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        ws.row_dimensions[2].height = 28

        # Row 3 — column headers
        for col_idx, (hdr, clr) in enumerate(meal_headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.font      = Font(bold=True, color='FFFFFF', size=10)
            cell.fill      = hfill(clr)
            cell.alignment = cc
            cell.border    = thin
        ws.row_dimensions[3].height = 22

        # One row per day
        for day_num in range(1, num_days + 1):
            row_num  = day_num + 3
            date_obj = _dt.date(year, month_num, day_num)
            day_name = DAYS[date_obj.weekday()]
            weekend  = date_obj.weekday() >= 5
            bg       = hfill('EBF3FF') if day_num % 2 == 0 else hfill('F8FAFF')
            wk_fill  = hfill('FFF8E1')

            dc = ws.cell(row=row_num, column=1, value=date_obj)
            dc.number_format = 'DD-MM-YYYY'
            dc.alignment = cc; dc.border = thin
            dc.fill = wk_fill if weekend else bg

            dy = ws.cell(row=row_num, column=2, value=day_name)
            dy.alignment = cc; dy.border = thin
            dy.fill = wk_fill if weekend else bg

            # Only first 2 days of first month get sample data
            for col_idx, sample in enumerate(sample_menus, 3):
                val = sample if (month_num == 1 and day_num <= 2) else ''
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = Alignment(vertical='top', wrap_text=True, indent=1)
                cell.border    = thin
                cell.fill      = wk_fill if weekend else bg

            ws.row_dimensions[row_num].height = 40

        ws.freeze_panes = 'C4'

    # Instructions sheet
    wi = wb.create_sheet('Instructions')
    wi.column_dimensions['A'].width = 28
    wi.column_dimensions['B'].width = 65

    title_cell = wi.cell(row=1, column=1,
                         value='GGI Hostel ERP — Annual Mess Menu Upload Guide')
    title_cell.font = Font(bold=True, color='FFFFFF', size=12)
    title_cell.fill = hfill('0D47A1')
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    wi.merge_cells('A1:B1')
    wi.row_dimensions[1].height = 26

    guide_rows = [
        ('Workbook structure',  'One sheet per month (Jan–Dec). Each sheet has dates for that month.'),
        ('Date* (Col A)',       'Required. Use a date cell (format DD-MM-YYYY). System auto-detects.'),
        ('Day (Col B)',         'Optional — for reference only. System ignores during import.'),
        ('Breakfast (Col C)',   'All breakfast items, comma-separated or Alt+Enter within cell.'),
        ('Lunch (Col D)',       'All lunch items.'),
        ('Evening Snacks (E)',  'Snack / tea-time items.'),
        ('Dinner (Col F)',      'All dinner items.'),
        ('How upload works',    'Upload once for the full year. System reads all sheets. '
                                'Entries are stored date-wise and auto-display on the correct date.'),
        ('Monthly auto-switch', 'When a month ends, the next month\'s menu appears automatically. '
                                'No re-upload needed unless the menu changes.'),
        ('Editing an entry',    'You can manually edit any date\'s entry from the Mess Management page. '
                                'That entry is marked "protected" and won\'t be overwritten by re-upload.'),
        ('Force replace',       'Use "Force Replace" during upload to overwrite protected (edited) entries '
                                'for a specific month.'),
    ]
    for r, (label, desc) in enumerate(guide_rows, start=3):
        bg = 'EBF3FF' if r % 2 == 0 else 'F8FAFF'
        for c, val in [(1, label), (2, desc)]:
            cell = wi.cell(row=r, column=c, value=val)
            cell.font      = Font(size=10)
            cell.fill      = hfill(bg)
            cell.alignment = Alignment(vertical='center', indent=1, wrap_text=True)
            cell.border    = thin
        wi.row_dimensions[r].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="mess_annual_menu_{year}.xlsx"'
    return resp


# ─── Mess Incharge Dashboard ──────────────────────────────────────────────────

@mess_only
def mess_incharge_dashboard(request):  # noqa: C901
    """Mess incharge — campus-wide daily wastage entry against today's menu."""
    import re as _re
    from datetime import timedelta
    from decimal import Decimal, InvalidOperation

    today     = dt_date.today()
    today_day = today.strftime('%A').lower()

    # ── POST: save campus-wide item-level wastage ─────────────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'save_item_wastage':
        # Delete today's campus-wide records and re-create from form
        MessItemWastage.objects.filter(hostel__isnull=True, date=today).delete()
        for meal_code, _ in MessMenu.MealType.choices:
            count_str = request.POST.get(f'item_count_{meal_code}', '0')
            count = int(count_str) if count_str.isdigit() else 0
            for i in range(count):
                item_name = request.POST.get(f'item_name_{meal_code}_{i}', '').strip()
                if not item_name:
                    continue
                unit = request.POST.get(f'unit_{meal_code}_{i}', 'kg').strip()
                try:
                    p = Decimal(request.POST.get(f'prepared_{meal_code}_{i}', '0') or '0')
                    c = Decimal(request.POST.get(f'consumed_{meal_code}_{i}', '0') or '0')
                    w_raw = request.POST.get(f'wasted_{meal_code}_{i}', '').strip()
                    w = Decimal(w_raw) if w_raw else max(p - c, Decimal('0'))
                except InvalidOperation:
                    continue
                MessItemWastage.objects.create(
                    hostel=None, date=today, meal_type=meal_code,
                    item_name=item_name, unit=unit,
                    prepared_qty=p, consumed_qty=c, wasted_qty=w,
                    filled_by=request.user,
                )
        messages.success(request, f'Wastage report saved for {today.strftime("%d %b %Y")}.')
        return redirect(request.path)

    # ── Today's menu — any source, first match per meal ───────────────────────
    # Sort null-hostel (campus-wide) records first so they take priority.
    daily_menus = {}
    _dm_list = sorted(
        MessDailyMenu.objects.filter(menu_date=today),
        key=lambda dm: (dm.hostel_id is not None)  # None=0 sorts before non-null=1
    )
    for dm in _dm_list:
        if dm.meal_type not in daily_menus:
            daily_menus[dm.meal_type] = dm

    weekly_menus = {}
    for wm in MessMenu.objects.filter(day=today_day, is_active=True):
        if wm.meal_type not in weekly_menus:
            weekly_menus[wm.meal_type] = wm

    # ── Existing campus-wide wastage for today ────────────────────────────────
    existing_items = {}
    for iw in MessItemWastage.objects.filter(hostel__isnull=True, date=today):
        existing_items.setdefault(iw.meal_type, []).append(iw)

    # ── Build meal grid ───────────────────────────────────────────────────────
    meal_grid = []
    meals_filled = 0
    for meal_code, meal_label in MessMenu.MealType.choices:
        dm = daily_menus.get(meal_code)
        wm = weekly_menus.get(meal_code)
        menu_obj = dm or wm
        parsed = []
        if menu_obj and menu_obj.items:
            parsed = [it.strip() for it in _re.split(r'[,\n]+', menu_obj.items) if it.strip()]

        saved = existing_items.get(meal_code, [])
        if saved:
            meals_filled += 1
            rows = [{'name': iw.item_name, 'unit': iw.unit,
                     'prepared': str(iw.prepared_qty),
                     'consumed': str(iw.consumed_qty),
                     'wasted':   str(iw.wasted_qty)} for iw in saved]
        else:
            rows = [{'name': it, 'unit': 'kg', 'prepared': '', 'consumed': '', 'wasted': ''}
                    for it in parsed]

        meal_grid.append({
            'meal': meal_code, 'label': meal_label,
            'menu': menu_obj,
            'source': 'daily' if dm else ('weekly' if wm else None),
            'rows':  rows,
            'filled': bool(saved),
        })

    # ── Last 7-day wastage history (campus-wide) ──────────────────────────────
    week_ago = today - timedelta(days=6)
    hist_qs  = (
        MessItemWastage.objects
        .filter(hostel__isnull=True, date__gte=week_ago)
        .select_related('filled_by')
        .order_by('-date', 'meal_type', 'item_name')
    )
    history_grouped = {}
    for iw in hist_qs:
        key = (iw.date, iw.meal_type)
        history_grouped.setdefault(key, []).append(iw)
    history_list = sorted(
        [{'date': k[0], 'meal_type': k[1], 'items': v,
          'total_prepared': sum(float(i.prepared_qty) for i in v),
          'total_consumed': sum(float(i.consumed_qty) for i in v),
          'total_wasted':   sum(float(i.wasted_qty)   for i in v),
          'updated_at':     max(i.updated_at for i in v),
          'filled_by':      v[0].filled_by}
         for k, v in history_grouped.items()],
        key=lambda x: (x['date'], x['meal_type']), reverse=True
    )

    # Last-updated timestamp across all today's records
    last_updated = None
    today_iw = list(MessItemWastage.objects.filter(hostel__isnull=True, date=today))
    if today_iw:
        last_updated = max(iw.updated_at for iw in today_iw)

    return render(request, 'staff/mess_incharge.html', {
        'today':           today,
        'meal_grid':       meal_grid,
        'meals_filled':    meals_filled,
        'total_meals':     len(meal_grid),
        'history_list':    history_list,
        'last_updated':    last_updated,
        'unit_choices':    MessItemWastage.UNIT_CHOICES,
        'meal_items_json': json.dumps({
            m['meal']: m['rows'] for m in meal_grid
        }),
    })


# ─── SuperAdmin — Mess Overview ───────────────────────────────────────────────

@superadmin_only
def superadmin_mess(request):
    """Legacy redirect — now unified in mess_management."""
    return redirect(f"{reverse('mess_management')}?tab=wastage")


@superadmin_only
def _superadmin_mess_legacy(request):  # noqa: C901
    """SuperAdmin: mess menu management, feedback review, wastage monitoring."""
    from datetime import timedelta
    today      = dt_date.today()
    month_start = today.replace(day=1)

    tab = request.GET.get('tab', 'overview')
    hostel_id = request.GET.get('hostel', '')
    hostel_obj = Hostel.objects.filter(pk=hostel_id).first() if hostel_id else None
    hostels    = Hostel.objects.all().order_by('name')

    # ── Overview stats ──
    fb_this_month = MessFeedback.objects.filter(date__gte=month_start)
    if hostel_obj:
        fb_this_month = fb_this_month.filter(hostel=hostel_obj)
    fb_count   = fb_this_month.count()
    fb_avg     = fb_this_month.aggregate(avg=Avg('rating'))['avg'] or 0

    wastage_this_month = MessWastageRecord.objects.filter(date__gte=month_start)
    if hostel_obj:
        wastage_this_month = wastage_this_month.filter(hostel=hostel_obj)
    total_wastage_kg = wastage_this_month.aggregate(total=Sum('wastage_kg'))['total'] or 0

    # ── Feedback list (tab=feedback) ──
    feedback_qs = (
        MessFeedback.objects
        .select_related('student', 'hostel')
        .order_by('-date', '-created_at')
    )
    if hostel_obj:
        feedback_qs = feedback_qs.filter(hostel=hostel_obj)
    date_from = request.GET.get('date_from')
    date_to   = request.GET.get('date_to')
    if date_from:
        try:
            feedback_qs = feedback_qs.filter(date__gte=dt_date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            feedback_qs = feedback_qs.filter(date__lte=dt_date.fromisoformat(date_to))
        except ValueError:
            pass
    meal_f = request.GET.get('meal', '')
    if meal_f:
        feedback_qs = feedback_qs.filter(meal_type=meal_f)
    fb_paginator = Paginator(feedback_qs, 25)
    fb_page      = fb_paginator.get_page(request.GET.get('page'))

    # ── Wastage list (tab=wastage) ──
    wastage_qs = (
        MessWastageRecord.objects
        .select_related('hostel', 'filled_by')
        .order_by('-date', 'meal_type')
    )
    if hostel_obj:
        wastage_qs = wastage_qs.filter(hostel=hostel_obj)
    wdate_from = request.GET.get('wdate_from')
    wdate_to   = request.GET.get('wdate_to')
    if wdate_from:
        try:
            wastage_qs = wastage_qs.filter(date__gte=dt_date.fromisoformat(wdate_from))
        except ValueError:
            pass
    if wdate_to:
        try:
            wastage_qs = wastage_qs.filter(date__lte=dt_date.fromisoformat(wdate_to))
        except ValueError:
            pass
    w_paginator = Paginator(wastage_qs, 30)
    w_page      = w_paginator.get_page(request.GET.get('wpage'))

    # ── Today's daily menu preview ──
    today_daily = MessDailyMenu.objects.filter(menu_date=today)
    if hostel_obj:
        today_daily = today_daily.filter(Q(hostel=hostel_obj) | Q(hostel__isnull=True))

    # ── Last wastage update time per hostel ──
    last_wastage = (
        MessWastageRecord.objects
        .values('hostel__name', 'hostel_id')
        .annotate(last_updated=Max('updated_at'))
        .order_by('-last_updated')[:10]
    )

    return render(request, 'admin/superadmin_mess.html', {
        'tab': tab, 'hostels': hostels, 'hostel': hostel_obj,
        'fb_count': fb_count, 'fb_avg': round(fb_avg, 1),
        'total_wastage_kg': total_wastage_kg,
        'fb_page': fb_page, 'w_page': w_page,
        'meals': MessMenu.MealType.choices,
        'today_daily': today_daily,
        'last_wastage': last_wastage,
        'today': today,
    })


# ─── Mess Feedback — Admin detail list ───────────────────────────────────────

@admin_only
def mess_feedback_list(request):
    """Legacy redirect — now inside mess_management."""
    return redirect(f"{reverse('mess_management')}?tab=feedback")


@admin_only
def _mess_feedback_list_legacy(request):
    """Admin view: paginated list of all student mess feedback."""
    qs = (
        MessFeedback.objects
        .select_related('student', 'hostel')
        .order_by('-date', '-created_at')
    )
    hostel_id = request.GET.get('hostel', '')
    meal_f    = request.GET.get('meal', '')
    rating_f  = request.GET.get('rating', '')
    if hostel_id:
        qs = qs.filter(hostel_id=hostel_id)
    if meal_f:
        qs = qs.filter(meal_type=meal_f)
    if rating_f:
        try:
            qs = qs.filter(rating=int(rating_f))
        except ValueError:
            pass

    hostels   = Hostel.objects.all().order_by('name')
    paginator = Paginator(qs, 30)
    page      = paginator.get_page(request.GET.get('page'))

    avg_by_meal = list(
        MessFeedback.objects.values('meal_type')
        .annotate(avg=Avg('rating'), count=Count('id'))
        .order_by('meal_type')
    )

    return render(request, 'admin/mess_feedback_list.html', {
        'page': page, 'hostels': hostels,
        'meals': MessMenu.MealType.choices,
        'avg_by_meal': avg_by_meal,
        'selected_hostel': hostel_id,
        'selected_meal':   meal_f,
        'selected_rating': rating_f,
    })


# ─── Discipline Module ────────────────────────────────────────────────────────

def _discipline_student_qs():
    return Student.objects.filter(is_resident=True).prefetch_related(
        Prefetch('allocations',
                 queryset=RoomAllocation.objects.filter(status='active').select_related('room__hostel'),
                 to_attr='active_allocs')
    ).order_by('name')


@login_required
def discipline_category_list(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        name   = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Category name is required.')
            return redirect('discipline_category_list')
        pk = request.POST.get('pk')
        cat = get_object_or_404(DisciplineCategory, pk=pk) if (action == 'edit' and pk) else DisciplineCategory(created_by=request.user)
        if action == 'delete' and pk:
            get_object_or_404(DisciplineCategory, pk=pk).delete()
            messages.success(request, 'Category deleted.')
            return redirect('discipline_category_list')
        cat.name               = name
        cat.description        = request.POST.get('description', '').strip()
        cat.severity           = request.POST.get('severity', 'medium')
        cat.default_warning    = request.POST.get('default_warning', 'verbal')
        cat.fine_mandatory     = 'fine_mandatory' in request.POST
        cat.notify_parent_auto = 'notify_parent_auto' in request.POST
        cat.is_active          = 'is_active' in request.POST
        cat.save()
        messages.success(request, f'Category "{cat.name}" saved.')
        return redirect('discipline_category_list')

    cats = DisciplineCategory.objects.all()
    return render(request, 'admin/discipline_categories.html', {
        'categories': cats,
        'severity_choices': DisciplineCategory.Severity.choices,
        'warning_choices': DisciplineRecord.WarningType.choices,
    })


@admin_only
def discipline_list(request):
    status_filter = request.GET.get('status', '')
    hostel_filter = request.GET.get('hostel', '')
    q = request.GET.get('q', '').strip()

    qs = DisciplineRecord.objects.select_related(
        'student', 'category', 'recorded_by'
    ).order_by('-created_at')

    if status_filter:
        qs = qs.filter(status=status_filter)
    if hostel_filter:
        qs = qs.filter(hostel_name__icontains=hostel_filter)
    if q:
        qs = qs.filter(
            Q(student__name__icontains=q) | Q(student__roll_number__icontains=q)
            | Q(description__icontains=q) | Q(category__name__icontains=q)
        )

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))

    counts = {
        'all':                DisciplineRecord.objects.count(),
        'pending':            DisciplineRecord.objects.filter(status='pending').count(),
        'under_investigation':DisciplineRecord.objects.filter(status='under_investigation').count(),
        'resolved':           DisciplineRecord.objects.filter(status='resolved').count(),
        'closed':             DisciplineRecord.objects.filter(status='closed').count(),
        'unpaid_fines':       DisciplineRecord.objects.filter(fine_amount__gt=0, fine_paid=False).count(),
    }
    hostels = Hostel.objects.values_list('name', flat=True)
    return render(request, 'admin/discipline.html', {
        'page': page, 'status_filter': status_filter,
        'hostel_filter': hostel_filter, 'q': q,
        'counts': counts, 'hostels': hostels,
        'statuses': DisciplineRecord.Status.choices,
    })


@admin_only
def discipline_add(request):
    if request.method == 'POST':
        student_id   = request.POST.get('student_id')
        category_id  = request.POST.get('category_id') or None
        description  = request.POST.get('description', '').strip()
        incident_date = request.POST.get('incident_date')
        incident_time = request.POST.get('incident_time') or None
        warning_type  = request.POST.get('warning_type', 'none')
        fine_amount   = request.POST.get('fine_amount', 0) or 0
        fine_due_date = request.POST.get('fine_due_date') or None
        notify_parent = 'notify_parent' in request.POST

        if not student_id or not category_id or not description or not incident_date:
            messages.error(request, 'Student, category, incident date, and description are required.')
            return redirect('discipline_add')

        student_obj = get_object_or_404(Student, pk=student_id)

        # Snapshot hostel/room from active allocation
        alloc = RoomAllocation.objects.filter(student=student_obj, status='active').select_related('room__hostel').first()
        rec = DisciplineRecord.objects.create(
            student       = student_obj,
            category_id   = category_id,
            description   = description,
            incident_date = incident_date,
            incident_time = incident_time,
            warning_type  = warning_type,
            fine_amount   = fine_amount,
            fine_due_date = fine_due_date,
            notify_parent = notify_parent,
            hostel_name   = alloc.room.hostel.name if alloc else '',
            room_number   = alloc.room.room_number if alloc else '',
            floor         = alloc.room.floor if alloc else '',
            recorded_by   = request.user,
        )

        # Audit log
        DisciplineAction.objects.create(
            record=rec, action='created',
            description=f'Record created by {request.user.name}. Category: {rec.category}. Warning: {warning_type}.',
            performed_by=request.user,
        )

        # Upload evidence
        for f in request.FILES.getlist('evidence'):
            ext = f.name.rsplit('.', 1)[-1].lower()
            ftype = 'image' if ext in ('jpg','jpeg','png','gif','webp') else 'document'
            ev = DisciplineEvidence(record=rec, file_type=ftype,
                                     original_name=f.name, uploaded_by=request.user)
            ev.file.save(f.name, f, save=True)
            DisciplineAction.objects.create(
                record=rec, action='evidence_added',
                description=f'Evidence uploaded: {f.name}',
                performed_by=request.user,
            )

        # Notify student
        if student_obj.user:
            send_notification(
                [student_obj.user], 'Discipline Notice',
                f'A discipline record ({rec.category}) has been filed against you for {incident_date}.',
                notif_type='system',
            )

        # Parent notification flag
        if notify_parent:
            rec.parent_notified_by = request.user
            rec.parent_notified_at = timezone.now()
            rec.save()
            DisciplineAction.objects.create(
                record=rec, action='parent_notified',
                description=f'Parent notification recorded by {request.user.name}.',
                performed_by=request.user,
            )

        messages.success(request, f'Discipline record filed for {student_obj.name}.')
        return redirect('discipline_detail', pk=rec.pk)

    students   = _discipline_student_qs()
    categories = DisciplineCategory.objects.filter(is_active=True)
    return render(request, 'admin/discipline_add.html', {
        'students':       students,
        'categories':     categories,
        'warning_choices': DisciplineRecord.WarningType.choices,
        'today':          timezone.now().date(),
    })


@admin_only
def discipline_detail(request, pk):
    rec = get_object_or_404(
        DisciplineRecord.objects.select_related(
            'student', 'category', 'recorded_by', 'resolved_by',
            'parent_notified_by', 'fine_paid_by'
        ).prefetch_related('evidence', 'actions__performed_by'),
        pk=pk
    )
    return render(request, 'admin/discipline_detail.html', {
        'rec': rec,
        'statuses': DisciplineRecord.Status.choices,
        'warning_choices': DisciplineRecord.WarningType.choices,
    })


@admin_only
@require_POST
def discipline_action_view(request, pk):
    rec    = get_object_or_404(DisciplineRecord, pk=pk)
    action = request.POST.get('action')

    if action == 'change_status':
        new_status = request.POST.get('status')
        note       = request.POST.get('note', '').strip()
        old_status = rec.get_status_display()
        rec.status = new_status
        if new_status in ('resolved', 'closed'):
            rec.resolution_notes = request.POST.get('resolution_notes', '').strip()
            rec.resolved_by  = request.user
            rec.resolved_at  = timezone.now()
        rec.save()
        DisciplineAction.objects.create(
            record=rec, action='status_changed',
            description=f'Status changed from {old_status} to {rec.get_status_display()}.' + (f' Note: {note}' if note else ''),
            performed_by=request.user,
        )
        messages.success(request, f'Status updated to {rec.get_status_display()}.')

    elif action == 'update_warning':
        rec.warning_type = request.POST.get('warning_type', 'none')
        rec.save()
        DisciplineAction.objects.create(
            record=rec, action='warning_issued',
            description=f'Warning updated to: {rec.get_warning_type_display()}.',
            performed_by=request.user,
        )
        messages.success(request, 'Warning updated.')

    elif action == 'update_fine':
        rec.fine_amount   = request.POST.get('fine_amount', 0) or 0
        rec.fine_due_date = request.POST.get('fine_due_date') or None
        rec.save()
        DisciplineAction.objects.create(
            record=rec, action='fine_added',
            description=f'Fine set to ₹{rec.fine_amount}, due {rec.fine_due_date or "N/A"}.',
            performed_by=request.user,
        )
        messages.success(request, 'Fine updated.')

    elif action == 'mark_paid':
        rec.fine_paid      = True
        rec.fine_paid_date = timezone.now().date()
        rec.fine_paid_by   = request.user
        rec.save()
        DisciplineAction.objects.create(
            record=rec, action='fine_paid',
            description=f'Fine of ₹{rec.fine_amount} marked as paid by {request.user.name}.',
            performed_by=request.user,
        )
        messages.success(request, 'Fine marked as paid.')

    elif action == 'add_note':
        note = request.POST.get('note', '').strip()
        if note:
            DisciplineAction.objects.create(
                record=rec, action='note_added',
                description=note, performed_by=request.user,
            )
            messages.success(request, 'Note added.')

    elif action == 'notify_parent':
        rec.notify_parent      = True
        rec.parent_notified_by = request.user
        rec.parent_notified_at = timezone.now()
        rec.save()
        DisciplineAction.objects.create(
            record=rec, action='parent_notified',
            description=f'Parent notified by {request.user.name}. Note: {request.POST.get("note","")}',
            performed_by=request.user,
        )
        messages.success(request, 'Parent notification recorded.')

    elif action == 'apply_restriction':
        rtype  = request.POST.get('restriction_type')
        reason = request.POST.get('reason', '').strip()
        if rtype and reason:
            StudentRestriction.objects.create(
                student=rec.student, restriction_type=rtype,
                reason=f'[Discipline: {rec.category}] {reason}',
                restricted_by=request.user,
            )
            DisciplineAction.objects.create(
                record=rec, action='restriction_applied',
                description=f'{dict(StudentRestriction.RestrictionType.choices).get(rtype)} restriction applied. Reason: {reason}',
                performed_by=request.user,
            )
            if rec.student.user:
                send_notification(
                    [rec.student.user],
                    'Restriction Applied',
                    f'Due to discipline action, your {rtype.replace("_"," ")} access has been restricted.',
                    notif_type='system',
                )
            messages.success(request, 'Restriction applied.')

    elif action == 'upload_evidence':
        for f in request.FILES.getlist('evidence'):
            ext   = f.name.rsplit('.', 1)[-1].lower()
            ftype = 'image' if ext in ('jpg','jpeg','png','gif','webp') else 'document'
            ev = DisciplineEvidence(record=rec, file_type=ftype,
                                     original_name=f.name, uploaded_by=request.user)
            ev.file.save(f.name, f, save=True)
        DisciplineAction.objects.create(
            record=rec, action='evidence_added',
            description=f'{len(request.FILES.getlist("evidence"))} evidence file(s) uploaded.',
            performed_by=request.user,
        )
        messages.success(request, 'Evidence uploaded.')

    return redirect('discipline_detail', pk=pk)


@admin_only
def discipline_report(request):
    from django.db.models.functions import TruncMonth
    total    = DisciplineRecord.objects.count()
    active   = DisciplineRecord.objects.filter(status__in=['pending','under_investigation']).count()
    resolved = DisciplineRecord.objects.filter(status__in=['resolved','closed']).count()
    unpaid   = DisciplineRecord.objects.filter(fine_amount__gt=0, fine_paid=False).count()
    total_fines   = DisciplineRecord.objects.aggregate(t=Sum('fine_amount'))['t'] or 0
    collected     = DisciplineRecord.objects.filter(fine_paid=True).aggregate(t=Sum('fine_amount'))['t'] or 0
    pending_fines = total_fines - collected

    # Category breakdown
    by_category = DisciplineCategory.objects.annotate(
        count=Count('records')
    ).filter(count__gt=0).order_by('-count')[:10]

    # Hostel breakdown
    by_hostel = DisciplineRecord.objects.values('hostel_name').annotate(
        count=Count('id')
    ).exclude(hostel_name='').order_by('-count')[:10]

    # Repeat offenders (2+ records)
    repeat_offenders = Student.objects.annotate(
        record_count=Count('discipline_records')
    ).filter(record_count__gte=2).order_by('-record_count')[:10]

    # Monthly trend (last 6 months) — list() so template |last filter works
    monthly = list(DisciplineRecord.objects.annotate(
        month=TruncMonth('incident_date')
    ).values('month').annotate(count=Count('id')).order_by('month'))

    return render(request, 'admin/discipline_report.html', {
        'total': total, 'active': active, 'resolved': resolved,
        'unpaid': unpaid, 'total_fines': total_fines,
        'collected': collected, 'pending_fines': pending_fines,
        'by_category': by_category, 'by_hostel': by_hostel,
        'repeat_offenders': repeat_offenders, 'monthly': monthly,
    })


# ─── Maintenance Logs ─────────────────────────────────────────────────────────

@admin_required
def maintenance_list(request):
    status_filter = request.GET.get('status', '')
    qs = MaintenanceLog.objects.select_related('room__hostel', 'created_by').order_by('-created_at')
    if status_filter:
        qs = qs.filter(status=status_filter)
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            room_id = request.POST.get('room_id')
            MaintenanceLog.objects.create(
                room_id        = room_id,
                category       = request.POST.get('category'),
                issue          = request.POST.get('issue', '').strip(),
                assigned_to    = request.POST.get('assigned_to', '').strip(),
                scheduled_date = request.POST.get('scheduled_date') or None,
                created_by     = request.user,
            )
            messages.success(request, 'Maintenance log created.')
            return redirect('maintenance_list')
        elif action == 'update_status':
            log_id     = request.POST.get('log_id')
            new_status = request.POST.get('new_status')
            update_data = {'status': new_status}
            if new_status == 'resolved':
                update_data['resolved_at'] = timezone.now()
            MaintenanceLog.objects.filter(pk=log_id).update(**update_data)
            messages.success(request, 'Status updated.')
            return redirect('maintenance_list')

    rooms = Room.objects.select_related('hostel').order_by('hostel__name', 'room_number')
    counts = {
        'all':         MaintenanceLog.objects.count(),
        'open':        MaintenanceLog.objects.filter(status='open').count(),
        'in_progress': MaintenanceLog.objects.filter(status='in_progress').count(),
        'on_hold':     MaintenanceLog.objects.filter(status='on_hold').count(),
        'resolved':    MaintenanceLog.objects.filter(status='resolved').count(),
        'closed':      MaintenanceLog.objects.filter(status='closed').count(),
    }
    return render(request, 'admin/maintenance.html', {
        'page': page, 'status_filter': status_filter,
        'rooms': rooms, 'counts': counts,
        'categories': MaintenanceLog.Category.choices,
        'statuses': MaintenanceLog.Status.choices,
    })


# ─── Maintenance Incharge Dashboard ──────────────────────────────────────────

@maintenance_only
def maintenance_incharge_dashboard(request):  # noqa: C901
    """Maintenance incharge: view assigned complaints, update status, add comments."""
    from datetime import timedelta
    today    = dt_date.today()
    now      = timezone.now()
    role     = request.user.role
    is_mi    = (role == User.Role.MAINTENANCE)
    is_admin = role in (User.Role.ADMIN, User.Role.SUPER_ADMIN, User.Role.WARDEN)

    # Base queryset — maintenance sees only complaints assigned to them
    if is_mi:
        base_qs = Complaint.objects.filter(
            assigned_to=request.user,
            status__in=[Complaint.Status.FORWARDED, Complaint.Status.IN_PROGRESS,
                        Complaint.Status.RESOLVED, Complaint.Status.CLOSED],
        )
    else:
        base_qs = Complaint.objects.filter(
            status__in=[Complaint.Status.FORWARDED, Complaint.Status.IN_PROGRESS,
                        Complaint.Status.RESOLVED, Complaint.Status.CLOSED],
        )

    base_qs = base_qs.select_related('student', 'room__hostel', 'assigned_to')

    # Filters
    f_status   = request.GET.get('status', '')
    f_category = request.GET.get('category', '')
    cutoff24   = now - timedelta(hours=24)
    qs = base_qs.order_by('forwarded_at')
    if f_status == 'overdue':
        qs = qs.filter(forwarded_at__lte=cutoff24,
                       status__in=[Complaint.Status.FORWARDED, Complaint.Status.IN_PROGRESS])
    elif f_status:
        qs = qs.filter(status=f_status)
    if f_category:
        qs = qs.filter(category=f_category)

    # Stats
    counts = {
        'total':       base_qs.count(),
        'forwarded':   base_qs.filter(status='forwarded').count(),
        'in_progress': base_qs.filter(status='in_progress').count(),
        'resolved':    base_qs.filter(status='resolved').count(),
        'overdue':     base_qs.filter(forwarded_at__lte=cutoff24,
                                      status__in=['forwarded', 'in_progress']).count(),
        'today':       base_qs.filter(created_at__date=today).count(),
    }

    paginator = Paginator(qs, 15)
    page_obj  = paginator.get_page(request.GET.get('page'))

    return render(request, 'staff/maintenance_incharge.html', {
        'page_obj':     page_obj,
        'counts':       counts,
        'statuses':     Complaint.Status.choices,
        'categories':   Complaint.Category.choices,
        'filters':      {'status': f_status, 'category': f_category},
        'is_mi':        is_mi,
        'is_admin':     is_admin,
        'today':        today,
        'now':          now,
    })


# ─── Bed Management ───────────────────────────────────────────────────────────

@admin_only
def bed_manage(request, room_pk):
    room = get_object_or_404(Room, pk=room_pk)
    beds = Bed.objects.filter(room=room)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            bed_number = request.POST.get('bed_number', '').strip()
            if bed_number:
                Bed.objects.get_or_create(room=room, bed_number=bed_number)
                messages.success(request, f'Bed {bed_number} added.')
        elif action == 'auto_create':
            for i in range(1, room.capacity + 1):
                Bed.objects.get_or_create(room=room, bed_number=str(i))
            messages.success(request, f'{room.capacity} beds auto-created.')
        elif action == 'update_status':
            Bed.objects.filter(pk=request.POST.get('bed_id'), room=room).update(
                status=request.POST.get('new_status'))
            messages.success(request, 'Bed status updated.')
        elif action == 'delete':
            Bed.objects.filter(pk=request.POST.get('bed_id'), room=room).delete()
            messages.success(request, 'Bed removed.')
        return redirect('bed_manage', room_pk=room_pk)

    return render(request, 'admin/bed_manage.html', {
        'room': room, 'beds': beds,
        'bed_statuses': Bed.Status.choices,
    })


# ─── Enhanced Admin Dashboard (replaces old one) ──────────────────────────────

@admin_required
def admin_dashboard_v2(request):
    today = dt_date.today()

    # ── Scope: warden sees only their hostel; admin can pick one or see all ──
    warden_hostel = get_hostel_scope(request.user)
    is_warden     = warden_hostel is not None

    if is_warden:
        selected_hostel = warden_hostel
    else:
        hostel_pk = request.GET.get('hostel', '')
        selected_hostel = Hostel.objects.filter(pk=hostel_pk).first() if hostel_pk else None

    def hs(qs):
        """Filter queryset by hostel scope when applicable."""
        if selected_hostel:
            return qs
        return qs

    # Build scoped querysets
    if selected_hostel:
        h = selected_hostel
        student_qs    = Student.objects.filter(is_resident=True, allocations__room__hostel=h, allocations__status='active').distinct()
        room_qs       = Room.objects.filter(hostel=h)
        attend_qs     = Attendance.objects.filter(date=today, student__allocations__room__hostel=h, student__allocations__status='active').distinct()
        gatepass_qs   = GatePass.objects.filter(student__allocations__room__hostel=h, student__allocations__status='active').distinct()
        complaint_qs  = Complaint.objects.filter(status='open', student__allocations__room__hostel=h, student__allocations__status='active').distinct()
        visitor_qs    = Visitor.objects.filter(hostel=h)
        maintain_qs   = MaintenanceLog.objects.filter(room__hostel=h)
        discipline_qs = DisciplineRecord.objects.filter(student__allocations__room__hostel=h, student__allocations__status='active').distinct()
        alloc_qs      = RoomAllocation.objects.filter(status='active', room__hostel=h)
    else:
        student_qs    = Student.objects.filter(is_resident=True)
        room_qs       = Room.objects.all()
        attend_qs     = Attendance.objects.filter(date=today)
        gatepass_qs   = GatePass.objects.all()
        complaint_qs  = Complaint.objects.filter(status='open')
        visitor_qs    = Visitor.objects.all()
        maintain_qs   = MaintenanceLog.objects.all()
        discipline_qs = DisciplineRecord.objects.all()
        alloc_qs      = RoomAllocation.objects.filter(status='active')

    total_residents = student_qs.count()
    present_today   = attend_qs.filter(status='present').count()
    absent_today    = attend_qs.filter(status='absent').count()
    leave_today     = attend_qs.filter(status='leave').count()
    marked_today    = attend_qs.count()

    # Auto-flag overstays
    for gp in gatepass_qs.filter(status='approved'):
        gp.check_overstay()

    # QR-verified: exit scanned by guard but entry not yet scanned = physically outside right now
    students_out = gatepass_qs.filter(
        exit_qr_scanned=True, entry_qr_scanned=False
    ).count()

    _room_alloc_qs = room_qs.annotate(
        _active=Count('allocations', filter=Q(allocations__status='active'))
    )
    partially_occupied_rooms = _room_alloc_qs.filter(_active__gt=0, _active__lt=F('capacity')).count()

    # Gender-wise counts
    boys_total    = student_qs.filter(gender='male').count()
    girls_total   = student_qs.filter(gender='female').count()
    boys_incampus  = student_qs.filter(gender='male', allocations__room__designation='incampus', allocations__status='active').distinct().count()
    boys_outcampus = student_qs.filter(gender='male', allocations__room__designation='outcampus', allocations__status='active').distinct().count()
    # Staff counts (exclude superadmin and student roles)
    from apps.accounts.models import User as _User
    staff_total   = _User.objects.exclude(role__in=('superadmin', 'student')).filter(is_active=True).count()
    warden_count  = _User.objects.filter(role='warden', is_active=True).count()

    stats = {
        'total_hostels':          Hostel.objects.count() if not selected_hostel else 1,
        'total_students':         total_residents,
        'total_rooms':            room_qs.count(),
        'vacant_rooms':           room_qs.filter(status='vacant').count(),
        'occupied_rooms':         room_qs.filter(status='occupied').count(),
        'partially_occupied_rooms': partially_occupied_rooms,
        'open_complaints':   complaint_qs.count(),
        'present_today':     present_today,
        'absent_today':      absent_today,
        'leave_today':       leave_today,
        'unmarked_today':    total_residents - marked_today,
        'attendance_done':   marked_today > 0,
        'students_out':      students_out,
        'pending_passes':    gatepass_qs.filter(status='pending').count(),
        'overstayed_passes': gatepass_qs.filter(is_overstayed=True, status='approved').count(),
        'pending_visitors':  visitor_qs.filter(status='pending').count(),
        'pending_apps':      RegistrationSubmission.objects.filter(status='submitted').count(),
        'open_maintenance':  maintain_qs.filter(status='open').count(),
        'active_discipline': discipline_qs.filter(status='active').count(),
        'is_warden':         is_warden,
        'boys_total':        boys_total,
        'boys_incampus':     boys_incampus,
        'boys_outcampus':    boys_outcampus,
        'girls_total':       girls_total,
        'staff_total':       staff_total,
        'warden_count':      warden_count,
    }

    # Hostel-wise gender breakdown
    hostel_wise_data = []
    for _hw in Hostel.objects.order_by('name'):
        _boys  = Student.objects.filter(allocations__room__hostel=_hw, allocations__status='active', gender='male').distinct().count()
        _girls = Student.objects.filter(allocations__room__hostel=_hw, allocations__status='active', gender='female').distinct().count()
        _total = _boys + _girls
        hostel_wise_data.append({'hostel': _hw, 'boys': _boys, 'girls': _girls, 'total': _total})

    # Hostel occupancy table — admin sees all, warden sees only theirs
    hostels_list = [selected_hostel] if selected_hostel else list(Hostel.objects.all())
    hostels_data = []
    for h in hostels_list:
        total_beds    = Bed.objects.filter(room__hostel=h).count()
        occupied_beds = Bed.objects.filter(room__hostel=h, status='occupied').count()
        total_rooms   = Room.objects.filter(hostel=h).count()
        vacant_rooms  = Room.objects.filter(hostel=h, status='vacant').count()
        pct = round(occupied_beds / total_beds * 100) if total_beds else 0
        hostels_data.append({
            'hostel': h, 'total_rooms': total_rooms, 'vacant_rooms': vacant_rooms,
            'total_beds': total_beds, 'occupied_beds': occupied_beds,
            'vacant_beds': total_beds - occupied_beds, 'pct': pct,
        })

    # Per-hostel today's gate pass + leave counts (added to hostels_data)
    for row in hostels_data:
        h = row['hostel']
        _af = Q(student__allocations__room__hostel=h, student__allocations__status='active')
        row['gp_today']       = GatePass.objects.filter(_af, departure_date=today).distinct().count()
        row['leave_today']    = LeaveApplication.objects.filter(
            _af, status='approved', from_date__lte=today, to_date__gte=today
        ).distinct().count()
        row['out_now']        = GatePass.objects.filter(
            _af, exit_qr_scanned=True, entry_qr_scanned=False
        ).distinct().count()

    # Overall today totals for stat cards
    _leave_qs = LeaveApplication.objects.filter(
        status='approved', from_date__lte=today, to_date__gte=today
    )
    if selected_hostel:
        _leave_qs = _leave_qs.filter(
            student__allocations__room__hostel=selected_hostel,
            student__allocations__status='active'
        )
    stats['today_gate_passes'] = gatepass_qs.filter(departure_date=today).count()
    stats['today_on_leave']    = _leave_qs.distinct().count()

    recent_allocations = alloc_qs.select_related('student', 'room__hostel').order_by('-created_at')[:5]
    recent_complaints  = complaint_qs.select_related('student').order_by('-created_at')[:5]
    announcements      = Announcement.objects.filter(is_active=True).filter(
        Q(expires_at__isnull=True) | Q(expires_at__gte=today)).order_by('-created_at')[:3]
    overstayed         = gatepass_qs.filter(
        is_overstayed=True, status='approved'
    ).select_related('student')[:5]
    pending_visitors   = visitor_qs.filter(status='pending').select_related('hostel', 'student')[:5]

    # Staff panel data for offcanvas
    from apps.accounts.models import User as _User2
    _hod_staff    = list(_User2.objects.filter(role='hod', is_active=True).order_by('name'))
    _hostel_staff = list(_User2.objects.filter(
        role__in=['admin', 'warden', 'security', 'maintenance', 'mess'], is_active=True
    ).order_by('role', 'name'))
    _custom_staff = list(_User2.objects.exclude(
        role__in=['superadmin', 'student', 'hod', 'admin', 'warden', 'security', 'maintenance', 'mess']
    ).filter(is_active=True).order_by('role', 'name'))

    response = render(request, 'admin/dashboard.html', {
        'stats': stats, 'today': today,
        'recent_allocations': recent_allocations,
        'recent_complaints':  recent_complaints,
        'announcements':      announcements,
        'hostels_data':       hostels_data,
        'hostel_wise_data':   hostel_wise_data,
        'overstayed':         overstayed,
        'pending_visitors':   pending_visitors,
        'all_hostels':        Hostel.objects.all(),
        'selected_hostel':    selected_hostel,
        'is_warden':          is_warden,
        'hod_staff':          _hod_staff,
        'hostel_staff':       _hostel_staff,
        'custom_staff':       _custom_staff,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return response


# ─── Gate Pass Report ─────────────────────────────────────────────────────────

@admin_required
def gate_pass_report(request):
    date_str = request.GET.get('date', '')
    try:
        report_date = dt_date.fromisoformat(date_str) if date_str else dt_date.today()
    except ValueError:
        report_date = dt_date.today()

    warden_hostel = get_hostel_scope(request.user)
    is_warden     = warden_hostel is not None
    if is_warden:
        hostels = [warden_hostel]
    else:
        hostel_pk = request.GET.get('hostel', '')
        hostels   = list(Hostel.objects.filter(pk=hostel_pk)) if hostel_pk else list(Hostel.objects.all().order_by('name'))

    hostel_rows = []
    for h in hostels:
        af    = Q(student__allocations__room__hostel=h, student__allocations__status='active')
        gp_qs = GatePass.objects.filter(af, departure_date=report_date).select_related('student', 'approved_by').distinct().order_by('student__name')
        hostel_rows.append({'hostel': h, 'gp_list': gp_qs, 'gp_count': gp_qs.count()})

    return render(request, 'admin/gate_pass_report.html', {
        'hostel_rows':   hostel_rows,
        'report_date':   report_date,
        'today':         dt_date.today(),
        'all_hostels':   Hostel.objects.all().order_by('name'),
        'sel_hostel_pk': request.GET.get('hostel', ''),
        'is_warden':     is_warden,
    })


# ─── Leave Report ─────────────────────────────────────────────────────────────

@admin_required
def leave_report(request):
    date_str = request.GET.get('date', '')
    try:
        report_date = dt_date.fromisoformat(date_str) if date_str else dt_date.today()
    except ValueError:
        report_date = dt_date.today()

    warden_hostel = get_hostel_scope(request.user)
    is_warden     = warden_hostel is not None
    if is_warden:
        hostels = [warden_hostel]
    else:
        hostel_pk = request.GET.get('hostel', '')
        hostels   = list(Hostel.objects.filter(pk=hostel_pk)) if hostel_pk else list(Hostel.objects.all().order_by('name'))

    hostel_rows = []
    for h in hostels:
        af       = Q(student__allocations__room__hostel=h, student__allocations__status='active')
        leave_qs = LeaveApplication.objects.filter(
            af, status='approved', from_date__lte=report_date, to_date__gte=report_date
        ).select_related('student', 'reviewed_by').distinct().order_by('student__name')
        hostel_rows.append({'hostel': h, 'leave_list': leave_qs, 'leave_count': leave_qs.count()})

    return render(request, 'admin/leave_report.html', {
        'hostel_rows':   hostel_rows,
        'report_date':   report_date,
        'today':         dt_date.today(),
        'all_hostels':   Hostel.objects.all().order_by('name'),
        'sel_hostel_pk': request.GET.get('hostel', ''),
        'is_warden':     is_warden,
    })


# ─── Outsiders ────────────────────────────────────────────────────────────────

@admin_required
def outsiders(request):
    date_str = request.GET.get('date', '')
    try:
        filter_date = dt_date.fromisoformat(date_str) if date_str else None
    except ValueError:
        filter_date = None

    warden_hostel = get_hostel_scope(request.user)
    is_warden     = warden_hostel is not None
    hostels       = [warden_hostel] if is_warden else list(
        Hostel.objects.select_related('warden').order_by('name')
    )

    total_outside = 0
    hostel_rows   = []

    for h in hostels:
        af = Q(student__allocations__room__hostel=h, student__allocations__status='active')

        if filter_date:
            # Historical: anyone whose exit QR was scanned on that date
            gp_qs = GatePass.objects.filter(
                af, exit_qr_scanned=True, exit_time__date=filter_date,
            ).select_related('student', 'approved_by', 'exit_allowed_by', 'entry_allowed_by').distinct()
        else:
            # Live: exit scanned but entry NOT yet scanned
            gp_qs = GatePass.objects.filter(
                af, exit_qr_scanned=True, entry_qr_scanned=False,
            ).select_related('student', 'approved_by', 'exit_allowed_by').distinct()

        gp_list = list(gp_qs.order_by('student__name'))

        # Attach active room (room_number + floor) in one extra query — avoids N+1
        s_ids = [gp.student_id for gp in gp_list]
        if s_ids:
            room_map = {
                a.student_id: a.room
                for a in RoomAllocation.objects.filter(
                    student_id__in=s_ids, status='active', room__hostel=h
                ).select_related('room')
            }
        else:
            room_map = {}
        for gp in gp_list:
            gp.student_room = room_map.get(gp.student_id)

        # Collect distinct departments for the per-card dept filter
        depts = sorted({gp.student.department for gp in gp_list if gp.student.department})

        count = len(gp_list)
        total_outside += count
        hostel_rows.append({
            'hostel': h,
            'warden': getattr(h, 'warden', None),
            'gp_list': gp_list,
            'count': count,
            'depts': depts,
        })

    return render(request, 'admin/outsiders.html', {
        'hostel_rows':   hostel_rows,
        'total_outside': total_outside,
        'filter_date':   filter_date,
        'today':         dt_date.today(),
        'is_warden':     is_warden,
    })


# ─── Hostel Management (Admin only) ──────────────────────────────────────────

@admin_only
def hostel_list(request):
    raw_hostels = Hostel.objects.select_related('warden').annotate(
        room_count=Count('rooms', distinct=True),
        resident_count=Count('rooms__allocations',
                             filter=Q(rooms__allocations__status='active'),
                             distinct=True),
    ).order_by('name')

    hostels_data = []
    for h in raw_hostels:
        floors      = Room.objects.filter(hostel=h).values_list('floor', flat=True).distinct().order_by('floor')
        total_beds  = Bed.objects.filter(room__hostel=h).count()
        occ_beds    = Bed.objects.filter(room__hostel=h, status='occupied').count()
        vacant_beds = total_beds - occ_beds
        vacant_rooms = Room.objects.filter(hostel=h, status='vacant').count()
        occ_rooms    = Room.objects.filter(hostel=h, status='occupied').count()
        maint_rooms  = Room.objects.filter(hostel=h, status='maintenance').count()
        pct = round(occ_beds / total_beds * 100) if total_beds else 0
        hostels_data.append({
            'hostel':        h,
            'room_count':    h.room_count,
            'resident_count':h.resident_count,
            'floor_count':   floors.count(),
            'floors':        list(floors),
            'total_beds':    total_beds,
            'occ_beds':      occ_beds,
            'vacant_beds':   vacant_beds,
            'vacant_rooms':  vacant_rooms,
            'occ_rooms':     occ_rooms,
            'maint_rooms':   maint_rooms,
            'pct':           pct,
        })

    all_wardens      = User.objects.filter(role=User.Role.WARDEN, is_active=True)
    all_mess_incharge = User.objects.filter(role=User.Role.MESS, is_active=True)
    return render(request, 'admin/hostels.html', {
        'hostels_data':      hostels_data,
        'all_wardens':       all_wardens,
        'all_mess_incharge': all_mess_incharge,
    })


@admin_only
def create_hostel(request):
    if request.method == 'POST':
        name           = request.POST.get('name', '').strip()
        htype          = request.POST.get('type', '')
        address        = request.POST.get('address', '').strip()
        warden_id      = request.POST.get('warden') or None
        mess_incharge_id = request.POST.get('mess_incharge') or None
        if not name or not htype:
            messages.error(request, 'Name and type are required.')
            return redirect('hostel_list')
        if Hostel.objects.filter(name__iexact=name).exists():
            messages.error(request, f'A hostel named "{name}" already exists.')
            return redirect('hostel_list')
        warden        = User.objects.filter(pk=warden_id,      role=User.Role.WARDEN).first() if warden_id else None
        mess_incharge = User.objects.filter(pk=mess_incharge_id, role=User.Role.MESS).first()  if mess_incharge_id else None
        Hostel.objects.create(name=name, type=htype, address=address,
                              warden=warden, mess_incharge=mess_incharge)
        messages.success(request, f'Hostel "{name}" created successfully.')
    return redirect('hostel_list')


@admin_only
def edit_hostel(request, pk):
    hostel = get_object_or_404(Hostel, pk=pk)
    if request.method == 'POST':
        hostel.name          = request.POST.get('name', hostel.name).strip()
        hostel.type          = request.POST.get('type', hostel.type)
        hostel.address       = request.POST.get('address', '').strip()
        warden_id            = request.POST.get('warden') or None
        mess_incharge_id     = request.POST.get('mess_incharge') or None
        hostel.warden        = User.objects.filter(pk=warden_id,       role=User.Role.WARDEN).first() if warden_id else None
        hostel.mess_incharge = User.objects.filter(pk=mess_incharge_id, role=User.Role.MESS).first()  if mess_incharge_id else None
        hostel.save()
        messages.success(request, f'Hostel "{hostel.name}" updated.')
        return redirect('hostel_list')
    all_wardens       = User.objects.filter(role=User.Role.WARDEN, is_active=True)
    all_mess_incharge = User.objects.filter(role=User.Role.MESS,   is_active=True)
    return render(request, 'admin/hostel_edit.html', {
        'hostel':            hostel,
        'all_wardens':       all_wardens,
        'all_mess_incharge': all_mess_incharge,
        'hostel_types':      Hostel.Type.choices,
    })


@admin_only
@require_POST
def delete_hostel(request, pk):
    hostel = get_object_or_404(Hostel, pk=pk)
    name = hostel.name
    if hostel.rooms.exists():
        messages.error(request, f'Cannot delete "{name}" — it has rooms assigned. Remove rooms first.')
        return redirect('hostel_list')
    hostel.delete()
    messages.success(request, f'Hostel "{name}" deleted.')
    return redirect('hostel_list')


@admin_only
@require_POST
def create_warden(request):
    name      = request.POST.get('name', '').strip()
    email     = request.POST.get('email', '').strip().lower()
    phone     = request.POST.get('phone', '').strip()
    password  = request.POST.get('password', '').strip()
    if not name or not email or not password:
        messages.error(request, 'Name, email, and password are required.')
        return redirect('hostel_list')
    if User.objects.filter(email=email).exists():
        messages.error(request, f'An account with email "{email}" already exists.')
        return redirect('hostel_list')
    User.objects.create_user(email=email, password=password, name=name, phone=phone, role='warden')
    messages.success(request, f'Warden account created for {name} ({email}).')
    return redirect('hostel_list')


# ─── Semester & Reallotment ───────────────────────────────────────────────────

@admin_only
def semester_list(request):
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel:
        return redirect('reallotment')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            name         = request.POST.get('name', '').strip()
            sem_type     = request.POST.get('sem_type', '')
            start_date   = request.POST.get('start_date')
            end_date     = request.POST.get('end_date') or None
            make_current = request.POST.get('is_current') == 'on'
            if not name or not sem_type or not start_date:
                messages.error(request, 'Name, type and start date are required.')
            else:
                sem = Semester.objects.create(
                    name=name, sem_type=sem_type, start_date=start_date,
                    end_date=end_date, is_current=make_current, created_by=request.user
                )
                messages.success(request, f'Semester "{sem.name}" created.')
        elif action == 'set_current':
            sem = get_object_or_404(Semester, pk=request.POST.get('sem_id'))
            sem.is_current = True
            sem.save()
            messages.success(request, f'"{sem.name}" is now the current semester.')
        elif action == 'delete':
            sem = get_object_or_404(Semester, pk=request.POST.get('sem_id'))
            if sem.allocations.exists():
                messages.error(request, f'Cannot delete "{sem.name}" — it has allocations linked to it.')
            else:
                sem.delete()
                messages.success(request, 'Semester deleted.')
        return redirect('semester_list')

    semesters    = Semester.objects.select_related('created_by').all()
    current      = Semester.current()
    active_count = RoomAllocation.objects.filter(status='active').count()
    return render(request, 'admin/semesters.html', {
        'semesters':    semesters,
        'current':      current,
        'active_count': active_count,
        'sem_types':    Semester.SemType.choices,
    })


@admin_only
@require_POST
def close_semester(request):
    warden_hostel = get_hostel_scope(request.user)
    hostel_pk     = request.POST.get('hostel_pk') or None
    today         = dt_date.today()

    qs = RoomAllocation.objects.filter(status='active').select_related('room', 'student')
    if warden_hostel:
        qs = qs.filter(room__hostel=warden_hostel)
    elif hostel_pk:
        qs = qs.filter(room__hostel_id=hostel_pk)

    count = qs.count()
    if count == 0:
        messages.warning(request, 'No active allocations found to close out.')
        return redirect(request.POST.get('next', 'reallotment'))

    room_ids    = list(qs.values_list('room_id', flat=True))
    student_ids = list(qs.values_list('student_id', flat=True))
    qs.update(status='checkout', check_out=today)
    Room.objects.filter(pk__in=room_ids).update(status='vacant')
    Student.objects.filter(pk__in=student_ids).update(is_resident=False)

    sem_id = request.POST.get('semester_id')
    if sem_id:
        Semester.objects.filter(pk=sem_id).update(is_current=False, closed_at=timezone.now())

    messages.success(request, f'Semester closed. {count} student(s) checked out. All rooms marked vacant.')
    return redirect(request.POST.get('next', 'reallotment'))


@admin_only
def reallotment(request):
    warden_hostel = get_hostel_scope(request.user)
    current_sem   = Semester.current()

    allocated_student_ids = RoomAllocation.objects.filter(
        status='active'
    ).values_list('student_id', flat=True)

    unallocated = Student.objects.filter(
        is_resident=True
    ).exclude(pk__in=allocated_student_ids).select_related('user')

    if warden_hostel:
        last_hostel_ids = RoomAllocation.objects.filter(
            student__in=unallocated, room__hostel=warden_hostel
        ).values_list('student_id', flat=True)
        unallocated = unallocated.filter(pk__in=last_hostel_ids)

    avail_rooms = Room.objects.filter(status='vacant').select_related('hostel').order_by('hostel__name', 'room_number')
    if warden_hostel:
        avail_rooms = avail_rooms.filter(hostel=warden_hostel)

    unallocated_data = []
    for s in unallocated.order_by('name'):
        last_alloc = RoomAllocation.objects.filter(student=s).order_by('-created_at').first()
        can_auto   = (
            last_alloc is not None and
            last_alloc.room.status == 'vacant' and
            last_alloc.room.current_occupants < last_alloc.room.capacity
        )
        unallocated_data.append({
            'student':     s,
            'last_room':   last_alloc.room if last_alloc else None,
            'last_hostel': last_alloc.room.hostel if last_alloc else None,
            'can_auto':    can_auto,
        })

    hostels      = Hostel.objects.all() if not warden_hostel else [warden_hostel]
    all_semesters = Semester.objects.all()

    return render(request, 'admin/reallotment.html', {
        'unallocated_data':      unallocated_data,
        'avail_rooms':           avail_rooms,
        'current_sem':           current_sem,
        'all_semesters':         all_semesters,
        'hostels':               hostels,
        'warden_hostel':         warden_hostel,
        'is_warden':             warden_hostel is not None,
        'today':                 dt_date.today(),
        'total_unallocated':     len(unallocated_data),
        'total_available_rooms': avail_rooms.count(),
    })


@admin_only
@require_POST
def auto_reallot(request):
    warden_hostel = get_hostel_scope(request.user)
    current_sem   = Semester.current()
    today         = dt_date.today()
    check_in_date = request.POST.get('check_in_date', str(today))

    allocated_student_ids = RoomAllocation.objects.filter(
        status='active'
    ).values_list('student_id', flat=True)
    unallocated = Student.objects.filter(is_resident=True).exclude(pk__in=allocated_student_ids)
    if warden_hostel:
        last_hostel_ids = RoomAllocation.objects.filter(
            student__in=unallocated, room__hostel=warden_hostel
        ).values_list('student_id', flat=True)
        unallocated = unallocated.filter(pk__in=last_hostel_ids)

    success_count = 0
    skip_count    = 0
    for student in unallocated:
        last_alloc = RoomAllocation.objects.filter(student=student).order_by('-created_at').first()
        if not last_alloc:
            skip_count += 1
            continue
        room = last_alloc.room
        if room.status != 'vacant' or room.current_occupants >= room.capacity:
            skip_count += 1
            continue
        RoomAllocation.objects.create(
            student=student, room=room, bed=last_alloc.bed,
            semester=current_sem, allocated_by=request.user,
            check_in=check_in_date,
        )
        if room.current_occupants >= room.capacity:
            room.status = 'occupied'
            room.save(update_fields=['status'])
        student.is_resident = True
        student.save(update_fields=['is_resident'])
        success_count += 1

    msg = f'Auto-reallotment done: {success_count} student(s) allocated.'
    if skip_count:
        msg += f' {skip_count} skipped (no previous room or room unavailable).'
    messages.success(request, msg)
    return redirect('reallotment')


@admin_only
@require_POST
def single_reallot(request):
    student_pk  = request.POST.get('student_pk')
    room_pk     = request.POST.get('room_pk')
    check_in    = request.POST.get('check_in_date', str(dt_date.today()))
    current_sem = Semester.current()

    student = get_object_or_404(Student, pk=student_pk)
    room    = get_object_or_404(Room, pk=room_pk)

    if RoomAllocation.objects.filter(student=student, status='active').exists():
        messages.warning(request, f'{student.name} already has an active allocation.')
        return redirect('reallotment')

    if room.current_occupants >= room.capacity:
        messages.error(request, f'Room {room.room_number} is at full capacity.')
        return redirect('reallotment')

    RoomAllocation.objects.create(
        student=student, room=room, semester=current_sem,
        allocated_by=request.user, check_in=check_in,
    )
    if room.current_occupants >= room.capacity:
        room.status = 'occupied'
        room.save(update_fields=['status'])
    student.is_resident = True
    student.save(update_fields=['is_resident'])

    messages.success(request, f'{student.name} allocated to Room {room.room_number} ({room.hostel.name}).')
    return redirect('reallotment')


# ─── Hostel / Floor / Room Detail ─────────────────────────────────────────────

@admin_required
def hostel_detail(request, pk):
    """Interactive floor/room management page for a hostel."""
    hostel = get_object_or_404(Hostel, pk=pk)
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel and warden_hostel.pk != hostel.pk:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    if request.method == 'POST':
        action = request.POST.get('action')

        # ── Add Room ──
        if action == 'add_room':
            room_number = request.POST.get('room_number', '').strip()
            floor       = int(request.POST.get('floor', 1))
            wing        = request.POST.get('wing', '').strip()
            room_type   = request.POST.get('room_type', 'Standard').strip()
            capacity    = int(request.POST.get('capacity', 2))
            if not room_number:
                messages.error(request, 'Room number is required.')
            elif Room.objects.filter(hostel=hostel, room_number=room_number).exists():
                messages.error(request, f'Room {room_number} already exists in this hostel.')
            else:
                Room.objects.create(
                    hostel=hostel, room_number=room_number, floor=floor,
                    wing=wing, room_type=room_type, capacity=capacity,
                    status='vacant'
                )
                messages.success(request, f'Room {room_number} added to Floor {floor}.')
            return redirect(f'{request.path}?floor={floor}')

        # ── Edit Room (inline) ──
        elif action == 'edit_room':
            room = get_object_or_404(Room, pk=request.POST.get('room_id'), hostel=hostel)
            room.room_number = request.POST.get('room_number', room.room_number).strip()
            room.floor       = int(request.POST.get('floor', room.floor))
            room.wing        = request.POST.get('wing', '').strip()
            room.room_type   = request.POST.get('room_type', room.room_type).strip()
            room.capacity    = int(request.POST.get('capacity', room.capacity))
            room.status      = request.POST.get('status', room.status)
            room.save()
            room.amenities.set(RoomAmenity.objects.filter(pk__in=request.POST.getlist('amenities')))
            messages.success(request, f'Room {room.room_number} updated.')
            return redirect(f'{request.path}?floor={room.floor}')

        # ── Delete Room ──
        elif action == 'delete_room':
            room = get_object_or_404(Room, pk=request.POST.get('room_id'), hostel=hostel)
            if room.allocations.filter(status='active').exists():
                messages.error(request, f'Room {room.room_number} has active residents — checkout first.')
            else:
                floor = room.floor
                room.delete()
                messages.success(request, f'Room deleted.')
                return redirect(f'{request.path}?floor={floor}')

        # ── Add Asset ──
        elif action == 'add_asset':
            room = get_object_or_404(Room, pk=request.POST.get('room_id'), hostel=hostel)
            name = request.POST.get('asset_name', '').strip()
            if name:
                RoomAsset.objects.create(
                    room=room, name=name,
                    quantity=int(request.POST.get('asset_qty', 1)),
                    condition=request.POST.get('asset_condition', 'good'),
                    asset_tag=request.POST.get('asset_tag', '').strip(),
                    notes=request.POST.get('asset_notes', '').strip(),
                    added_by=request.user
                )
                messages.success(request, f'Asset added to Room {room.room_number}.')
            return redirect(f'{request.path}?floor={room.floor}')

        # ── Edit Asset ──
        elif action == 'edit_asset':
            asset = get_object_or_404(RoomAsset, pk=request.POST.get('asset_id'))
            floor = asset.room.floor
            asset.name      = request.POST.get('asset_name', asset.name).strip()
            asset.quantity  = int(request.POST.get('asset_qty', asset.quantity))
            asset.condition = request.POST.get('asset_condition', asset.condition)
            asset.asset_tag = request.POST.get('asset_tag', '').strip()
            asset.notes     = request.POST.get('asset_notes', '').strip()
            asset.save()
            messages.success(request, f'Asset "{asset.name}" updated.')
            return redirect(f'{request.path}?floor={floor}')

        # ── Delete Asset ──
        elif action == 'delete_asset':
            asset = get_object_or_404(RoomAsset, pk=request.POST.get('asset_id'))
            floor = asset.room.floor
            asset.delete()
            messages.success(request, 'Asset removed.')
            return redirect(f'{request.path}?floor={floor}')

        # ── Bulk Add Rooms ──
        elif action == 'bulk_add_rooms':
            floor      = int(request.POST.get('floor', 1))
            prefix     = request.POST.get('prefix', '').strip()
            start_num  = int(request.POST.get('start_num', 1))
            count      = min(int(request.POST.get('count', 1)), 50)  # cap at 50
            room_type  = request.POST.get('room_type', 'Standard').strip()
            capacity   = int(request.POST.get('capacity', 2))
            wing       = request.POST.get('wing', '').strip()
            created, skipped = 0, []
            for i in range(count):
                room_number = f'{prefix}{start_num + i}'
                if Room.objects.filter(hostel=hostel, room_number=room_number).exists():
                    skipped.append(room_number)
                else:
                    Room.objects.create(
                        hostel=hostel, room_number=room_number, floor=floor,
                        wing=wing, room_type=room_type, capacity=capacity, status='vacant'
                    )
                    created += 1
            if created:
                messages.success(request, f'{created} room{"s" if created>1 else ""} added to Floor {floor}.')
            if skipped:
                messages.warning(request, f'Skipped {len(skipped)} existing rooms: {", ".join(skipped[:5])}{"…" if len(skipped)>5 else ""}')
            return redirect(f'{request.path}?floor={floor}')

        # ── Add Bed ──
        elif action == 'add_bed':
            room = get_object_or_404(Room, pk=request.POST.get('room_id'), hostel=hostel)
            bed_number = request.POST.get('bed_number', '').strip()
            if bed_number:
                Bed.objects.get_or_create(room=room, bed_number=bed_number,
                                          defaults={'status': 'available'})
                messages.success(request, f'Bed {bed_number} added.')
            return redirect(f'{request.path}?floor={room.floor}')

        # ── Auto-create Beds ──
        elif action == 'auto_beds':
            room = get_object_or_404(Room, pk=request.POST.get('room_id'), hostel=hostel)
            for i in range(1, room.capacity + 1):
                Bed.objects.get_or_create(room=room, bed_number=str(i),
                                          defaults={'status': 'available'})
            messages.success(request, f'{room.capacity} beds created for Room {room.room_number}.')
            return redirect(f'{request.path}?floor={room.floor}')

        return redirect(request.path)

    # ── GET ──
    rooms = Room.objects.filter(hostel=hostel).prefetch_related(
        'amenities', 'beds', 'assets', 'allocations__student'
    ).order_by('floor', 'room_number')

    # Group by floor
    floors = {}
    for room in rooms:
        f = room.floor
        if f not in floors:
            floors[f] = {'rooms': [], 'total': 0, 'occupied': 0, 'vacant': 0, 'maintenance': 0}
        floors[f]['rooms'].append(room)
        floors[f]['total'] += 1
        if room.status == 'occupied':   floors[f]['occupied'] += 1
        elif room.status == 'vacant':   floors[f]['vacant'] += 1
        elif room.status == 'maintenance': floors[f]['maintenance'] += 1

    floors_list  = [{'floor': f, **data} for f, data in sorted(floors.items())]
    active_floor = int(request.GET.get('floor', floors_list[0]['floor'] if floors_list else 1))
    all_amenities = RoomAmenity.objects.all()

    return render(request, 'admin/hostel_detail.html', {
        'hostel':            hostel,
        'floors_list':       floors_list,
        'active_floor':      active_floor,
        'total_rooms':       rooms.count(),
        'total_floors':      len(floors),
        'all_amenities':     all_amenities,
        'status_choices':    Room.Status.choices,
        'bed_statuses':      Bed.Status.choices,
        'asset_conditions':  RoomAsset.Condition.choices,
        'maint_categories':  MaintenanceLog.Category.choices,
        'room_categories':   ROOM_CATEGORIES_GROUPED,
        'room_cat_capacity': ROOM_CATEGORY_CAPACITY,
    })


@admin_required
def floor_detail(request, hostel_pk, floor_num):
    """All rooms on a specific floor."""
    hostel = get_object_or_404(Hostel, pk=hostel_pk)
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel and warden_hostel.pk != hostel.pk:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    rooms = Room.objects.filter(hostel=hostel, floor=floor_num).prefetch_related(
        'amenities', 'beds',
        'allocations__student'
    ).order_by('room_number')

    return render(request, 'admin/floor_detail.html', {
        'hostel':    hostel,
        'floor_num': floor_num,
        'rooms':     rooms,
    })


@admin_required
def room_detail(request, pk):
    """Full room detail page — info, beds, students, assets, maintenance. All editable."""
    room    = get_object_or_404(Room, pk=pk)
    hostel  = room.hostel
    warden_hostel = get_hostel_scope(request.user)
    if warden_hostel and warden_hostel.pk != hostel.pk:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    if request.method == 'POST':
        action = request.POST.get('action')

        # ── Edit room info ──
        if action == 'edit_room':
            room.room_number = request.POST.get('room_number', room.room_number).strip()
            room.floor       = int(request.POST.get('floor', room.floor))
            room.wing        = request.POST.get('wing', '').strip()
            room.room_type   = request.POST.get('room_type', room.room_type).strip()
            room.capacity    = int(request.POST.get('capacity', room.capacity))
            room.status      = request.POST.get('status', room.status)
            room.save()
            room.amenities.set(RoomAmenity.objects.filter(pk__in=request.POST.getlist('amenities')))
            messages.success(request, 'Room details updated.')

        # ── Add / edit bed ──
        elif action == 'add_bed':
            bed_number = request.POST.get('bed_number', '').strip()
            if bed_number:
                Bed.objects.get_or_create(room=room, bed_number=bed_number,
                                          defaults={'status': 'available'})
                messages.success(request, f'Bed {bed_number} added.')
        elif action == 'update_bed':
            Bed.objects.filter(pk=request.POST.get('bed_id'), room=room).update(
                status=request.POST.get('bed_status')
            )
            messages.success(request, 'Bed status updated.')
        elif action == 'delete_bed':
            Bed.objects.filter(pk=request.POST.get('bed_id'), room=room).delete()
            messages.success(request, 'Bed removed.')
        elif action == 'auto_beds':
            for i in range(1, room.capacity + 1):
                Bed.objects.get_or_create(room=room, bed_number=str(i),
                                          defaults={'status': 'available'})
            messages.success(request, f'{room.capacity} beds auto-created.')

        # ── Assets ──
        elif action == 'add_asset':
            name     = request.POST.get('asset_name', '').strip()
            qty      = int(request.POST.get('asset_qty', 1))
            cond     = request.POST.get('asset_condition', 'good')
            tag      = request.POST.get('asset_tag', '').strip()
            notes    = request.POST.get('asset_notes', '').strip()
            if name:
                RoomAsset.objects.create(
                    room=room, name=name, quantity=qty,
                    condition=cond, asset_tag=tag, notes=notes, added_by=request.user
                )
                messages.success(request, f'Asset "{name}" added.')
        elif action == 'edit_asset':
            asset = get_object_or_404(RoomAsset, pk=request.POST.get('asset_id'), room=room)
            asset.name      = request.POST.get('asset_name', asset.name).strip()
            asset.quantity  = int(request.POST.get('asset_qty', asset.quantity))
            asset.condition = request.POST.get('asset_condition', asset.condition)
            asset.asset_tag = request.POST.get('asset_tag', '').strip()
            asset.notes     = request.POST.get('asset_notes', '').strip()
            asset.save()
            messages.success(request, f'Asset "{asset.name}" updated.')
        elif action == 'delete_asset':
            RoomAsset.objects.filter(pk=request.POST.get('asset_id'), room=room).delete()
            messages.success(request, 'Asset removed.')

        # ── Maintenance ──
        elif action == 'add_maintenance':
            MaintenanceLog.objects.create(
                room=room,
                category    = request.POST.get('category'),
                issue       = request.POST.get('issue', '').strip(),
                assigned_to = request.POST.get('assigned_to', '').strip(),
                scheduled_date = request.POST.get('scheduled_date') or None,
                created_by  = request.user,
            )
            messages.success(request, 'Maintenance log added.')

        return redirect('room_detail', pk=pk)

    allocations   = room.allocations.filter(status='active').select_related('student')
    beds          = room.beds.all()
    assets        = room.assets.all()
    all_amenities = RoomAmenity.objects.all()
    maintenance   = room.maintenance_logs.order_by('-created_at')[:10]
    selected_amenities = list(room.amenities.values_list('pk', flat=True))

    return render(request, 'admin/room_detail.html', {
        'room':               room,
        'hostel':             hostel,
        'allocations':        allocations,
        'beds':               beds,
        'assets':             assets,
        'all_amenities':      all_amenities,
        'selected_amenities': selected_amenities,
        'maintenance':        maintenance,
        'status_choices':     Room.Status.choices,
        'bed_statuses':       Bed.Status.choices,
        'asset_conditions':   RoomAsset.Condition.choices,
        'maint_categories':   MaintenanceLog.Category.choices,
        'maint_statuses':     MaintenanceLog.Status.choices,
    })


# ─── Bulk Upload: Hostels / Rooms ─────────────────────────────────────────────

@admin_only
def bulk_upload_sample(request):  # noqa: C901
    """Download styled sample Excel — new category system, no Wing column, with common facilities."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rooms'

    thin = Border(left=Side(style='thin'),   right=Side(style='thin'),
                  top=Side(style='thin'),    bottom=Side(style='thin'))

    def hfill(h): return PatternFill('solid', fgColor=h)
    cc = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lc = Alignment(horizontal='left',   vertical='center', indent=1)

    # ── Column layout (A–H, 8 cols; no start/end numbers) ──
    #  A: Hostel Name   B: Hostel Type   C: No. of Floors
    #  D: Floor         E: Rooms on Floor
    #  F: Room Category  G: Room Count
    #  H: Designation (incampus/outcampus)
    for col_ltr, w in zip('ABCDEFGH', [18, 13, 12, 7, 14, 18, 9, 18]):
        ws.column_dimensions[col_ltr].width = w

    # ── Row 1: title banner ──
    tc = ws.cell(row=1, column=1,
                 value='  GGI Hostel ERP  —  Room Bulk Upload Template')
    tc.font      = Font(bold=True, color='FFFFFF', size=13)
    tc.fill      = hfill('0D47A1')
    tc.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 26

    # ── Row 2: section group labels ──
    for sc, ec, lbl, clr in [
        (1, 3, 'HOSTEL INFORMATION', '1565C0'),
        (4, 5, 'FLOOR INFO',         '0F4C81'),
        (6, 7, 'ROOM BREAKDOWN',     '1B5E20'),
        (8, 8, 'DESIGNATION',        '6A1B9A'),
    ]:
        c = ws.cell(row=2, column=sc, value=lbl)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = hfill(clr); c.alignment = cc; c.border = thin
        if sc != ec:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
    ws.row_dimensions[2].height = 18

    # ── Row 3: column headers ──
    hdrs = [
        ('Hostel Name*',               '1565C0'),
        ('Hostel Type*\n(boys/girls/staff)', '1565C0'),
        ('Number\nof Floors',          '1565C0'),
        ('Floor',                      '0F4C81'),
        ('Number of\nRooms on Floor',  '0F4C81'),
        ('Room\nCategory',             '1B5E20'),
        ('Room\nCount',                '1B5E20'),
        ('Designation\n(incampus/outcampus)', '6A1B9A'),
    ]
    for col, (lbl, clr) in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=col, value=lbl)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = hfill(clr); c.alignment = cc; c.border = thin
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = 'A4'

    # ── Category colour map ──
    CAT_STYLES = {
        '6S+CW':           ('E0F7FA', '006064'),
        '5S+CW':           ('B2EBF2', '00838F'),
        '5S+WW':           ('E0F2F1', '00695C'),
        '5S+WW+AC':        ('B2DFDB', '00695C'),
        '4S+CWR':          ('E8F5E9', '1B5E20'),
        '4S+CWR+AC':       ('C8E6C9', '1B5E20'),
        '4S+WR':           ('DCEDC8', '33691E'),
        '4S+WR+AC':        ('F1F8E9', '33691E'),
        '3S+CWR':          ('EDE7F6', '4A148C'),
        '3S+CWR+AC':       ('D1C4E9', '4A148C'),
        '3S+WR':           ('E8EAF6', '283593'),
        '3S+WR+AC':        ('F3E5F5', '6A1B9A'),
        '2S+CWR':          ('FCE4EC', 'B71C1C'),
        '2S+CWR+AC':       ('FFCDD2', 'B71C1C'),
        '2S+WR':           ('FBE9E7', 'BF360C'),
        '2S+WR+AC':        ('FFF3E0', 'E65100'),
        '1S+CWR':          ('E0F2F1', '004D40'),
        '1S+CWR+AC':       ('B2DFDB', '004D40'),
        '1S+WR':           ('E3F2FD', '0D47A1'),
        '1S+WR+AC':        ('BBDEFB', '0D47A1'),
        'GYM':             ('FFF8E1', 'E65100'),
        'SPORTS ROOM':     ('E8F5E9', '1B5E20'),
        'ENTERTAINMENT':   ('FCE4EC', 'C62828'),
        'MESS':            ('E3F2FD', '1565C0'),
        'STUDY ROOM':      ('F3E5F5', '6A1B9A'),
        'SITTING ARENA':   ('FFF3E0', 'E65100'),
        'STORE ROOM':      ('ECEFF1', '37474F'),
        'WARDEN OFFICE':   ('FFF3E0', 'BF360C'),
        'LAUNDRY ROOM':    ('E8EAF6', '283593'),
    }

    # ── Sample data ──
    HOSTELS = [
        {
            'name': 'BH-1', 'type': 'boys', 'floors': 3,
            'floor_data': [
                {'lbl': 'F1', 'total': 20, 'desig': 'incampus',
                 'rows': [('5S+CW', 5), ('4S+CWR', 10), ('3S+WR+AC', 5)]},
                {'lbl': 'F2', 'total': 20, 'desig': 'incampus',
                 'rows': [('5S+WW', 8), ('4S+WR+AC', 7), ('3S+CWR+AC', 5)]},
                {'lbl': 'F3', 'total': 20, 'desig': 'outcampus',
                 'rows': [('6S+CW', 5), ('4S+CWR+AC', 10), ('2S+WR+AC', 5)]},
            ],
            'common': [
                ('GYM',           1),
                ('SPORTS ROOM',   1),
                ('ENTERTAINMENT', 1),
                ('MESS',          1),
                ('STUDY ROOM',    2),
                ('WARDEN OFFICE', 1),
                ('LAUNDRY ROOM',  1),
            ],
        },
    ]

    H_BG = ['F0F7FF', 'FFF8F0']
    F_BG = ['E3F2FD', 'FFF3E0']

    cur_row = 4

    for h_idx, hostel in enumerate(HOSTELS):
        h_start = cur_row
        h_bg    = hfill(H_BG[h_idx % 2])

        # ── Student room floors ──
        for f_idx, fd in enumerate(hostel['floor_data']):
            f_start = cur_row
            f_bg    = hfill(F_BG[f_idx % 2])

            for rt_idx, (rt_lbl, rt_cnt) in enumerate(fd['rows']):
                r = cur_row
                bg_hex, fc_hex = CAT_STYLES.get(rt_lbl, ('F5F5F5', '333333'))

                # Cols A–C: hostel info (value only on first row of hostel)
                for col in range(1, 4):
                    c = ws.cell(row=r, column=col)
                    c.fill = h_bg; c.border = thin; c.alignment = cc
                    if f_idx == 0 and rt_idx == 0:
                        c.value = [hostel['name'], hostel['type'], hostel['floors']][col - 1]
                        c.font  = Font(bold=True,
                                       color='0D47A1' if col == 1 else '333333',
                                       size=11 if col == 1 else 10)

                # Cols D–E: floor info (value only on first row of floor)
                for col in range(4, 6):
                    c = ws.cell(row=r, column=col)
                    c.fill = f_bg; c.border = thin; c.alignment = cc
                    if rt_idx == 0:
                        c.value = [fd['lbl'], fd['total']][col - 4]
                        c.font  = Font(bold=True, color='0F4C81', size=10)

                # Col F: Room Category
                f = ws.cell(row=r, column=6, value=rt_lbl)
                f.font = Font(bold=True, color=fc_hex, size=11)
                f.fill = hfill(bg_hex); f.alignment = cc; f.border = thin

                # Col G: Room Count
                g = ws.cell(row=r, column=7, value=rt_cnt)
                g.font = Font(bold=True, color='111111', size=13)
                g.fill = hfill(bg_hex); g.alignment = cc; g.border = thin

                # Col H: Designation (incampus/outcampus) — value only on first row of floor
                h = ws.cell(row=r, column=8)
                h.fill = hfill('F3E5F5'); h.border = thin; h.alignment = cc
                if rt_idx == 0:
                    h.value = fd.get('desig', '')
                    h.font  = Font(bold=True, color='6A1B9A', size=10)

                ws.row_dimensions[r].height = 22
                cur_row += 1

            # Merge floor cols D–E and H across all category rows of this floor
            f_end = cur_row - 1
            for col in [4, 5, 8]:
                ws.merge_cells(start_row=f_start, start_column=col,
                               end_row=f_end,     end_column=col)
                ws.cell(row=f_start, column=col).alignment = cc

        # ── Common / amenity facilities section ──
        if hostel.get('common'):
            # Sub-header row
            sub = ws.cell(row=cur_row, column=1, value='  Common Facilities')
            sub.font = Font(bold=True, italic=True, color='0D47A1', size=10)
            sub.fill = hfill('DBEAFE'); sub.alignment = lc
            for col in range(1, 9):
                ws.cell(row=cur_row, column=col).fill   = hfill('DBEAFE')
                ws.cell(row=cur_row, column=col).border = thin
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row,   end_column=8)
            ws.row_dimensions[cur_row].height = 17
            cur_row += 1

            for cat, cnt in hostel['common']:
                r = cur_row
                bg_hex, fc_hex = CAT_STYLES.get(cat, ('F5F5F5', '333333'))

                # Cols A–C: part of hostel merge (blank border only)
                for col in range(1, 4):
                    ws.cell(row=r, column=col).fill   = h_bg
                    ws.cell(row=r, column=col).border = thin

                # Col D: Floor = GF (each common room is its own floor entry)
                d = ws.cell(row=r, column=4, value='GF')
                d.font = Font(bold=True, color='0F4C81', size=10)
                d.fill = hfill('DBEAFE'); d.alignment = cc; d.border = thin

                # Col E: dash (not applicable for common rooms)
                e = ws.cell(row=r, column=5, value='—')
                e.font = Font(color='AAAAAA', size=10)
                e.fill = hfill('DBEAFE'); e.alignment = cc; e.border = thin

                # Col F: Category
                f = ws.cell(row=r, column=6, value=cat)
                f.font = Font(bold=True, color=fc_hex, size=11)
                f.fill = hfill(bg_hex); f.alignment = cc; f.border = thin

                # Col G: Count
                g = ws.cell(row=r, column=7, value=cnt)
                g.font = Font(bold=True, color='111111', size=12)
                g.fill = hfill(bg_hex); g.alignment = cc; g.border = thin

                # Col H: Designation — blank for common rooms
                h = ws.cell(row=r, column=8, value='')
                h.fill = hfill('F3E5F5'); h.alignment = cc; h.border = thin

                ws.row_dimensions[r].height = 20
                cur_row += 1

        # Merge hostel cols A–C across all rows of this hostel
        h_end = cur_row - 1
        for col in range(1, 4):
            ws.merge_cells(start_row=h_start, start_column=col,
                           end_row=h_end,     end_column=col)
            ws.cell(row=h_start, column=col).alignment = cc

        # Thin grey separator between hostels
        if h_idx < len(HOSTELS) - 1:
            for col in range(1, 9):
                s = ws.cell(row=cur_row, column=col)
                s.fill = hfill('B0BEC5'); s.border = thin
            ws.row_dimensions[cur_row].height = 4
            cur_row += 1

    # ── Sheet 2: Instructions ──
    wi = wb.create_sheet('Instructions')
    wi.column_dimensions['A'].width = 3
    wi.column_dimensions['B'].width = 26
    wi.column_dimensions['C'].width = 12
    wi.column_dimensions['D'].width = 62

    ti = wi.cell(row=1, column=1, value='  GGI Hostel ERP  —  Bulk Room Upload Guide')
    ti.font = Font(bold=True, color='FFFFFF', size=13)
    ti.fill = hfill('0D47A1')
    ti.alignment = Alignment(horizontal='left', vertical='center')
    wi.merge_cells('A1:D1'); wi.row_dimensions[1].height = 28

    sf  = PatternFill('solid', fgColor='E3F2FD')
    sfn = Font(bold=True, color='0D47A1', size=11)

    def isec(r, txt):
        c = wi.cell(row=r, column=1, value=txt)
        c.font = sfn; c.fill = sf
        c.alignment = Alignment(vertical='center', indent=1)
        wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        wi.row_dimensions[r].height = 22

    def irow(r, col, req, desc):
        wi.cell(row=r, column=2, value=col).font = Font(bold=True, size=10)
        rc = wi.cell(row=r, column=3, value=req)
        rc.font = Font(bold=True, size=10,
                       color='B71C1C' if req == 'Required' else '2E7D32')
        wi.cell(row=r, column=4, value=desc).font = Font(size=10)
        wi.row_dimensions[r].height = 18
 
    r = 3
    isec(r, '  COLUMN GUIDE'); r += 1
    hbg = PatternFill('solid', fgColor='BBDEFB')
    for col, val in [(2, 'Column'), (3, 'Required?'), (4, 'What to enter')]:
        hc = wi.cell(row=r, column=col, value=val)
        hc.font = Font(bold=True, size=10, color='0D47A1')
        hc.fill = hbg; hc.alignment = cc
    wi.row_dimensions[r].height = 18; r += 1

    for i, (cn, rq, ds) in enumerate([
        ('A — Hostel Name',      'Required', 'Name of the hostel block, e.g. BH-1, GH-2. Auto-created if it does not exist.'),
        ('B — Hostel Type',      'Required', '"boys", "girls", or "staff" (case-insensitive). Defaults to "boys" if invalid.'),
        ('C — Number of Floors', 'Optional', 'Total floors in this hostel — informational only, not enforced by the system.'),
        ('D — Floor',            'Required', 'Floor label: F1, F2 … or GF for Ground Floor. Auto-converts to a floor number.'),
        ('E — Rooms on Floor',   'Optional', 'Total rooms on this floor — informational only.'),
        ('F — Room Category',    'Required', 'Category code from the full list on this sheet. e.g. 5S+CW, 4S+CWR, 3S+WR+AC, GYM, WARDEN OFFICE.'),
        ('G — Room Count',       'Required', 'Number of rooms of this category to create. Room numbers are auto-generated.'),
        ('H — Designation',      'Optional', '"incampus" or "outcampus" for the floor. Leave blank for common/staff rooms.'),
    ]):
        irow(r, cn, rq, ds)
        if i % 2 == 0:
            for c in range(1, 5): wi.cell(row=r, column=c).fill = PatternFill('solid', fgColor='F5F5F5')
        r += 1

    r += 1
    isec(r, '  STUDENT ROOM CATEGORY CODES'); r += 1
    cat_hbg = PatternFill('solid', fgColor='E8F5E9')
    for col, val in [(2, 'Category Code'), (3, 'Capacity'), (4, 'Full Description')]:
        hc = wi.cell(row=r, column=col, value=val)
        hc.font = Font(bold=True, size=10, color='1B5E20')
        hc.fill = cat_hbg; hc.alignment = cc
    wi.row_dimensions[r].height = 18; r += 1

    for i, (code, cap, desc) in enumerate([
        ('6S+CW',     6, '6-Seater + Common Washroom'),
        ('5S+CW',     5, '5-Seater + Common Washroom'),
        ('5S+WW',     5, '5-Seater + With (Attached) Washroom'),
        ('5S+WW+AC',  5, '5-Seater + With Washroom + Air Conditioning'),
        ('4S+CWR',    4, '4-Seater + Common Washroom'),
        ('4S+CWR+AC', 4, '4-Seater + Common Washroom + Air Conditioning'),
        ('4S+WR',     4, '4-Seater + Attached Washroom'),
        ('4S+WR+AC',  4, '4-Seater + Attached Washroom + Air Conditioning'),
        ('3S+CWR',    3, '3-Seater + Common Washroom'),
        ('3S+CWR+AC', 3, '3-Seater + Common Washroom + Air Conditioning'),
        ('3S+WR',     3, '3-Seater + Attached Washroom'),
        ('3S+WR+AC',  3, '3-Seater + Attached Washroom + Air Conditioning'),
        ('2S+CWR',    2, '2-Seater + Common Washroom'),
        ('2S+CWR+AC', 2, '2-Seater + Common Washroom + Air Conditioning'),
        ('2S+WR',     2, '2-Seater + Attached Washroom'),
        ('2S+WR+AC',  2, '2-Seater + Attached Washroom + Air Conditioning'),
        ('1S+CWR',    1, 'Single Room + Common Washroom'),
        ('1S+CWR+AC', 1, 'Single Room + Common Washroom + Air Conditioning'),
        ('1S+WR',     1, 'Single Room + Attached Washroom'),
        ('1S+WR+AC',  1, 'Single Room + Attached Washroom + Air Conditioning'),
    ]):
        wi.cell(row=r, column=2, value=code).font = Font(bold=True, size=10)
        wi.cell(row=r, column=3, value=str(cap)).font = Font(bold=True, size=10, color='0D47A1')
        wi.cell(row=r, column=4, value=desc).font = Font(size=10)
        if i % 2 == 0:
            for c in range(1, 5): wi.cell(row=r, column=c).fill = PatternFill('solid', fgColor='F5F5F5')
        wi.row_dimensions[r].height = 17; r += 1

    r += 1
    isec(r, '  COMMON / AMENITY ROOM CODES  (no beds, capacity = 0)'); r += 1
    fac_hbg = PatternFill('solid', fgColor='FFF8E1')
    for col, val in [(2, 'Category Code'), (3, 'Capacity'), (4, 'Notes / Example')]:
        hc = wi.cell(row=r, column=col, value=val)
        hc.font = Font(bold=True, size=10, color='E65100')
        hc.fill = fac_hbg; hc.alignment = cc
    wi.row_dimensions[r].height = 18; r += 1

    for i, (code, desc) in enumerate([
        ('GYM',             'Gym / Fitness Room  —  room number auto-generated as PREFIX-GYM-GF'),
        ('SPORTS ROOM',     'Sports Room  —  auto-generated as PREFIX-SPR-GF'),
        ('ENTERTAINMENT',   'Entertainment / TV Room  —  auto-generated as PREFIX-ENT-GF'),
        ('MESS',            'Mess Hall / Dining  —  auto-generated as PREFIX-MESS-GF'),
        ('STUDY ROOM',      'Study Room  —  multiple rooms auto-numbered: PREFIX-STR-GF, PREFIX-STR-GF-2 …'),
        ('SITTING ARENA',   'Sitting Area / Lounge  —  auto-generated as PREFIX-SIT-GF'),
        ('STORE ROOM',      'Store Room / Storage  —  auto-generated as PREFIX-STO-GF'),
        ('WARDEN OFFICE',   'Warden\'s Office  —  auto-generated as PREFIX-WO-GF'),
        ('LAUNDRY ROOM',    'Laundry Room  —  auto-generated as PREFIX-LDY-GF'),
    ]):
        wi.cell(row=r, column=2, value=code).font = Font(bold=True, size=10)
        wi.cell(row=r, column=3, value='0').font = Font(bold=True, size=10, color='6A1B9A')
        wi.cell(row=r, column=4, value=desc).font = Font(size=10)
        if i % 2 == 0:
            for c in range(1, 5): wi.cell(row=r, column=c).fill = PatternFill('solid', fgColor='FFF8E1')
        wi.row_dimensions[r].height = 17; r += 1

    r += 1
    isec(r, '  KEY RULES'); r += 1
    for clr, rule in [
        ('B71C1C', 'ROOM NUMBERS ARE AUTO-GENERATED: no start/end columns needed. Rooms are numbered as PREFIX-F1-001, PREFIX-F1-002 …'),
        ('B71C1C', 'Floor numbering: GF = 001–099, F1 = 101–199, F2 = 201–299, etc. Numbers continue sequentially across categories.'),
        ('1B5E20', 'COMMON ROOMS (GYM, MESS, WARDEN OFFICE, LAUNDRY ROOM…): each row gets Floor=GF and a Count. Room names are auto-generated.'),
        ('1B5E20', 'Multiple common rooms of the same type on the same floor are auto-suffixed: PREFIX-STR-GF, PREFIX-STR-GF-2, etc.'),
        ('1565C0', 'DESIGNATION (col H): enter "incampus" or "outcampus" once per floor row. Leave blank for common/staff rooms.'),
        ('1565C0', 'Zero counts are valid — a category row with Count=0 is simply skipped.'),
        ('1565C0', 'Beds are auto-created for student rooms (e.g. 4S+CWR → 4 beds). No beds created for capacity-0 common rooms.'),
    ]:
        wi.cell(row=r, column=2, value='▶').font = Font(bold=True, size=11, color=clr)
        wi.cell(row=r, column=4, value=rule).font = Font(size=10)
        wi.row_dimensions[r].height = 18; r += 1

    r += 1
    isec(r, '  HOW TO USE'); r += 1
    for step, desc in [
        ('Step 1', 'Fill Hostel Name (A) and Hostel Type (B) once — they are merged across all rows of that hostel.'),
        ('Step 2', 'For each student-room floor, add one row per category. Enter Floor (D) only in the first row (merged).'),
        ('Step 3', 'Enter Designation (H) once per floor — "incampus" or "outcampus". Leave blank for common or staff rooms.'),
        ('Step 4', 'Fill Category (F) and Count (G) for each row. Room numbers are auto-generated — no start/end needed.'),
        ('Step 5', 'For common rooms (Gym, Mess, Warden Office, Laundry…): add one row each with Floor=GF, Category, Count.'),
        ('Step 6', 'To add another hostel: leave a blank separator row, then start the next hostel block below it.'),
        ('Step 7', 'Delete sample data, fill in your own data, save as .xlsx. Upload via Hostels → Bulk Upload.'),
        ('Step 8', 'Review the result — created hostels, rooms, beds, and any skipped rows are shown after upload.'),
    ]:
        wi.cell(row=r, column=2, value=step).font = Font(bold=True, size=10, color='1565C0')
        wi.cell(row=r, column=4, value=desc).font = Font(size=10)
        wi.row_dimensions[r].height = 20; r += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="hostel_rooms_bulk_upload_sample.xlsx"'
    )
    wb.save(response)
    return response


@admin_only
def bulk_upload_rooms(request):  # noqa: C901
    """Upload + process hierarchical floor-wise Excel (new category system, no Wing column)."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb['Rooms'] if 'Rooms' in wb.sheetnames else wb.active
        except Exception as e:
            messages.error(request, f'Could not read file: {e}')
            return redirect('bulk_upload_rooms')

        import re as _re

        # Column layout (0-based):
        # A=0 Hostel Name, B=1 Hostel Type, C=2 No. of Floors,
        # D=3 Floor, E=4 Rooms on Floor,
        # F=5 Room Category, G=6 Room Count,
        # H=7 Designation (incampus/outcampus — optional)
        RT_MAP = {
            # 6-Seater
            '6S+CW':              ('6-Seater (Common WR)',        6),
            # 5-Seater variants
            '5S+CW':              ('5-Seater (Common WR)',        5),
            '5S+WW':              ('5-Seater (With WR)',          5),
            '5S+WW+AC':           ('5-Seater (With WR, AC)',      5),
            # 4-Seater variants
            '4S+CWR':             ('4-Seater (Common WR)',        4),
            '4S+CWR+AC':          ('4-Seater (Common WR, AC)',    4),
            '4S+WR':              ('4-Seater (Attached WR)',      4),
            '4S+WR+AC':           ('4-Seater (Attached WR, AC)', 4),
            # 3-Seater variants
            '3S+CWR':             ('3-Seater (Common WR)',        3),
            '3S+CWR+AC':          ('3-Seater (Common WR, AC)',    3),
            '3S+WR':              ('3-Seater (Attached WR)',      3),
            '3S+WR+AC':           ('3-Seater (Attached WR, AC)', 3),
            # 2-Seater variants
            '2S+CWR':             ('2-Seater (Common WR)',        2),
            '2S+CWR+AC':          ('2-Seater (Common WR, AC)',    2),
            '2S+WR':              ('2-Seater (Attached WR)',      2),
            '2S+WR+AC':           ('2-Seater (Attached WR, AC)', 2),
            # 1-Seater (single room) variants
            '1S+CWR':             ('Single (Common WR)',          1),
            '1S+CWR+AC':          ('Single (Common WR, AC)',      1),
            '1S+WR':              ('Single (Attached WR)',        1),
            '1S+WR+AC':           ('Single (Attached WR, AC)',   1),
            # Common / amenity rooms — capacity 0, no beds created
            'GYM':                ('Gym',               0),
            'SPORTS ROOM':        ('Sports Room',       0),
            'SPORTS':             ('Sports Room',       0),
            'ENTERTAINMENT ROOM': ('Entertainment Room', 0),
            'ENTERTAINMENT':      ('Entertainment Room', 0),
            'ENT':                ('Entertainment Room', 0),
            'MESS':               ('Mess Hall',          0),
            'MESS HALL':          ('Mess Hall',          0),
            'STUDY ROOM':         ('Study Room',         0),
            'STUDY':              ('Study Room',         0),
            'SITTING ARENA':      ('Sitting Arena',      0),
            'SITTING':            ('Sitting Arena',      0),
            'STORE ROOM':         ('Store Room',         0),
            'STORE':              ('Store Room',         0),
            'STOREROOM':          ('Store Room',         0),
            'WARDEN OFFICE':      ('Warden Office',      0),
            'WARDEN':             ('Warden Office',      0),
            'LAUNDRY ROOM':       ('Laundry Room',       0),
            'LAUNDRY':            ('Laundry Room',       0),
        }

        # Short codes for common rooms used in auto-generated room numbers
        COMMON_SHORT = {
            'Gym': 'GYM', 'Sports Room': 'SPR', 'Entertainment Room': 'ENT',
            'Mess Hall': 'MESS', 'Study Room': 'STR', 'Sitting Arena': 'SIT',
            'Store Room': 'STO', 'Warden Office': 'WO', 'Laundry Room': 'LDY',
        }

        def floor_label_to_int(label):
            lbl = str(label).strip().upper()
            if lbl in ('GF', 'G', 'GROUND', 'GROUND FLOOR', 'B', 'B1', 'BASEMENT'):
                return 0
            nums = _re.sub(r'[^0-9]', '', lbl)
            return int(nums) if nums else 1

        def hostel_prefix(name):
            """Derive a short prefix from the hostel name for room numbering."""
            return _re.sub(r'\s+', '', str(name).upper())

        # Rows 1–3 are title / group-headers / column-headers; data starts at row 4.
        cur_hostel_name  = None
        cur_hostel_type  = 'boys'
        cur_floor_lbl    = None
        cur_designation  = ''
        cur_floor_row    = None

        floor_counts = []   # list of (rt_label, count) — flexible number of rows per floor
        pending      = []
        skipped_rows = []

        def flush_floor():
            nonlocal floor_counts
            if cur_floor_lbl and floor_counts:
                pending.append({
                    'hostel': cur_hostel_name, 'type': cur_hostel_type,
                    'floor': cur_floor_lbl,
                    'designation': cur_designation,
                    'counts': floor_counts[:], 'row': cur_floor_row,
                })
            floor_counts = []

        for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            vals = list(row) + [None] * 10

            if not any(v is not None for v in vals[:8]):
                continue

            # Hostel context — col A (0)
            if vals[0] is not None:
                flush_floor()
                cur_hostel_name = str(vals[0]).strip()

            if vals[1] is not None:
                cur_hostel_type = str(vals[1]).strip().lower()
                if cur_hostel_type not in ('boys', 'girls', 'staff'):
                    cur_hostel_type = 'boys'

            # Floor context — col D (3)
            if vals[3] is not None:
                flush_floor()
                cur_floor_lbl   = str(vals[3]).strip()
                cur_designation = str(vals[7]).strip().lower() if vals[7] else ''
                if cur_designation not in ('incampus', 'outcampus'):
                    cur_designation = ''
                cur_floor_row  = row_num
                floor_counts   = []

            # Room category + count — col F (5) and col G (6)
            if vals[5] is not None and vals[6] is not None and cur_floor_lbl:
                rt_lbl = str(vals[5]).strip().upper()
                try:
                    cnt = int(vals[6])
                except (TypeError, ValueError):
                    cnt = 0
                floor_counts.append((rt_lbl, cnt))

        flush_floor()  # flush the last floor

        # ── Create rooms from pending floor entries ──
        created_hostels  = 0
        created_rooms    = 0
        created_beds     = 0
        hostel_cache     = {}
        # Track per-hostel per-floor room sequence for auto-numbering
        floor_seq_cache  = {}   # (hostel_name, floor_int) → next_seq

        for entry in pending:
            row_num = entry['row']

            if not entry['hostel']:
                skipped_rows.append({'row': row_num, 'reason': 'Hostel Name missing',
                                     'data': str(entry.get('floor', ''))})
                continue

            floor_num = floor_label_to_int(entry['floor'])
            hn = entry['hostel']
            prefix = hostel_prefix(hn)

            if hn not in hostel_cache:
                hostel, h_created = Hostel.objects.get_or_create(
                    name=hn, defaults={'type': entry['type']}
                )
                if h_created:
                    created_hostels += 1
                hostel_cache[hn] = hostel
            hostel = hostel_cache[hn]

            # Determine starting sequence for this floor
            seq_key = (hn, floor_num)
            if seq_key not in floor_seq_cache:
                # Start from floor * 100 + 1 (GF = 001, F1 = 101, F2 = 201, …)
                floor_seq_cache[seq_key] = floor_num * 100 + 1

            common_type_seq = {}  # track count per common-room type for unique naming

            for rt_lbl, count in entry['counts']:
                if count == 0:
                    continue
                type_name, capacity = RT_MAP.get(rt_lbl, ('Unknown', 2))

                for _ in range(count):
                    if capacity == 0:
                        # Common room: use category-based name
                        short = COMMON_SHORT.get(type_name, rt_lbl[:3])
                        common_type_seq[short] = common_type_seq.get(short, 0) + 1
                        suffix = '' if common_type_seq[short] == 1 else f'-{common_type_seq[short]}'
                        floor_pfx = 'GF' if floor_num == 0 else f'F{floor_num}'
                        room_number = f'{prefix}-{short}-{floor_pfx}{suffix}'
                    else:
                        seq = floor_seq_cache[seq_key]
                        room_number = f'{prefix}-{seq:03d}'
                        floor_seq_cache[seq_key] += 1

                    if Room.objects.filter(hostel=hostel, room_number=room_number).exists():
                        skipped_rows.append({'row': row_num, 'reason': 'Room already exists',
                                             'data': f'{hn} / {room_number}'})
                    else:
                        room = Room.objects.create(
                            hostel=hostel, room_number=room_number,
                            floor=floor_num, room_type=type_name,
                            capacity=capacity, status='vacant',
                            designation=entry.get('designation', ''),
                        )
                        # Auto-create beds for student rooms only
                        for bed_n in range(1, capacity + 1):
                            Bed.objects.create(room=room, bed_number=str(bed_n),
                                               status='available')
                            created_beds += 1
                        created_rooms += 1

        if created_rooms or created_hostels:
            msg = f'Upload complete: {created_rooms} room(s) and {created_beds} bed(s) created'
            if created_hostels:
                msg += f', {created_hostels} new hostel(s) created'
            messages.success(request, msg + '.')
        else:
            messages.warning(request, 'No new rooms were created.')

        return render(request, 'admin/bulk_upload_rooms.html', {
            'done': True,
            'created_rooms':   created_rooms,
            'created_hostels': created_hostels,
            'created_beds':    created_beds,
            'skipped_rows':    skipped_rows,
        })

    return render(request, 'admin/bulk_upload_rooms.html', {'done': False})

# ═══════════════════════════════════════════════════════════════════════════════
# QR GATE PASS MODULE
# ═══════════════════════════════════════════════════════════════════════════════

def _gp_security_required(view_fn):
    """Decorator: admin, warden, or security guard can access."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.can_manage_gate_pass:
            messages.error(request, 'Access denied.')
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


def _generate_qr_image(token):
    """Return a base64-encoded PNG of the QR code for the given token."""
    import qrcode as _qrcode
    qr = _qrcode.QRCode(
        version=None,
        error_correction=_qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=4,
    )
    qr.add_data(token)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


def _make_exit_token(gp):
    return signing.dumps(
        {'gp': str(gp.pk), 'type': 'exit', 'nonce': str(uuid.uuid4())},
        salt='gatepass-exit'
    )


def _make_entry_token(gp):
    return signing.dumps(
        {'gp': str(gp.pk), 'type': 'entry', 'nonce': str(uuid.uuid4())},
        salt='gatepass-entry'
    )


# ─── Student: Generate Exit QR ───────────────────────────────────────────────

@login_required
def generate_exit_qr(request, pk):
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir
    gp = get_object_or_404(GatePass, pk=pk, student=student)
    if gp.status != GatePass.Status.APPROVED:
        messages.error(request, 'Gate pass must be approved before generating exit QR.')
        return redirect('my_gate_passes')  
    if gp.exit_time:
        messages.error(request, 'Exit already recorded for this gate pass.')
        return redirect('my_gate_passes')
    # Guard must have permitted exit
    if not gp.guard_exit_permitted or not gp.guard_exit_permitted_at:
        messages.error(request, 'Security guard has not permitted your exit yet. Please go to the gate first.')
        return redirect('my_gate_passes')
    # Permission window: 5 minutes
    elapsed = (timezone.now() - gp.guard_exit_permitted_at).total_seconds()
    if elapsed > 300:
        messages.error(request, 'Exit permission expired. Ask the security guard to permit again.')
        return redirect('my_gate_passes')
    token = _make_exit_token(gp)
    gp.exit_qr_token        = token
    gp.exit_qr_generated_at = timezone.now()  
    gp.exit_qr_scanned      = False 
    gp.save(update_fields=['exit_qr_token', 'exit_qr_generated_at', 'exit_qr_scanned'])
    response = render(request, 'student/gate_pass_qr.html', {
        'gp': gp, 'qr_type': 'exit', 'token': token,
        'qr_image': _generate_qr_image(token),
        'title': 'Exit QR Code',
        'subtitle': 'Show this to the security guard while leaving.',
        'warning': 'Expires in 5 minutes. Single use only.',
        'expires_in': 300,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


# ─── Student: Generate Entry QR ──────────────────────────────────────────────

@login_required
def generate_entry_qr(request, pk):
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir
    gp = get_object_or_404(GatePass, pk=pk, student=student)
    if not gp.exit_time:
        messages.error(request, 'Exit has not been recorded yet.')
        return redirect('my_gate_passes')
    if gp.entry_time:
        messages.error(request, 'Entry already recorded for this gate pass.')
        return redirect('my_gate_passes')
    # Guard must have permitted entry
    if not gp.guard_entry_permitted or not gp.guard_entry_permitted_at:
        messages.error(request, 'Security guard has not permitted your entry yet. Please go to the gate first.')
        return redirect('my_gate_passes')
    # Permission window: 5 minutes
    elapsed = (timezone.now() - gp.guard_entry_permitted_at).total_seconds()
    if elapsed > 300:
        messages.error(request, 'Entry permission expired. Ask the security guard to permit again.')
        return redirect('my_gate_passes')      
    token = _make_entry_token(gp)
    gp.entry_qr_token        = token
    gp.entry_qr_generated_at = timezone.now()
    gp.entry_qr_scanned      = False
    gp.save(update_fields=['entry_qr_token', 'entry_qr_generated_at', 'entry_qr_scanned'])
    response = render(request, 'student/gate_pass_qr.html', {
        'gp': gp, 'qr_type': 'entry', 'token': token,
        'qr_image': _generate_qr_image(token),
        'title': 'Entry QR Code',
        'subtitle': 'Show this to the security guard when returning to hostel.',
        'warning': 'Expires in 5 minutes. Single use only.',
        'expires_in': 300,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


# ─── Security Dashboard ───────────────────────────────────────────────────────

@_gp_security_required
def security_dashboard(request):
    import datetime
    warden_hostel = get_hostel_scope(request.user)
    qs = GatePass.objects.select_related('student', 'approved_by').filter(
        exit_time__isnull=False, entry_time__isnull=True
    )
    if warden_hostel:
        qs = qs.filter(
            student__allocations__room__hostel=warden_hostel,
            student__allocations__status='active'
        ).distinct()

    now = timezone.now()

    def _base_qs():
        q = GatePass.objects.select_related('student', 'approved_by', 'guard_exit_permitted_by')
        if warden_hostel:
            q = q.filter(
                student__allocations__room__hostel=warden_hostel,
                student__allocations__status='active'
            ).distinct()
        return q

    # 1. Approved, no exit yet, guard NOT yet permitted exit
    awaiting_exit = _base_qs().filter(
        status=GatePass.Status.APPROVED,
        exit_time__isnull=True,
        guard_exit_permitted=False,
    )

    # 2. Exit permitted by guard but QR not scanned yet (within 10-min window)
    exit_permitted_raw = _base_qs().filter(
        status=GatePass.Status.APPROVED,
        exit_time__isnull=True,
        guard_exit_permitted=True,
    )
    exit_permitted = []
    for gp in exit_permitted_raw:
        secs_left = max(0, 600 - int((now - gp.guard_exit_permitted_at).total_seconds()))
        exit_permitted.append({'gp': gp, 'expires_in': secs_left})

    # 3. Currently outside — split into on-time and overdue
    outside_qs = _base_qs().filter(exit_time__isnull=False, entry_time__isnull=True)
    currently_outside = []
    late_returns = []
    for gp in outside_qs:
        try:
            expected_dt = timezone.make_aware(
                datetime.datetime.combine(gp.expected_return_date, gp.expected_return_time)
            )
        except Exception:
            expected_dt = None
        is_late = expected_dt and now > expected_dt
        hours_out = round((now - gp.exit_time).total_seconds() / 3600, 1)
        entry = {'gp': gp, 'expected_dt': expected_dt, 'hours_out': hours_out}
        (late_returns if is_late else currently_outside).append(entry)

    today = dt_date.today()
    today_log = _base_qs().filter(exit_time__date=today).order_by('-exit_time')

    guards = User.objects.filter(role=User.Role.SECURITY, is_active=True)

    return render(request, 'security/dashboard.html', {
        'awaiting_exit':     awaiting_exit,
        'exit_permitted':    exit_permitted,
        'currently_outside': currently_outside,
        'late_returns':      late_returns,
        'awaiting_entry':    outside_qs.filter(guard_entry_permitted=False),
        'today_log':         today_log,
        'now':               now,
        'warden_hostel':     warden_hostel,
        'guards':            guards, 
        'is_admin':          request.user.role == User.Role.ADMIN,
    })


# ─── Scan Interface ───────────────────────────────────────────────────────────

@_gp_security_required
def scan_qr(request):
    if request.method == 'POST':
        token = request.POST.get('token', '').strip()
        if not token:
            messages.error(request, 'No token provided.')
            return redirect('scan_qr')
        request.session['qr_token'] = token
        return redirect('verify_qr')
    return render(request, 'security/scan.html')


# ─── Verify QR ───────────────────────────────────────────────────────────────

@_gp_security_required
def verify_qr(request):
    # Token can come from session (manual entry) or POST (camera form)
    token = (
        request.session.pop('qr_token', None)
        or request.POST.get('token', '')
        or request.GET.get('token', '')
    ).strip()
    if not token:
        messages.error(request, 'No token.')
        return redirect('scan_qr')

    gp = None
    qr_type = None
    error_msg = None

    # Try exit
    try:
        data = signing.loads(token, salt='gatepass-exit', max_age=300)
        if data.get('type') == 'exit':
            candidate = GatePass.objects.select_related('student', 'approved_by').filter(pk=data['gp']).first()
            if candidate:
                if candidate.exit_qr_token != token:
                    error_msg = 'This QR is outdated. Ask the student to regenerate.'
                elif candidate.exit_qr_scanned:
                    error_msg = 'This exit QR has already been used (one-time only).'
                elif candidate.exit_time:
                    error_msg = 'Exit already recorded for this gate pass.'
                else:
                    gp = candidate
                    qr_type = 'exit'
    except signing.SignatureExpired:
        error_msg = 'Exit QR has expired (valid 5 minutes). Ask the student to generate a new one.'
    except signing.BadSignature:
        pass

    # Try entry
    if gp is None and error_msg is None:
        try:
            data = signing.loads(token, salt='gatepass-entry', max_age=300)
            if data.get('type') == 'entry':
                candidate = GatePass.objects.select_related('student', 'approved_by').filter(pk=data['gp']).first()
                if candidate:
                    if candidate.entry_qr_token != token:
                        error_msg = 'This QR is outdated. Ask the student to regenerate.'
                    elif candidate.entry_qr_scanned:
                        error_msg = 'This entry QR has already been used (one-time only).'
                    elif candidate.entry_time:
                        error_msg = 'Entry already recorded for this gate pass.'
                    elif not candidate.exit_time:
                        error_msg = 'Exit has not been recorded for this gate pass.'
                    else:
                        gp = candidate
                        qr_type = 'entry'
        except signing.SignatureExpired:
            error_msg = 'Entry QR has expired (valid 5 minutes). Ask the student to generate a new one.'
        except signing.BadSignature:
            pass

    if error_msg:
        return render(request, 'security/verify.html', {'error': error_msg, 'token': token})
    if gp is None:
        return render(request, 'security/verify.html', {
            'error': 'Invalid QR code. Token could not be verified.', 'token': token
        })

    allocation = gp.student.allocations.filter(status='active').select_related('room__hostel').first()
    return render(request, 'security/verify.html', {
        'gp': gp, 'qr_type': qr_type, 'token': token,
        'allocation': allocation, 'error': None,
    })
 

# ─── Guard Permit Exit (before QR generation) ────────────────────────────────

@_gp_security_required
@require_POST
def guard_permit_exit(request):
    pk = request.POST.get('pk', '').strip()
    gp = get_object_or_404(GatePass, pk=pk)
    if gp.status != GatePass.Status.APPROVED or gp.exit_time:
        messages.error(request, 'Gate pass is not eligible for exit permission.')
        return redirect('security_dashboard')
    now = timezone.now()
    gp.guard_exit_permitted    = True
    gp.guard_exit_permitted_at = now
    gp.guard_exit_permitted_by = request.user
    gp.save(update_fields=['guard_exit_permitted', 'guard_exit_permitted_at', 'guard_exit_permitted_by'])
    send_notification(
        [gp.student.user],
        'Exit Permitted — Generate Your Exit QR Now',
        f'Security guard has permitted your exit. Open "My Gate Passes" and generate your Exit QR immediately — it expires in 5 minutes.',
        notif_type='gate_pass',
    )
    messages.success(request, f'Exit permitted for {gp.student.name}. Student has been notified.')
    return redirect('security_dashboard')


# ─── Guard Permit Entry (before QR generation) ───────────────────────────────

@_gp_security_required
@require_POST
def guard_permit_entry(request):
    pk = request.POST.get('pk', '').strip()
    gp = get_object_or_404(GatePass, pk=pk)
    if not gp.exit_time or gp.entry_time:
        messages.error(request, 'Gate pass is not eligible for entry permission.')
        return redirect('security_dashboard')
    now = timezone.now()
    gp.guard_entry_permitted    = True
    gp.guard_entry_permitted_at = now
    gp.guard_entry_permitted_by = request.user
    gp.save(update_fields=['guard_entry_permitted', 'guard_entry_permitted_at', 'guard_entry_permitted_by'])
    send_notification(
        [gp.student.user],
        'Entry Permitted — Generate Your Entry QR Now',
        f'Security guard has permitted your entry. Open "My Gate Passes" and generate your Entry QR immediately — it expires in 5 minutes.',
        notif_type='gate_pass',
    )
    messages.success(request, f'Entry permitted for {gp.student.name}. Student has been notified.')
    return redirect('security_dashboard')


# ─── Guard Allow Exit ─────────────────────────────────────────────────────────

@_gp_security_required
@require_POST
def guard_allow_exit(request):
    token = request.POST.get('token', '').strip()
    try:
        data = signing.loads(token, salt='gatepass-exit', max_age=300)
    except signing.SignatureExpired:
        messages.error(request, 'Exit QR expired (valid 5 minutes). Ask student to generate a new one.')
        return redirect('scan_qr')
    except signing.BadSignature:
        messages.error(request, 'Invalid or tampered token.')
        return redirect('scan_qr')

    gp = get_object_or_404(GatePass, pk=data['gp'])
    if gp.exit_qr_token != token or gp.exit_qr_scanned or gp.exit_time:
        messages.error(request, 'QR already used or invalid.')
        return redirect('scan_qr')

    now       = timezone.now()
    local_now = timezone.localtime(now)
    gp.exit_qr_scanned = True
    gp.exit_time       = now
    gp.exit_allowed_by = request.user
    gp.save(update_fields=['exit_qr_scanned', 'exit_time', 'exit_allowed_by'])

    send_notification(
        [gp.student.user],
        'Gate Pass — Exit Recorded',
        f'Your exit has been recorded at {local_now.strftime("%d %b %Y, %I:%M %p")}. '
        f'Please return by {gp.expected_return_date.strftime("%d %b")} {gp.expected_return_time.strftime("%I:%M %p")}.',
        notif_type='gate_pass',
    )
    for u in User.objects.filter(role__in=[User.Role.WARDEN, User.Role.ADMIN], is_active=True):
        send_notification(
            [u],
            f'Exit — {gp.student.name}',
            f'{gp.student.name} ({gp.student.roll_number}) exited at {local_now.strftime("%I:%M %p")}.',
            notif_type='gate_pass',
        )
    messages.success(request, f'Exit recorded for {gp.student.name} at {local_now.strftime("%I:%M %p")}.')
    return redirect('security_dashboard')


# ─── Guard Mark Entry ─────────────────────────────────────────────────────────

@_gp_security_required
@require_POST
def guard_mark_entry(request):
    token = request.POST.get('token', '').strip()
    try:
        data = signing.loads(token, salt='gatepass-entry', max_age=300)
    except signing.SignatureExpired:
        messages.error(request, 'Entry QR expired. Ask student to regenerate.')
        return redirect('scan_qr')
    except signing.BadSignature:
        messages.error(request, 'Invalid or tampered token.')
        return redirect('scan_qr')

    gp = get_object_or_404(GatePass, pk=data['gp'])
    if gp.entry_qr_token != token or gp.entry_qr_scanned or gp.entry_time or not gp.exit_time:
        messages.error(request, 'QR already used, invalid, or exit not recorded.')
        return redirect('scan_qr')

    now       = timezone.now()
    local_now = timezone.localtime(now)
    gp.entry_qr_scanned   = True
    gp.entry_time         = now
    gp.entry_allowed_by   = request.user
    gp.actual_return_date = local_now.date()
    gp.actual_return_time = local_now.time()
    gp.status             = GatePass.Status.RETURNED
    gp.is_overstayed      = False
    gp.save(update_fields=[
        'entry_qr_scanned', 'entry_time', 'entry_allowed_by',
        'actual_return_date', 'actual_return_time', 'status', 'is_overstayed'
    ])

    send_notification(
        [gp.student.user],
        'Gate Pass — Entry Recorded',
        f'Welcome back! Your entry has been logged at {local_now.strftime("%d %b %Y, %I:%M %p")}.',
        notif_type='gate_pass',
    )
    for u in User.objects.filter(role__in=[User.Role.WARDEN, User.Role.ADMIN], is_active=True):
        send_notification(
            [u],
            f'Return — {gp.student.name}',
            f'{gp.student.name} ({gp.student.roll_number}) returned at {local_now.strftime("%I:%M %p")}.',
            notif_type='gate_pass',
        )
    messages.success(request, f'Entry recorded for {gp.student.name} at {local_now.strftime("%I:%M %p")}.')
    return redirect('security_dashboard')

   
# ─── Admin: Create Security Guard ─────────────────────────────────────────────

@admin_only
@require_POST
def create_security_guard(request):
    name     = request.POST.get('name', '').strip()
    email    = request.POST.get('email', '').strip()
    phone    = request.POST.get('phone', '').strip()
    password = request.POST.get('password', '')
    if not name or not email or not password:
        messages.error(request, 'Name, email and password are required.')
        return redirect('security_dashboard')
    if User.objects.filter(email=email).exists():
        messages.error(request, f'A user with email {email} already exists.')
        return redirect('security_dashboard')
    User.objects.create_user(email=email, password=password, name=name, phone=phone, role=User.Role.SECURITY)
    messages.success(request, f'Security guard account created for {name}.')
    return redirect('security_dashboard')


# ─── Security: Announcements ──────────────────────────────────────────────────

def _security_required(view_fn):
    """Decorator: only security guards (not admin/warden) can access."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role != User.Role.SECURITY:
            messages.error(request, 'Access denied.')
            return redirect('dashboard')
        return view_fn(request, *args, **kwargs)
    wrapper.__name__ = view_fn.__name__
    return wrapper


@_security_required
def security_announcements(request):
    today = dt_date.today()
    announcements = (
        Announcement.objects
        .filter(is_active=True, audience__in=['all', 'staff_all', 'security'])
        .filter(Q(expires_at__isnull=True) | Q(expires_at__gte=today))
        .select_related('category', 'created_by')
        .order_by('-created_at')
    )
    return render(request, 'security/announcements.html', {'announcements': announcements})


# ─── Leave Applications ───────────────────────────────────────────────────────

def _get_active_restriction(student, rtype):
    """Return the first active restriction blocking rtype ('gate_pass' or 'leave'), or None."""
    return StudentRestriction.objects.filter(
        student=student, is_active=True,
        restriction_type__in=[rtype, StudentRestriction.RestrictionType.BOTH]
    ).select_related('restricted_by').first()


def _get_quota_status(student, ptype):
    """
    Check quota policy for the student.
    Returns dict: {blocked, used, limit, period, policy} or None if no active policy.
    ptype = 'gate_pass' or 'leave'
    """
    from datetime import date as _date
    import datetime

    # Specific student policy takes priority over global
    policy = QuotaPolicy.objects.filter(
        is_active=True,
        policy_type__in=[ptype, 'both'],
        applies_to_all=False,
        student=student,
    ).order_by('-created_at').first()

    if not policy:
        policy = QuotaPolicy.objects.filter(
            is_active=True,
            policy_type__in=[ptype, 'both'],
            applies_to_all=True,
        ).order_by('-created_at').first()

    if not policy:
        return None

    today = _date.today()
    if policy.period == 'weekly':
        # ISO week start (Monday)
        start = today - datetime.timedelta(days=today.weekday())
        end   = start + datetime.timedelta(days=6)
    elif policy.period == 'monthly':
        start = today.replace(day=1)
        end   = (today.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)
    else:  # yearly
        start = today.replace(month=1, day=1)
        end   = today.replace(month=12, day=31)

    if ptype == 'gate_pass':
        used = GatePass.objects.filter(
            student=student,
            created_at__date__gte=start,
            created_at__date__lte=end,
        ).exclude(status='rejected').count()
    else:
        used = LeaveApplication.objects.filter(
            student=student,
            created_at__date__gte=start,
            created_at__date__lte=end,
        ).exclude(status='rejected').count()

    return {
        'policy':  policy,
        'used':    used,
        'limit':   policy.limit,
        'period':  policy.get_period_display(),
        'blocked': used >= policy.limit,
        'remaining': max(0, policy.limit - used),
    }


def _is_working_day(date_val):
    """True if date_val is a working day per the academic calendar.
    Falls back to Mon–Fri if no calendar entry exists."""
    from datetime import date as date_type
    import datetime
    if isinstance(date_val, str):
        date_val = datetime.date.fromisoformat(date_val)
    event = AcademicCalendarEvent.objects.filter(date=date_val).first()
    if event:
        return not event.is_non_working
    return date_val.weekday() < 5  # Mon=0 … Fri=4


def _get_hod(department_name):
    """Return the HOD User for the given department name, or None."""
    from apps.accounts.models import User as _User
    return _User.objects.filter(
        role=_User.Role.HOD,
        department__iexact=department_name.strip()
    ).first()


@login_required
def apply_leave(request):
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir
    restriction = _get_active_restriction(student, 'leave')
    if restriction:
        return render(request, 'student/restricted.html', {
            'restriction': restriction, 'action': 'apply for leave'
        })
    quota = _get_quota_status(student, 'leave')
    if quota and quota['blocked']:
        return render(request, 'student/quota_exceeded.html', {
            'quota': quota, 'action': 'leave', 'back_url': 'my_leaves'
        })
    allocation = RoomAllocation.objects.filter(student=student, status='active').select_related('room__hostel').first()
    categories = LeaveCategory.objects.filter(is_active=True)

    def _ctx(extra=None):
        base = {
            'leave_categories': categories, 'student': student,
            'allocation': allocation, 'quota': quota,
        }
        if extra:
            base.update(extra)
        return base

    if request.method == 'POST':
        from_date   = request.POST.get('from_date')
        to_date     = request.POST.get('to_date')
        reason      = request.POST.get('reason', '').strip()
        category_id = request.POST.get('category') or None
        if not from_date or not to_date or not reason:
            messages.error(request, 'All fields are required.')
            return render(request, 'student/apply_leave.html', _ctx())

        working_day = _is_working_day(from_date)
        leave = LeaveApplication.objects.create(
            student=student,
            category_id=category_id,
            from_date=from_date,
            to_date=to_date,
            reason=reason,
            is_working_day_leave=working_day,
        )

        if working_day:
            # Route to HOD of student's department first
            hod = _get_hod(student.department)
            if hod:
                send_notification(
                    [hod],
                    f'Leave Application — {student.name} ({student.department})',
                    f'{student.name} ({student.roll_number}) applied for working-day leave from {from_date} to {to_date}.',
                    notif_type='leave',
                    link='/leave/hod/',
                )
            else:
                # No HOD configured — fall through to warden
                send_notification(
                    _staff_targets(student),
                    f'New Leave Application — {student.name}',
                    f'{student.name} ({student.roll_number}) applied for leave from {from_date} to {to_date}. (No HOD found for {student.department})',
                    notif_type='leave',
                    link='/leave/',
                )
        else:
            # Non-working day — goes directly to warden
            send_notification(
                _staff_targets(student),
                f'New Leave Application — {student.name}',
                f'{student.name} ({student.roll_number}) applied for leave from {from_date} to {to_date}.',
                notif_type='leave',
                link='/leave/',
            )

        messages.success(request, 'Leave application submitted successfully.')
        return redirect('my_leaves')
    return render(request, 'student/apply_leave.html', _ctx())


@login_required
def my_leaves(request):
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir
    leaves = LeaveApplication.objects.filter(student=student).order_by('-created_at')
    return render(request, 'student/my_leaves.html', {'leaves': leaves})


# ─── Medical Records ─────────────────────────────────────────────────────────

@login_required
def my_medical_records(request):
    """Student: view & submit their own medical records."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    if request.method == 'POST':
        from .models import MedicalRecord
        date_val = request.POST.get('date', '').strip()
        desc     = request.POST.get('description', '').strip()
        if date_val and desc:
            MedicalRecord.objects.create(
                student     = student,
                record_type = request.POST.get('record_type', 'illness'),
                date        = date_val,
                description = desc,
                hospital    = request.POST.get('hospital', '').strip(),
                doctor      = request.POST.get('doctor', '').strip(),
                medicines   = request.POST.get('medicines', '').strip(),
                ambulance_used   = 'ambulance_used' in request.POST,
                ambulance_reason = request.POST.get('ambulance_reason', '').strip(),
                pickup_location  = request.POST.get('pickup_location', '').strip(),
                drop_location    = request.POST.get('drop_location', '').strip(),
            )
            messages.success(request, 'Medical record submitted.')
        return redirect('my_medical_records')

    from .models import MedicalRecord
    records = MedicalRecord.objects.filter(student=student).order_by('-date')
    return render(request, 'student/my_medical_records.html', {
        'records':      records,
        'record_types': MedicalRecord.RecordType.choices,
    })


@admin_only
def admin_medical_records(request):
    """Admin: view all medical records with ambulance usage stats."""
    from .models import MedicalRecord
    warden_hostel = get_hostel_scope(request.user)
    qs = MedicalRecord.objects.select_related('student', 'student__allocations__room__hostel', 'verified_by').order_by('-date')
    if warden_hostel:
        qs = qs.filter(student__allocations__room__hostel=warden_hostel, student__allocations__status='active').distinct()

    type_filter    = request.GET.get('type', '')
    hostel_filter  = request.GET.get('hostel', '')
    verified_filter= request.GET.get('verified', '')
    if type_filter:
        qs = qs.filter(record_type=type_filter)
    if hostel_filter:
        qs = qs.filter(student__allocations__room__hostel_id=hostel_filter, student__allocations__status='active').distinct()
    if verified_filter == '1':
        qs = qs.filter(is_verified=True)
    elif verified_filter == '0':
        qs = qs.filter(is_verified=False)

    if request.method == 'POST':
        rec_pk = request.POST.get('rec_pk', '').strip()
        if rec_pk:
            rec = get_object_or_404(MedicalRecord, pk=rec_pk)
            rec.is_verified = True
            rec.verified_by = request.user
            rec.admin_notes = request.POST.get('admin_notes', '').strip()
            rec.save(update_fields=['is_verified', 'verified_by', 'admin_notes'])
            messages.success(request, 'Record verified.')
        return redirect(request.get_full_path())

    ambulance_count = MedicalRecord.objects.filter(ambulance_used=True)
    if warden_hostel:
        ambulance_count = ambulance_count.filter(
            student__allocations__room__hostel=warden_hostel, student__allocations__status='active'
        ).distinct()

    paginator = Paginator(qs, 25)
    page      = paginator.get_page(request.GET.get('page'))
    all_hostels = Hostel.objects.filter(is_active=True).order_by('name')
    return render(request, 'admin/medical_records.html', {
        'page': page, 'type_filter': type_filter,
        'hostel_filter': hostel_filter, 'verified_filter': verified_filter,
        'record_types': MedicalRecord.RecordType.choices,
        'all_hostels': all_hostels,
        'ambulance_total': ambulance_count.count(),
        'warden_hostel': warden_hostel,
    })


@admin_required
def leave_list(request):
    g = request.GET
    status_filter   = g.get('status', '')
    search          = g.get('q', '').strip()
    hostel_id       = g.get('hostel', '')
    department      = g.get('department', '').strip()
    category_id     = g.get('category', '')
    room_number     = g.get('room', '').strip()
    date_from       = g.get('date_from', '').strip()
    date_to         = g.get('date_to', '').strip()
    year_filter     = g.get('year', '')
    semester_filter = g.get('semester', '')
    gender_filter   = g.get('gender', '')
    state_filter    = g.get('state', '').strip()

    warden_hostel = get_hostel_scope(request.user)

    _alloc_prefetch = Prefetch(
        'student__allocations',
        queryset=RoomAllocation.objects.filter(status='active').select_related('room__hostel'),
        to_attr='active_allocs',
    )
    base_qs = LeaveApplication.objects.select_related(
        'student', 'reviewed_by', 'category',
        'warden_approved_by', 'hod_approved_by',
    ).prefetch_related(_alloc_prefetch)
    if warden_hostel:
        base_qs = base_qs.filter(
            student__allocations__room__hostel=warden_hostel,
            student__allocations__status='active',
        ).distinct()

    qs = base_qs.order_by('-created_at')

    if status_filter:
        qs = qs.filter(status=status_filter)
    if search:
        qs = qs.filter(
            Q(student__name__icontains=search) |
            Q(student__roll_number__icontains=search) |
            Q(leave_id__icontains=search)
        )
    if hostel_id:
        qs = qs.filter(
            student__allocations__room__hostel_id=hostel_id,
            student__allocations__status='active',
        ).distinct()
    if department:
        qs = qs.filter(student__department__icontains=department)
    if category_id:
        qs = qs.filter(category_id=category_id)
    if room_number:
        qs = qs.filter(
            student__allocations__room__room_number__icontains=room_number,
            student__allocations__status='active',
        ).distinct()
    if date_from:
        qs = qs.filter(from_date__gte=date_from)
    if date_to:
        qs = qs.filter(from_date__lte=date_to)
    if year_filter:
        qs = qs.filter(student__year=year_filter)
    if semester_filter:
        qs = qs.filter(student__semester=semester_filter)
    if gender_filter:
        qs = qs.filter(student__gender=gender_filter)
    if state_filter:
        qs = qs.filter(student__state__iexact=state_filter)

    counts = {
        'all':      base_qs.count(),
        'pending':  base_qs.filter(status='pending').count(),
        'approved': base_qs.filter(status='approved').count(),
        'rejected': base_qs.filter(status='rejected').count(),
    }

    paginator = Paginator(qs, 20)
    page = paginator.get_page(g.get('page'))

    pending_extensions = LeaveExtensionRequest.objects.filter(
        status='pending',
    ).select_related('leave__student', 'leave__category').order_by('-created_at')[:10]
    pending_hod = LeaveApplication.objects.filter(
        is_working_day_leave=True, hod_status='', status='pending',
    ).select_related('student', 'category').order_by('-created_at')[:10]

    all_hostels      = Hostel.objects.order_by('name') if not warden_hostel else []
    leave_categories = LeaveCategory.objects.filter(is_active=True).order_by('name')
    departments      = Student.objects.values_list('department', flat=True).distinct().order_by('department')
    states           = Student.objects.exclude(state='').values_list('state', flat=True).distinct().order_by('state')
    years            = Student.Year.choices

    active_filters = any([search, hostel_id, department, category_id, room_number, date_from, date_to, year_filter, semester_filter, gender_filter, state_filter])

    return render(request, 'admin/leave_list.html', {
        'page':              page,
        'status_filter':     status_filter,
        'counts':            counts,
        'statuses':          LeaveApplication.Status.choices,
        'pending_extensions': pending_extensions,
        'pending_hod':       pending_hod,
        'warden_hostel':     warden_hostel,
        # filter values
        'q':                 search,
        'hostel_id':         hostel_id,
        'f_department':      department,
        'f_category':        category_id,
        'f_room':            room_number,
        'date_from':         date_from,
        'date_to':           date_to,
        'year_filter':       year_filter,
        'semester_filter':   semester_filter,
        'gender_filter':     gender_filter,
        'state_filter':      state_filter,
        # filter options
        'all_hostels':       all_hostels,
        'leave_categories':  leave_categories,
        'departments':       departments,
        'states':            states,
        'years':             years,
        'active_filters':    active_filters,
    })

  
@admin_required
@require_POST
def leave_action(request, pk):
    leave   = get_object_or_404(LeaveApplication, pk=pk)
    action  = request.POST.get('action')
    remarks = request.POST.get('admin_remarks', '').strip()
    if action == 'approve':
        leave.status        = LeaveApplication.Status.APPROVED
        leave.admin_remarks = remarks
        leave.reviewed_by   = request.user
        leave.reviewed_at   = timezone.now()
        leave.save()
        if leave.student.user:
            send_notification(
                [leave.student.user],
                'Leave Application Approved',
                f'Your leave from {leave.from_date} to {leave.to_date} has been approved.' + (f' Remarks: {remarks}' if remarks else ''),
                notif_type='leave',
            )
        messages.success(request, f'Leave approved for {leave.student.name}.')
    elif action == 'reject':
        leave.status        = LeaveApplication.Status.REJECTED
        leave.admin_remarks = remarks
        leave.reviewed_by   = request.user
        leave.reviewed_at   = timezone.now()
        leave.save()
        if leave.student.user:
            send_notification(
                [leave.student.user],
                'Leave Application Rejected',
                f'Your leave from {leave.from_date} to {leave.to_date} has been rejected.' + (f' Reason: {remarks}' if remarks else ''),
                notif_type='leave',
            )
        messages.warning(request, f'Leave rejected for {leave.student.name}.')
    return redirect('leave_list')


@login_required
def hod_leave_list(request):
    """HOD sees pending leave applications for their department."""
    from apps.accounts.models import User as _User
    if request.user.role not in (_User.Role.HOD, _User.Role.SUPER_ADMIN, _User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    dept = request.user.department
    qs = LeaveApplication.objects.filter(
        is_working_day_leave=True,
        student__department__iexact=dept,
    ).select_related('student', 'category', 'hod_approved_by').order_by('-created_at')

    tab = request.GET.get('tab', 'pending')
    if tab == 'pending':
        qs = qs.filter(hod_status='')
    elif tab == 'approved':
        qs = qs.filter(hod_status='approved')
    elif tab == 'rejected':
        qs = qs.filter(hod_status='rejected')

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))
    pending_count = LeaveApplication.objects.filter(
        is_working_day_leave=True,
        student__department__iexact=dept,
        hod_status=''
    ).count()
    return render(request, 'hod/leave_list.html', {
        'page': page, 'tab': tab, 'pending_count': pending_count, 'dept': dept,
    })


@login_required
@require_POST
def leave_hod_action(request, pk):
    """HOD approves or rejects a working-day leave application."""
    from apps.accounts.models import User as _User
    if request.user.role not in (_User.Role.HOD, _User.Role.SUPER_ADMIN, _User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    leave  = get_object_or_404(LeaveApplication, pk=pk)
    action = request.POST.get('action')
    remarks = request.POST.get('remarks', '').strip()

    if action == 'approve':
        leave.hod_status      = LeaveApplication.Status.APPROVED
        leave.hod_approved_by = request.user
        leave.hod_approved_at = timezone.now()
        leave.save()
        # Forward to warden after HOD approval
        send_notification(
            _staff_targets(leave.student),
            f'Leave Application — HOD Approved — {leave.student.name}',
            f'HOD approved working-day leave for {leave.student.name} ({leave.student.roll_number}). '
            f'Period: {leave.from_date} to {leave.to_date}. Pending warden approval.',
            notif_type='leave',
            link='/leave/',
        )
        if leave.student.user:
            send_notification(
                [leave.student.user],
                'Leave — HOD Approved',
                f'Your leave ({leave.from_date} to {leave.to_date}) has been approved by your HOD and is now pending warden approval.',
                notif_type='leave',
            )
        messages.success(request, f'Leave approved for {leave.student.name}. Forwarded to warden.')
    elif action == 'reject':
        leave.hod_status      = LeaveApplication.Status.REJECTED
        leave.hod_approved_by = request.user
        leave.hod_approved_at = timezone.now()
        leave.status          = LeaveApplication.Status.REJECTED
        leave.admin_remarks   = remarks
        leave.save()
        if leave.student.user:
            send_notification(
                [leave.student.user],
                'Leave Application Rejected by HOD',
                f'Your leave ({leave.from_date} to {leave.to_date}) was rejected by your HOD.' + (f' Reason: {remarks}' if remarks else ''),
                notif_type='leave',
            )
        messages.warning(request, f'Leave rejected for {leave.student.name}.')
    return redirect('hod_leave_list')


# ═══════════════════════════════════════════════════════════════════════════════
# SUPER ADMIN MODULE
# ═══════════════════════════════════════════════════════════════════════════════

@superadmin_only
def superadmin_dashboard(request):
    # SuperAdmin sees the same full dashboard as admin — delegate directly.
    # is_admin_or_warden includes SUPER_ADMIN so the @admin_required check passes.
    return admin_dashboard_v2(request)


@superadmin_only
def superadmin_permissions(request):
    import re as _re
    BUILTIN_ROLES = [
        ('admin',       'Admin 2'),
        ('warden',      'Warden'),
        ('security',    'Security Guard'),
        ('maintenance', 'Maintenance Incharge'),
        ('mess',        'Mess Management'),
    ]

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'create_role':
            role_name = request.POST.get('role_name', '').strip()
            if not role_name:
                messages.error(request, 'Role name is required.')
            else:
                key = _re.sub(r'[^a-z0-9]+', '_', role_name.lower()).strip('_')[:40]
                reserved = {r[0] for r in BUILTIN_ROLES} | {'superadmin', 'student'}
                if not key:
                    messages.error(request, 'Role name produced an invalid key. Use letters/numbers.')
                elif key in reserved:
                    messages.error(request, f'"{role_name}" conflicts with a built-in role.')
                elif CustomStaffRole.objects.filter(key=key).exists():
                    messages.error(request, f'A role with key "{key}" already exists.')
                else:
                    allowed = request.POST.getlist('modules_new')
                    CustomStaffRole.objects.create(name=role_name, key=key, created_by=request.user)
                    RoleModulePermission.objects.update_or_create(
                        role=key,
                        defaults={'allowed_modules': allowed, 'updated_by': request.user},
                    )
                    messages.success(request, f'Custom role "{role_name}" created.')

        elif action == 'delete_role':
            key        = request.POST.get('role_key', '').strip()
            is_custom  = CustomStaffRole.objects.filter(key=key).exists()
            role_label = dict(BUILTIN_ROLES).get(key) or \
                         (CustomStaffRole.objects.filter(key=key).values_list('name', flat=True).first()) or key
            user_count = User.objects.filter(role=key).count()
            if user_count > 0:
                messages.warning(request,
                    f'Cannot delete "{role_label}": {user_count} staff member(s) are assigned to it. '
                    f'Reassign or deactivate them first from Staff Accounts.')
            elif is_custom:
                CustomStaffRole.objects.filter(key=key).delete()
                RoleModulePermission.objects.filter(role=key).delete()
                messages.success(request, f'Custom role "{role_label}" deleted.')
            else:
                # Built-in: mark as inactive (can be restored)
                RoleModulePermission.objects.update_or_create(
                    role=key,
                    defaults={'is_active': False, 'updated_by': request.user},
                )
                messages.success(request, f'Role "{role_label}" removed. You can restore it anytime.')

        elif action == 'restore_role':
            key = request.POST.get('role_key', '').strip()
            role_label = dict(BUILTIN_ROLES).get(key, key)
            RoleModulePermission.objects.update_or_create(
                role=key,
                defaults={'is_active': True, 'updated_by': request.user},
            )
            messages.success(request, f'Role "{role_label}" restored.')

        else:  # save permissions for all active roles + student
            active_builtin = []
            for r_key, _ in BUILTIN_ROLES:
                perm = RoleModulePermission.objects.filter(role=r_key).first()
                if not perm or perm.is_active:
                    active_builtin.append(r_key)
            all_role_keys = (active_builtin
                             + list(CustomStaffRole.objects.values_list('key', flat=True))
                             + ['student'])
            for role_key in all_role_keys:
                allowed = request.POST.getlist('modules_' + role_key)
                RoleModulePermission.objects.update_or_create(
                    role=role_key,
                    defaults={'allowed_modules': allowed, 'is_active': True, 'updated_by': request.user},
                )
            messages.success(request, 'Module permissions updated successfully.')

        return redirect('superadmin_permissions')

    builtin_label = dict(BUILTIN_ROLES)
    roles_perms   = []
    deleted_roles  = []

    for role_key, role_label in BUILTIN_ROLES:
        perm    = RoleModulePermission.objects.filter(role=role_key).first()
        active  = (not perm) or perm.is_active
        allowed = set(perm.allowed_modules if perm else DEFAULT_ROLE_MODULES.get(role_key, []))
        entry   = {
            'key': role_key, 'label': role_label, 'allowed': allowed,
            'custom': False, 'deletable': True,
            'user_count': User.objects.filter(role=role_key).count(),
            'modules_list': STAFF_MODULES,
        }
        if active:
            roles_perms.append(entry)
        else:
            deleted_roles.append(entry)

    for cr in CustomStaffRole.objects.all():
        perm    = RoleModulePermission.objects.filter(role=cr.key).first()
        allowed = set(perm.allowed_modules if perm else [])
        roles_perms.append({
            'key': cr.key, 'label': cr.name, 'allowed': allowed,
            'custom': True, 'deletable': True,
            'user_count': User.objects.filter(role=cr.key).count(),
            'modules_list': STAFF_MODULES,
        })

    # Student portal permissions (always shown, not deletable)
    student_perm = RoleModulePermission.objects.filter(role='student').first()
    student_allowed = set(
        student_perm.allowed_modules if student_perm
        else DEFAULT_ROLE_MODULES.get('student', [])
    )
    roles_perms.append({
        'key': 'student', 'label': 'Student Portal', 'allowed': student_allowed,
        'custom': False, 'deletable': False,
        'user_count': User.objects.filter(role='student').count(),
        'modules_list': STUDENT_MODULES,
        'is_student': True,
    })

    return render(request, 'admin/superadmin_permissions.html', {
        'roles_perms':   roles_perms,
        'deleted_roles': deleted_roles,
        'all_modules':   STAFF_MODULES,
    })


@superadmin_only
def staff_accounts(request):
    BUILTIN_ROLES = [
        ('admin',       'Admin 2'),
        ('warden',      'Warden'),
        ('security',    'Security Guard'),
        ('maintenance', 'Maintenance Incharge'),
        ('mess',        'Mess Management'),
    ]
    # Exclude built-in roles that the superadmin has deleted (is_active=False)
    inactive_keys = set(
        RoleModulePermission.objects.filter(
            role__in=[r[0] for r in BUILTIN_ROLES], is_active=False
        ).values_list('role', flat=True)
    )
    active_builtin = [(k, l) for k, l in BUILTIN_ROLES if k not in inactive_keys]
    custom_roles   = list(CustomStaffRole.objects.values_list('key', 'name'))
    all_roles      = active_builtin + custom_roles

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            email    = request.POST.get('email', '').strip()
            name     = request.POST.get('name', '').strip()
            role     = request.POST.get('role', '').strip()
            phone    = request.POST.get('phone', '').strip()
            password = request.POST.get('password', '').strip()
            valid_roles = [r[0] for r in all_roles]
            if not all([email, name, role, password]):
                messages.error(request, 'Email, Name, Role and Password are required.')
            elif len(password) < 6:
                messages.error(request, 'Password must be at least 6 characters.')
            elif role not in valid_roles:
                messages.error(request, 'Invalid role selected.')
            elif User.objects.filter(email=email).exists():
                messages.error(request, 'An account with this email already exists.')
            else:
                User.objects.create_user(
                    email=email, password=password,
                    name=name, role=role, phone=phone,
                )
                role_label = dict(all_roles).get(role, role)
                messages.success(request, f'Account created for {name} ({role_label}).')

        elif action == 'toggle_active':
            try:
                target = User.objects.get(pk=request.POST.get('user_id'))
                if target.is_super_admin:
                    messages.error(request, 'Cannot deactivate an Admin 1 account.')
                else:
                    target.is_active = not target.is_active
                    target.save(update_fields=['is_active'])
                    state = 'activated' if target.is_active else 'deactivated'
                    messages.success(request, target.name + ' ' + state + '.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')

        elif action == 'reset_password':
            try:
                target       = User.objects.get(pk=request.POST.get('user_id'))
                new_password = request.POST.get('new_password', '').strip()
                if len(new_password) < 6:
                    messages.error(request, 'Password must be at least 6 characters.')
                else:
                    target.set_password(new_password)
                    target.save()
                    messages.success(request, 'Password reset for ' + target.name + '.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')

        elif action == 'delete_staff':
            try:
                target = User.objects.get(pk=request.POST.get('user_id'))
                if target.is_super_admin:
                    messages.error(request, 'Cannot delete an Admin 1 account.')
                else:
                    name = target.name
                    target.delete()
                    messages.success(request, f'Account for "{name}" has been permanently deleted.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')

        elif action == 'create_role':
            import re as _re
            role_name = request.POST.get('role_name', '').strip()
            if not role_name:
                messages.error(request, 'Role name is required.')
            else:
                key = _re.sub(r'[^a-z0-9]+', '_', role_name.lower()).strip('_')[:40]
                reserved = {'superadmin', 'admin', 'warden', 'student', 'security',
                            'maintenance', 'mess'}
                if not key:
                    messages.error(request, 'Role name produced an invalid key.')
                elif key in reserved:
                    messages.error(request, f'"{role_name}" conflicts with a built-in role.')
                elif CustomStaffRole.objects.filter(key=key).exists():
                    messages.error(request, f'Role "{role_name}" already exists.')
                else:
                    allowed = request.POST.getlist('modules_new')
                    CustomStaffRole.objects.create(
                        name=role_name, key=key, created_by=request.user
                    )
                    RoleModulePermission.objects.update_or_create(
                        role=key,
                        defaults={'allowed_modules': allowed, 'updated_by': request.user},
                    )
                    messages.success(request, f'Role "{role_name}" created. You can now assign staff to it.')

        return redirect('staff_accounts')

    all_role_keys = [r[0] for r in all_roles]
    staff = User.objects.filter(role__in=all_role_keys).order_by('role', 'name')
    role_label_map = dict(all_roles)

    return render(request, 'admin/staff_accounts.html', {
        'staff':          staff,
        'builtin_roles':  active_builtin,
        'custom_roles':   custom_roles,
        'role_label_map': role_label_map,
        'all_modules':    STAFF_MODULES,
    })


@login_required
def staff_dashboard(request):
    """Landing dashboard for maintenance, mess, and any custom staff roles."""
    system_roles = {'superadmin', 'admin', 'warden', 'student', 'security'}
    if request.user.role in system_roles:
        return redirect('dashboard')
    custom_role_name = None
    if request.user.is_custom_staff_role:
        cr = CustomStaffRole.objects.filter(key=request.user.role).first()
        custom_role_name = cr.name if cr else request.user.role.replace('_', ' ').title()
    if request.user.role == User.Role.MESS:
        return redirect('mess_incharge_dashboard')
    if request.user.role == User.Role.MAINTENANCE:
        return redirect('maintenance_incharge_dashboard')
    return render(request, 'admin/staff_dashboard.html', {
        'role':             request.user.role,
        'custom_role_name': custom_role_name,
    })


# ─── Staff Profile views ───────────────────────────────────────────────────────

STAFF_PROFILE_FIELDS = [
    'employee_id', 'gender', 'blood_group',
    'address', 'city', 'state', 'pincode',
    'emergency_name', 'emergency_phone', 'emergency_relation',
    'qualification', 'department',
    'aadhar_number', 'college_id_number', 'bio',
]

NON_STUDENT_ROLES = {'admin', 'warden', 'security', 'maintenance', 'mess'}


def _is_staff_member(user):
    return (not user.is_super_admin and
            user.role != 'student' and
            (user.role in NON_STUDENT_ROLES or user.is_custom_staff_role))  


@login_required
def my_staff_profile(request):
    if not _is_staff_member(request.user):
        return redirect('dashboard')

    profile, _ = StaffProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        for field in STAFF_PROFILE_FIELDS:
            val = request.POST.get(field, '').strip()
            setattr(profile, field, val if val else '')
        # date fields — keep None if blank
        for date_field in ('date_of_birth', 'date_of_joining'):
            raw = request.POST.get(date_field, '').strip()
            setattr(profile, date_field, raw if raw else None)
        exp = request.POST.get('experience_years', '').strip()
        profile.experience_years = int(exp) if exp.isdigit() else None
        # aadhar: strip non-digits, store raw 12-digit number
        aadhar = ''.join(filter(str.isdigit, request.POST.get('aadhar_number', '')))
        profile.aadhar_number = aadhar[:12] if aadhar else ''
        # phone on user model
        phone = request.POST.get('phone', '').strip()
        request.user.phone = phone
        request.user.save(update_fields=['phone'])
        # document uploads
        for doc_field in ('aadhar_doc', 'college_id_doc', 'appointment_letter', 'qualification_doc'):
            if request.POST.get(f'clear_{doc_field}'):
                old = getattr(profile, doc_field)
                if old:
                    old.delete(save=False)
                setattr(profile, doc_field, None)
            elif doc_field in request.FILES:
                old = getattr(profile, doc_field)
                if old:
                    old.delete(save=False)
                setattr(profile, doc_field, request.FILES[doc_field])
        profile.save()
        messages.success(request, 'Profile updated successfully.')
        return redirect('my_staff_profile')

    return render(request, 'staff/staff_profile_edit.html', {
        'profile': profile,
        'user':    request.user,
    })


@superadmin_only
def staff_directory(request):
    BUILTIN = {'admin', 'warden', 'security', 'maintenance', 'mess'}
    custom_keys = set(CustomStaffRole.objects.values_list('key', flat=True))
    all_staff_roles = BUILTIN | custom_keys
    staff_users = User.objects.filter(
        role__in=all_staff_roles
    ).select_related('staff_profile').order_by('role', 'name')

    # Attach profile (or None) and completion
    staff_list = []
    for u in staff_users:
        try:
            p = u.staff_profile
        except StaffProfile.DoesNotExist:
            p = None
        staff_list.append({
            'user':       u,
            'profile':    p,
            'pct':        p.completion_pct if p else 0,
            'label':      p.completion_label if p else 'No Profile',
        })

    return render(request, 'admin/staff_directory.html', {'staff_list': staff_list})


@superadmin_only
def staff_profile_detail(request, uid):
    target_user = get_object_or_404(User, pk=uid)
    try:
        profile = target_user.staff_profile
    except StaffProfile.DoesNotExist:
        profile = None
    return render(request, 'admin/staff_profile_detail.html', {
        'target':  target_user,
        'profile': profile,
    })



# ═══════════════════════════════════════════════════════════════════════════════
# ASSETS MANAGEMENT MODULE
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def asset_dashboard(request):
    role = request.user.role
    if role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    f_hostel    = request.GET.get('hostel', '')
    f_type      = request.GET.get('asset_type', '')
    f_condition = request.GET.get('condition', '')
    f_status    = request.GET.get('status', '')
    f_q         = request.GET.get('q', '').strip()

    qs = Asset.objects.select_related('hostel', 'room', 'asset_type').all()
    if f_hostel:    qs = qs.filter(hostel_id=f_hostel)
    if f_type:      qs = qs.filter(asset_type_id=f_type)
    if f_condition: qs = qs.filter(condition=f_condition)
    if f_status:    qs = qs.filter(status=f_status)
    if f_q:
        qs = qs.filter(Q(name__icontains=f_q) | Q(asset_code__icontains=f_q) |
                       Q(vendor_name__icontains=f_q))

    counts = {
        'total':             Asset.objects.count(),
        'active':            Asset.objects.filter(status='active').count(),
        'damaged':           Asset.objects.filter(condition='damaged').count(),
        'maintenance':       Asset.objects.filter(condition='under_maintenance').count(),
        'lost':              Asset.objects.filter(status='lost').count(),
        'pending_transfers': AssetTransfer.objects.filter(status='pending').count(),
    }

    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'admin/assets.html', {
        'page_obj': page_obj, 'counts': counts,
        'hostels': Hostel.objects.all(),
        'asset_types': AssetType.objects.all(),
        'conditions': Asset.Condition.choices,
        'statuses': Asset.Status.choices,
        'filters': {'hostel': f_hostel, 'asset_type': f_type,
                    'condition': f_condition, 'status': f_status, 'q': f_q},
    })


@login_required
def asset_create(request):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    if request.method == 'POST':
        name       = request.POST.get('name', '').strip()
        asset_code = request.POST.get('asset_code', '').strip()
        if not name or not asset_code:
            messages.error(request, 'Name and asset code are required.')
            return redirect('asset_create')
        if Asset.objects.filter(asset_code=asset_code).exists():
            messages.error(request, f'Asset code "{asset_code}" already exists.')
            return redirect('asset_create')
        floor_val = request.POST.get('floor', '')
        asset = Asset.objects.create(
            name=name, asset_code=asset_code,
            asset_type_id=request.POST.get('asset_type') or None,
            hostel_id=request.POST.get('hostel') or None,
            floor=int(floor_val) if floor_val else None,
            room_id=request.POST.get('room') or None,
            condition=request.POST.get('condition', 'good'),
            vendor_name=request.POST.get('vendor_name', '').strip(),
            vendor_contact=request.POST.get('vendor_contact', '').strip(),
            vendor_invoice=request.POST.get('vendor_invoice', '').strip(),
            purchase_date=request.POST.get('purchase_date') or None,
            purchase_cost=request.POST.get('purchase_cost') or None,
            assigned_date=request.POST.get('assigned_date') or None,
            notes=request.POST.get('notes', '').strip(),
            added_by=request.user,
        )
        AssetLog.objects.create(
            asset=asset, action=AssetLog.Action.CREATED,
            description=f'Asset created at {asset.location_display}.',
            performed_by=request.user,
        )
        messages.success(request, f'Asset "{asset.name}" created.')
        return redirect('asset_detail', pk=asset.pk)
    return render(request, 'admin/asset_form.html', {
        'hostels': Hostel.objects.all(), 'asset_types': AssetType.objects.all(),
        'conditions': Asset.Condition.choices,
        'rooms': Room.objects.select_related('hostel').order_by('hostel__name', 'room_number'),
        'action': 'Create',
    })


@login_required
def asset_edit(request, pk):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        old_cond = asset.condition
        asset.name           = request.POST.get('name', asset.name).strip()
        asset.asset_type_id  = request.POST.get('asset_type') or None
        asset.hostel_id      = request.POST.get('hostel') or None
        fv = request.POST.get('floor', '')
        asset.floor          = int(fv) if fv else None
        asset.room_id        = request.POST.get('room') or None
        asset.condition      = request.POST.get('condition', asset.condition)
        asset.vendor_name    = request.POST.get('vendor_name', '').strip()
        asset.vendor_contact = request.POST.get('vendor_contact', '').strip()
        asset.vendor_invoice = request.POST.get('vendor_invoice', '').strip()
        asset.purchase_date  = request.POST.get('purchase_date') or None
        pc = request.POST.get('purchase_cost', '')
        asset.purchase_cost  = pc if pc else None
        asset.assigned_date  = request.POST.get('assigned_date') or None
        asset.notes          = request.POST.get('notes', '').strip()
        asset.save()
        action = AssetLog.Action.CONDITION_CHANGED if old_cond != asset.condition else AssetLog.Action.UPDATED
        desc = (f'Condition changed: {old_cond} → {asset.condition}.'
                if action == AssetLog.Action.CONDITION_CHANGED else 'Details updated.')
        AssetLog.objects.create(asset=asset, action=action, description=desc, performed_by=request.user)
        messages.success(request, 'Asset updated.')
        return redirect('asset_detail', pk=asset.pk)
    return render(request, 'admin/asset_form.html', {
        'asset': asset, 'hostels': Hostel.objects.all(),
        'asset_types': AssetType.objects.all(), 'conditions': Asset.Condition.choices,
        'rooms': Room.objects.select_related('hostel').order_by('hostel__name', 'room_number'),
        'action': 'Edit',
    })


@login_required
def asset_detail(request, pk):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    asset = get_object_or_404(
        Asset.objects.select_related('hostel', 'room', 'asset_type', 'added_by', 'discarded_by'), pk=pk)
    logs      = asset.logs.select_related('performed_by').order_by('-created_at')
    transfers = asset.transfers.select_related(
        'from_hostel', 'from_room', 'to_hostel', 'to_room', 'requested_by', 'approved_by').order_by('-created_at')
    return render(request, 'admin/asset_detail.html', {
        'asset': asset, 'logs': logs, 'transfers': transfers,
        'hostels': Hostel.objects.all(),
        'rooms': Room.objects.select_related('hostel').order_by('hostel__name', 'room_number'),
        'conditions': Asset.Condition.choices,
    })


@login_required
@require_POST
def asset_discard(request, pk):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    asset  = get_object_or_404(Asset, pk=pk)
    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, 'Discard reason is required.')
        return redirect('asset_detail', pk=pk)
    asset.status = Asset.Status.DISCARDED
    asset.condition = Asset.Condition.DISCARDED
    asset.discard_reason = reason
    asset.discarded_at   = timezone.now()
    asset.discarded_by   = request.user
    asset.save()
    AssetLog.objects.create(asset=asset, action=AssetLog.Action.DISCARDED,
                            description=f'Discarded. Reason: {reason}', performed_by=request.user)
    messages.success(request, f'Asset "{asset.name}" discarded.')
    return redirect('asset_dashboard')


@login_required
def asset_transfer_request(request, pk):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        to_hostel_id  = request.POST.get('to_hostel', '')
        to_floor_val  = request.POST.get('to_floor', '')
        to_room_id    = request.POST.get('to_room', '')
        reason        = request.POST.get('reason', '').strip()
        transfer_date = request.POST.get('transfer_date', str(timezone.now().date()))
        if not reason:
            messages.error(request, 'Transfer reason is required.')
            return redirect('asset_detail', pk=pk)
        is_superadmin = request.user.role == User.Role.SUPER_ADMIN
        t_status = AssetTransfer.Status.COMPLETED if is_superadmin else AssetTransfer.Status.PENDING
        transfer = AssetTransfer.objects.create(
            asset=asset,
            from_hostel=asset.hostel, from_room=asset.room, from_floor=asset.floor,
            to_hostel_id=to_hostel_id or None, to_room_id=to_room_id or None,
            to_floor=int(to_floor_val) if to_floor_val else None,
            reason=reason, transfer_date=transfer_date,
            requested_by=request.user, status=t_status,
        )
        if is_superadmin:
            transfer.approved_by = request.user
            transfer.approved_at = timezone.now()
            transfer.save()
            asset.hostel_id = to_hostel_id or None
            asset.room_id   = to_room_id or None
            asset.floor     = int(to_floor_val) if to_floor_val else None
            asset.assigned_date = transfer_date
            asset.save()
            AssetLog.objects.create(asset=asset, action=AssetLog.Action.TRANSFERRED,
                                    description=f'Transferred to {asset.location_display}.', performed_by=request.user)
            messages.success(request, 'Asset transferred.')
        else:
            AssetLog.objects.create(asset=asset, action=AssetLog.Action.TRANSFERRED,
                                    description='Transfer requested — awaiting approval.', performed_by=request.user)
            messages.success(request, 'Transfer request submitted.')
        return redirect('asset_detail', pk=pk)
    return redirect('asset_detail', pk=pk)


@login_required
@require_POST
def asset_transfer_action(request, transfer_pk):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    transfer = get_object_or_404(AssetTransfer, pk=transfer_pk)
    action   = request.POST.get('action', '')
    remarks  = request.POST.get('remarks', '').strip()
    if action == 'approve' and transfer.status == AssetTransfer.Status.PENDING:
        transfer.status = AssetTransfer.Status.COMPLETED
        transfer.approved_by = request.user
        transfer.approved_at = timezone.now()
        transfer.remarks     = remarks
        transfer.save()
        a = transfer.asset
        a.hostel_id = transfer.to_hostel_id
        a.room_id   = transfer.to_room_id
        a.floor     = transfer.to_floor
        a.assigned_date = str(transfer.transfer_date)
        a.save()
        AssetLog.objects.create(asset=a, action=AssetLog.Action.TRANSFERRED,
                                description=f'Transfer approved. Now at {a.location_display}.', performed_by=request.user)
        messages.success(request, 'Transfer approved.')
    elif action == 'reject' and transfer.status == AssetTransfer.Status.PENDING:
        transfer.status  = AssetTransfer.Status.REJECTED
        transfer.remarks = remarks
        transfer.save()
        messages.info(request, 'Transfer rejected.')
    return redirect('asset_detail', pk=transfer.asset_id)


@login_required
def asset_bulk_upload(request):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please select a CSV file.')
            return redirect('asset_bulk_upload')
        decoded = csv_file.read().decode('utf-8-sig')
        reader  = csv.DictReader(io.StringIO(decoded))
        created = 0
        errors  = []
        for i, row in enumerate(reader, start=2):
            code = row.get('asset_code', '').strip()
            name = row.get('name', '').strip()
            if not code or not name:
                errors.append(f'Row {i}: asset_code and name required.')
                continue
            if Asset.objects.filter(asset_code=code).exists():
                errors.append(f'Row {i}: "{code}" already exists — skipped.')
                continue
            hostel = None
            hn = row.get('hostel', '').strip()
            if hn:
                hostel = Hostel.objects.filter(name__iexact=hn).first()
            at = None
            tn = row.get('asset_type', '').strip()
            if tn:
                at, _ = AssetType.objects.get_or_create(name=tn)
            fv = row.get('floor', '')
            Asset.objects.create(
                asset_code=code, name=name, asset_type=at, hostel=hostel,
                floor=int(fv) if fv else None,
                condition=row.get('condition', 'good'),
                vendor_name=row.get('vendor_name', ''),
                purchase_date=row.get('purchase_date') or None,
                purchase_cost=row.get('purchase_cost') or None,
                notes=row.get('notes', ''),
                added_by=request.user,
            )
            created += 1
        if created:
            messages.success(request, f'{created} asset(s) imported.')
        for err in errors[:5]:
            messages.warning(request, err)
        return redirect('asset_dashboard')
    return render(request, 'admin/asset_bulk_upload.html')


def asset_bulk_sample(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="asset_upload_sample.csv"'
    w = csv.writer(response)
    w.writerow(['asset_code', 'name', 'asset_type', 'hostel', 'floor', 'condition',
                'vendor_name', 'purchase_date', 'purchase_cost', 'notes'])
    w.writerow(['AST-001', 'Ceiling Fan', 'Electrical', 'Boys Hostel 1', '1', 'good',
                'Usha Industries', '2024-01-15', '1200.00', 'Double blade fan'])
    w.writerow(['AST-002', 'Study Table', 'Furniture', 'Girls Hostel 1', '2', 'fair',
                'Local Vendor', '2023-06-10', '2500.00', ''])
    return response


@login_required
def asset_type_list(request):
    if request.user.role not in (User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', 'bi-box-fill').strip()
        desc = request.POST.get('description', '').strip()
        if name:
            AssetType.objects.get_or_create(name=name, defaults={'icon': icon, 'description': desc})
            messages.success(request, f'Asset type "{name}" created.')
        return redirect('asset_type_list')
    types = AssetType.objects.annotate(asset_count=Count('assets'))
    return render(request, 'admin/asset_types.html', {'types': types})


# ═══════════════════════════════════════════════════════════════════════════════
# GATE PASS CATEGORY MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def gp_category_list(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        name   = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Name required.')
            return redirect('gp_category_list')
        cat = get_object_or_404(GatePassCategory, pk=request.POST.get('pk')) if action == 'edit' else GatePassCategory(created_by=request.user)
        cat.name        = name
        cat.description = request.POST.get('description', '').strip()
        cat.start_time  = request.POST.get('start_time') or None
        cat.end_time    = request.POST.get('end_time') or None
        cat.max_hours   = int(request.POST.get('max_hours', 12) or 12)
        cat.requires_warden_approval     = 'req_warden' in request.POST
        cat.requires_hod_approval        = 'req_hod' in request.POST
        cat.requires_admin_approval      = 'req_admin' in request.POST
        cat.requires_superadmin_approval = 'req_superadmin' in request.POST
        cat.is_emergency = 'is_emergency' in request.POST
        cat.is_active    = 'is_active' in request.POST
        cat.save()
        messages.success(request, f'Category "{cat.name}" saved.')
        return redirect('gp_category_list')
    categories = GatePassCategory.objects.annotate(pass_count=Count('gate_passes'))
    return render(request, 'admin/gp_categories.html', {'categories': categories})


@login_required
@require_POST
def gp_category_toggle(request, pk):
    if request.user.role != User.Role.SUPER_ADMIN:
        return redirect('dashboard')
    cat = get_object_or_404(GatePassCategory, pk=pk)
    cat.is_active = not cat.is_active
    cat.save()
    return redirect('gp_category_list')


# ─── Gate Pass Sequence Reset (superadmin) ───────────────────────────────────

@login_required
def gp_sequence_reset(request):
    """Superadmin can reset the GP sequence counter for a hostel prefix."""
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    from .models import GatePassSequence
    sequences = GatePassSequence.objects.all().order_by('hostel_prefix')

    if request.method == 'POST':
        prefix = request.POST.get('hostel_prefix', '').strip()
        if prefix:
            seq_obj, _ = GatePassSequence.objects.get_or_create(
                hostel_prefix=prefix, defaults={'current_seq': 1}
            )
            seq_obj.current_seq = 1
            seq_obj.reset_at = timezone.now()
            seq_obj.reset_by = request.user
            seq_obj.save()
            messages.success(request, f'Gate pass sequence reset for {prefix}. Next GP will be {prefix}-00001.')
        return redirect('gp_sequence_reset')

    all_hostels = Hostel.objects.all().order_by('name')
    return render(request, 'admin/gp_sequence_reset.html', {
        'sequences': sequences,
        'all_hostels': all_hostels,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# QUOTA POLICY MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def quota_policy_list(request):
    """Admin 1 manages gate pass / leave quota policies."""
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    if request.method == 'POST':
        action = request.POST.get('action', 'create')

        if action == 'delete':
            get_object_or_404(QuotaPolicy, pk=request.POST.get('pk')).delete()
            messages.success(request, 'Policy deleted.')
            return redirect('quota_policy_list')

        if action == 'toggle':
            p = get_object_or_404(QuotaPolicy, pk=request.POST.get('pk'))
            p.is_active = not p.is_active
            p.save()
            messages.success(request, f'Policy {"enabled" if p.is_active else "disabled"}.')
            return redirect('quota_policy_list')

        # create / edit
        pk          = request.POST.get('pk')
        policy_type = request.POST.get('policy_type')
        period      = request.POST.get('period')
        limit       = int(request.POST.get('limit', 1) or 1)
        scope       = request.POST.get('scope', 'all')
        student_id  = request.POST.get('student_id') or None

        policy = get_object_or_404(QuotaPolicy, pk=pk) if pk else QuotaPolicy(created_by=request.user)
        policy.policy_type    = policy_type
        policy.period         = period
        policy.limit          = limit
        policy.applies_to_all = (scope == 'all')
        policy.student        = Student.objects.filter(pk=student_id).first() if student_id else None
        policy.is_active      = True
        policy.save()
        messages.success(request, 'Quota policy saved.')
        return redirect('quota_policy_list')

    policies = QuotaPolicy.objects.select_related('student', 'created_by').order_by('-created_at')
    students = Student.objects.filter(is_resident=True).order_by('name')
    return render(request, 'superadmin/quota_policies.html', {
        'policies':     policies,
        'students':     students,
        'policy_types': QuotaPolicy.PolicyType.choices,
        'periods':      QuotaPolicy.Period.choices,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# ACADEMIC CALENDAR
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_ptu_excel(file_obj):
    """
    Parse IKG PTU academic calendar Excel format.
    Layout: intro row → header row → data rows
    Columns repeat in groups of 3: Date/Day | Event | Working Days count/"H"
    Returns list of dicts: {date, event_type, title}
    """
    import openpyxl
    import re
    from datetime import date as dt

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    EXAM_KEYWORDS   = {'mst', 'exam', 'test', 'assessment', 'result', 'submission', 'viva', 'practical'}
    FESTIVAL_KEYWORDS = {'jayanti', 'diwali', 'holi', 'eid', 'christmas', 'navratri', 'puja',
                         'dussehra', 'baisakhi', 'vaisakhi', 'lohri', 'shivratri', 'janmashtami',
                         'independence', 'republic', 'gandhi', 'ambedkar', 'guru', 'path'}

    def _guess_type(event_str, working_val):
        """Map event text + working-day column value to an EventType."""
        ev_low = (event_str or '').lower()
        if 'week off' in ev_low:
            return 'sunday'
        if str(working_val).strip().upper() == 'H':
            kw = ev_low
            if any(k in kw for k in FESTIVAL_KEYWORDS):
                return 'festival'
            if any(k in kw for k in EXAM_KEYWORDS):
                return 'exam'
            return 'holiday'
        # working day — check for exam events
        if any(k in ev_low for k in EXAM_KEYWORDS):
            return 'exam'
        return 'working'

    results = []
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row: first row containing 'Date/Day' (case-insensitive)
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if any(str(c or '').strip().lower() == 'date/day' for c in row):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find the header row with 'Date/Day' columns.")

    # Identify which columns are "Date/Day" columns (every 3rd starting from first match)
    header = all_rows[header_row_idx]
    date_cols = [i for i, c in enumerate(header) if str(c or '').strip().lower() == 'date/day']

    # Parse data rows (skip header row itself)
    year_hint = None
    for row in all_rows[header_row_idx + 1:]:
        for dc in date_cols:
            event_col  = dc + 1
            working_col = dc + 2
            date_cell  = str(row[dc] or '').strip() if dc < len(row) else ''
            event_cell = str(row[event_col] or '').strip() if event_col < len(row) else ''
            work_cell  = str(row[working_col] or '').strip() if working_col < len(row) else ''

            if not date_cell:
                continue

            # Parse date like "Thu, Jan 1" or "Thu, Jan 01"
            m = re.match(r'\w+,?\s+(\w+)\s+(\d+)', date_cell)
            if not m:
                continue
            month_str, day_str = m.group(1), m.group(2)

            # Determine year — use current year; switch to next year after December
            MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                      'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
            month_num = MONTHS.get(month_str.lower()[:3])
            if not month_num:
                continue

            if year_hint is None:
                year_hint = timezone.now().year
            # When we cross from Dec → Jan, advance the year
            if results and month_num == 1 and results[-1]['date'].month == 12:
                year_hint += 1

            try:
                date_obj = dt(year_hint, month_num, int(day_str))
            except ValueError:
                continue

            # Skip completely blank rows (no event, no working-day marker)
            if not event_cell and not work_cell:
                continue

            etype = _guess_type(event_cell, work_cell)
            title = event_cell if event_cell and event_cell.lower() != 'week off' else (
                'Week Off' if 'week off' in event_cell.lower() else
                date_obj.strftime('%A')
            )
            if not title:
                title = etype.replace('_', ' ').title()

            results.append({'date': date_obj, 'event_type': etype, 'title': title})

    return results

@login_required
def academic_calendar(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    import calendar as cal_mod
    import csv
    import io
    from datetime import date as dt, timedelta

    if request.method == 'POST':
        action = request.POST.get('action', 'add')

        # ── Delete single event ──────────────────────────────────────────────
        if action == 'delete':
            get_object_or_404(AcademicCalendarEvent, pk=request.POST.get('pk')).delete()
            messages.success(request, 'Event removed.')
            return redirect('academic_calendar')

        # ── Mark a date range ────────────────────────────────────────────────
        if action == 'add_range':
            from_str   = request.POST.get('range_from', '').strip()
            to_str     = request.POST.get('range_to', '').strip()
            event_type = request.POST.get('range_event_type', 'holiday')
            title      = request.POST.get('range_title', '').strip()
            if not from_str or not to_str or not title:
                messages.error(request, 'From date, to date, and title are required.')
                return redirect('academic_calendar')
            from_date = dt.fromisoformat(from_str)
            to_date   = dt.fromisoformat(to_str)
            if to_date < from_date:
                messages.error(request, 'To date must be on or after From date.')
                return redirect('academic_calendar')
            count = 0
            cursor = from_date
            while cursor <= to_date:
                AcademicCalendarEvent.objects.update_or_create(
                    date=cursor,
                    defaults={'event_type': event_type, 'title': title,
                              'description': request.POST.get('range_description', '').strip(),
                              'marked_by': request.user}
                )
                cursor += timedelta(days=1)
                count += 1
            messages.success(request, f'{count} days marked as "{title}".')
            return redirect('academic_calendar')

        # ── File upload (Excel .xlsx or CSV) ────────────────────────────────
        if action == 'upload_csv':
            f = request.FILES.get('csv_file')
            if not f:
                messages.error(request, 'No file selected.')
                return redirect('academic_calendar')

            fname = f.name.lower()

            # ── Excel upload (PTU format or simple xlsx) ──────────────────
            if fname.endswith('.xlsx') or fname.endswith('.xls'):
                try:
                    records = _parse_ptu_excel(f)
                    if not records:
                        messages.error(request, 'No valid calendar data found in the Excel file. Make sure the header row contains "Date/Day" columns.')
                        return redirect('academic_calendar')
                    created = 0
                    for rec in records:
                        AcademicCalendarEvent.objects.update_or_create(
                            date=rec['date'],
                            defaults={
                                'event_type': rec['event_type'],
                                'title':      rec['title'],
                                'marked_by':  request.user,
                            }
                        )
                        created += 1
                    messages.success(request, f'Academic calendar imported: {created} day(s) processed from Excel.')
                except Exception as e:
                    messages.error(request, f'Excel parse error: {e}')
                return redirect('academic_calendar')

            # ── CSV upload (simple format) ────────────────────────────────
            try:
                text    = f.read().decode('utf-8-sig')
                reader  = csv.DictReader(io.StringIO(text))
                created = skipped = 0
                errors  = []
                for i, row in enumerate(reader, start=2):
                    from_str = (row.get('from_date') or row.get('date', '')).strip()
                    to_str   = (row.get('to_date') or row.get('date', '')).strip()
                    etype    = row.get('type', 'holiday').strip().lower()
                    title    = row.get('title', '').strip()
                    desc     = row.get('description', '').strip()
                    if not from_str or not title:
                        skipped += 1
                        continue
                    valid_types = {c[0] for c in AcademicCalendarEvent.EventType.choices}
                    if etype not in valid_types:
                        errors.append(f'Row {i}: unknown type "{etype}"')
                        skipped += 1
                        continue
                    try:
                        from_date = dt.fromisoformat(from_str)
                        to_date   = dt.fromisoformat(to_str) if to_str else from_date
                    except ValueError:
                        errors.append(f'Row {i}: invalid date "{from_str}"')
                        skipped += 1
                        continue
                    cursor = from_date
                    while cursor <= to_date:
                        AcademicCalendarEvent.objects.update_or_create(
                            date=cursor,
                            defaults={'event_type': etype, 'title': title,
                                      'description': desc, 'marked_by': request.user}
                        )
                        cursor += timedelta(days=1)
                        created += 1
                msg = f'Imported {created} day(s).'
                if skipped:
                    msg += f' {skipped} row(s) skipped.'
                if errors:
                    msg += ' Errors: ' + '; '.join(errors[:3])
                messages.success(request, msg)
            except Exception as e:
                messages.error(request, f'CSV parse error: {e}')
            return redirect('academic_calendar')

        # ── Mark single date ─────────────────────────────────────────────────
        date_val   = request.POST.get('date', '').strip()
        event_type = request.POST.get('event_type', 'holiday')
        title      = request.POST.get('title', '').strip()
        if not date_val or not title:
            messages.error(request, 'Date and title required.')
            return redirect('academic_calendar')
        AcademicCalendarEvent.objects.update_or_create(
            date=date_val,
            defaults={'event_type': event_type, 'title': title,
                      'description': request.POST.get('description', '').strip(),
                      'marked_by': request.user}
        )
        messages.success(request, f'Event saved for {date_val}.')
        return redirect('academic_calendar')

    year  = int(request.GET.get('year',  timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    if month < 1:  month, year = 12, year - 1
    if month > 12: month, year = 1, year + 1

    first_day = dt(year, month, 1)
    events = AcademicCalendarEvent.objects.filter(date__year=year, date__month=month)
    events_by_date = {e.date.day: e for e in events}
    cal_grid = cal_mod.monthcalendar(year, month)

    return render(request, 'admin/academic_calendar.html', {
        'year': year, 'month': month,
        'month_name': first_day.strftime('%B'),
        'cal': cal_grid, 'events_by_date': events_by_date,
        'recent_events': AcademicCalendarEvent.objects.order_by('-date')[:20],
        'event_types': AcademicCalendarEvent.EventType.choices,
        'today': timezone.now().date(),
        'prev_year':  year if month > 1 else year - 1,
        'prev_month': month - 1 if month > 1 else 12,
        'next_year':  year if month < 12 else year + 1,
        'next_month': month + 1 if month < 12 else 1,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# LEAVE CATEGORY MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def academic_calendar_sample(request):
    """Download a sample CSV template for bulk academic calendar import."""
    if request.user.role != User.Role.SUPER_ADMIN:
        return redirect('dashboard')
    import csv
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="academic_calendar_sample.csv"'
    writer = csv.writer(response)
    writer.writerow(['from_date', 'to_date', 'type', 'title', 'description'])
    writer.writerow(['2025-06-01', '2025-06-01', 'working', 'Classes Begin', 'Semester 1 starts'])
    writer.writerow(['2025-08-15', '2025-08-15', 'holiday', 'Independence Day', 'National Holiday'])
    writer.writerow(['2025-10-02', '2025-10-02', 'holiday', 'Gandhi Jayanti', ''])
    writer.writerow(['2025-10-20', '2025-10-30', 'exam', 'Mid Semester Exams', ''])
    writer.writerow(['2025-12-25', '2026-01-05', 'holiday', 'Winter Break', ''])
    writer.writerow(['2026-01-06', '2026-01-06', 'working', 'Classes Resume', 'Semester 2 begins'])
    writer.writerow(['2026-01-26', '2026-01-26', 'holiday', 'Republic Day', 'National Holiday'])
    return response


@login_required
def leave_category_list(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        name   = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Name required.')
            return redirect('leave_category_list')
        cat = get_object_or_404(LeaveCategory, pk=request.POST.get('pk')) if action == 'edit' else LeaveCategory(created_by=request.user)
        cat.name        = name
        cat.description = request.POST.get('description', '').strip()
        cat.max_days    = int(request.POST.get('max_days', 7) or 7)
        cat.start_time  = request.POST.get('start_time') or None
        cat.end_time    = request.POST.get('end_time') or None
        cat.requires_warden_approval     = 'req_warden' in request.POST
        cat.requires_hod_approval        = 'req_hod' in request.POST
        cat.requires_admin_approval      = 'req_admin' in request.POST
        cat.requires_superadmin_approval = 'req_superadmin' in request.POST
        cat.is_active = 'is_active' in request.POST
        cat.save()
        messages.success(request, f'Leave category "{cat.name}" saved.')
        return redirect('leave_category_list')
    categories = LeaveCategory.objects.annotate(leave_count=Count('leave_applications'))
    return render(request, 'admin/leave_categories.html', {'categories': categories})


@login_required
@require_POST
def leave_category_toggle(request, pk):
    if request.user.role != User.Role.SUPER_ADMIN:
        return redirect('dashboard')
    cat = get_object_or_404(LeaveCategory, pk=pk)
    cat.is_active = not cat.is_active
    cat.save()
    return redirect('leave_category_list')


@login_required
def leave_extension_request(request, leave_pk):
    student, redir = _get_student_or_redirect(request)
    if redir: return redir
    leave = get_object_or_404(LeaveApplication, pk=leave_pk, student=student, status='approved')
    if request.method == 'POST':
        new_to_date = request.POST.get('new_to_date', '').strip()
        reason      = request.POST.get('reason', '').strip()
        if not new_to_date or not reason:
            messages.error(request, 'New date and reason required.')
            return redirect('leave_extension_request', leave_pk=leave_pk)
        from datetime import date as dt
        try:
            ndate = dt.fromisoformat(new_to_date)
        except ValueError:
            messages.error(request, 'Invalid date.')
            return redirect('leave_extension_request', leave_pk=leave_pk)
        if ndate <= leave.to_date:
            messages.error(request, 'New date must be after current end date.')
            return redirect('leave_extension_request', leave_pk=leave_pk)
        LeaveExtensionRequest.objects.create(leave=leave, new_to_date=ndate, reason=reason)
        send_notification(_staff_targets(student),
                          f'Leave Extension — {student.name}',
                          f'{student.name} requested extension until {ndate}.',
                          notif_type='leave', link=f'/leave/{leave.pk}/action/')
        messages.success(request, 'Extension request submitted.')
        return redirect('my_leaves')
    return render(request, 'student/leave_extension.html', {'leave': leave})


@login_required
@require_POST
def leave_extension_action(request, ext_pk):
    role = request.user.role
    if role not in (User.Role.WARDEN, User.Role.ADMIN, User.Role.SUPER_ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    ext    = get_object_or_404(LeaveExtensionRequest, pk=ext_pk, status='pending')
    action = request.POST.get('action', '')
    if action == 'approve':
        ext.status = LeaveExtensionRequest.Status.APPROVED
        ext.approved_by = request.user
        ext.approved_at = timezone.now()
        ext.remarks     = request.POST.get('remarks', '').strip()
        ext.save()
        ext.leave.to_date = ext.new_to_date
        ext.leave.save(update_fields=['to_date'])
        messages.success(request, f'Extension approved to {ext.new_to_date}.')
    elif action == 'reject':
        ext.status  = LeaveExtensionRequest.Status.REJECTED
        ext.remarks = request.POST.get('remarks', '').strip()
        ext.save()
        messages.info(request, 'Extension rejected.')
    return redirect('leave_list')


@login_required
@require_POST
def leave_warden_action(request, pk):
    if request.user.role not in (User.Role.WARDEN, User.Role.SUPER_ADMIN, User.Role.ADMIN):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    leave   = get_object_or_404(LeaveApplication, pk=pk)
    action  = request.POST.get('action', '')
    remarks = request.POST.get('remarks', '').strip()

    # Block warden action on working-day leaves that haven't had HOD approval yet
    hod = _get_hod(leave.student.department)
    if leave.is_working_day_leave and hod and leave.hod_status != 'approved':
        messages.error(request, 'This is a working-day leave — HOD must approve it first.')
        return redirect('leave_list')

    if action == 'warden_approve':
        leave.warden_status      = 'approved'
        leave.warden_approved_by = request.user
        leave.warden_approved_at = timezone.now()
        leave.admin_remarks      = remarks
        cat = leave.category
        if not cat or not cat.requires_admin_approval:
            leave.status      = LeaveApplication.Status.APPROVED
            leave.reviewed_by = request.user
            leave.reviewed_at = timezone.now()
        leave.save()
        messages.success(request, 'Leave approved by warden.')
    elif action == 'warden_reject':
        leave.warden_status      = 'rejected'
        leave.warden_approved_by = request.user
        leave.warden_approved_at = timezone.now()
        leave.status             = LeaveApplication.Status.REJECTED
        leave.admin_remarks      = remarks
        leave.reviewed_by        = request.user
        leave.reviewed_at        = timezone.now()
        leave.save()
        messages.info(request, 'Leave rejected.')
    return redirect('leave_list')


# ═══════════════════════════════════════════════════════════════════════════════
# STUDENT RESTRICTIONS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def restriction_list(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    q = request.GET.get('q', '').strip()
    filter_type = request.GET.get('type', '')
    filter_status = request.GET.get('status', 'active')

    qs = StudentRestriction.objects.select_related('student', 'restricted_by', 'lifted_by').order_by('-restricted_at')
    if q:
        qs = qs.filter(Q(student__name__icontains=q) | Q(student__roll_number__icontains=q))
    if filter_type:
        qs = qs.filter(restriction_type=filter_type)
    if filter_status == 'active':
        qs = qs.filter(is_active=True)
    elif filter_status == 'lifted':
        qs = qs.filter(is_active=False)

    students_qs = Student.objects.filter(allocations__status='active').select_related().distinct().order_by('name')
    if q:
        students_qs = students_qs.filter(Q(name__icontains=q) | Q(roll_number__icontains=q))

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'superadmin/restrictions.html', {
        'page': page,
        'q': q,
        'filter_type': filter_type,
        'filter_status': filter_status,
        'restriction_types': StudentRestriction.RestrictionType.choices,
        'students': students_qs[:50],
        'active_count': StudentRestriction.objects.filter(is_active=True).count(),
    })


@login_required
@require_POST
def restriction_add(request):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    student_id      = request.POST.get('student_id')
    restriction_type = request.POST.get('restriction_type')
    reason          = request.POST.get('reason', '').strip()

    if not student_id or not restriction_type or not reason:
        messages.error(request, 'Student, restriction type, and reason are all required.')
        return redirect('restriction_list')

    student = get_object_or_404(Student, pk=student_id)

    # Deactivate any same-type existing active restriction for this student
    StudentRestriction.objects.filter(
        student=student, is_active=True,
        restriction_type=restriction_type
    ).update(is_active=False, lifted_by=request.user, lifted_at=timezone.now(),
              lift_reason='Superseded by new restriction')

    StudentRestriction.objects.create(
        student=student,
        restriction_type=restriction_type,
        reason=reason,
        restricted_by=request.user,
        is_active=True,
    )

    # Notify the student
    if student.user:
        label = dict(StudentRestriction.RestrictionType.choices).get(restriction_type, restriction_type)
        send_notification(
            [student.user],
            f'Restriction Applied — {label}',
            f'You have been restricted from {label.lower()} applications. Reason: {reason}',
            notif_type='system',
        )

    messages.success(request, f'Restriction applied for {student.name}.')
    return redirect('restriction_list')


@login_required
@require_POST
def restriction_lift(request, pk):
    if request.user.role != User.Role.SUPER_ADMIN:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    restriction = get_object_or_404(StudentRestriction, pk=pk)
    lift_reason = request.POST.get('lift_reason', '').strip()
    if not lift_reason:
        messages.error(request, 'Please provide a reason for lifting the restriction.')
        return redirect('restriction_list')

    restriction.is_active   = False
    restriction.lifted_by   = request.user
    restriction.lifted_at   = timezone.now()
    restriction.lift_reason = lift_reason
    restriction.save()

    if restriction.student.user:
        label = restriction.get_restriction_type_display()
        send_notification(
            [restriction.student.user],
            f'Restriction Lifted — {label}',
            f'Your {label.lower()} restriction has been lifted. You may now apply again.',
            notif_type='system',
        )

    messages.success(request, f'Restriction lifted for {restriction.student.name}.')
    return redirect('restriction_list')


# ═══════════════════════════════════════════════════════════════════════════════
#  DOWNLOAD REPORTS CENTER
# ═══════════════════════════════════════════════════════════════════════════════

import io as _io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def _pdf_response(filename):
    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _xlsx_response(filename):
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _xlsx_header_style():
    fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=11)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return fill, font, align


def _pdf_base_style():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Heading1'],
                                  fontSize=16, textColor=rl_colors.HexColor('#1565C0'),
                                  spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle('ReportSub', parent=styles['Normal'],
                                fontSize=9, textColor=rl_colors.HexColor('#64748B'),
                                spaceAfter=12, alignment=TA_CENTER)
    return styles, title_style, sub_style


def _pdf_table_style(header_color='#1565C0'):
    hc = rl_colors.HexColor(header_color)
    return TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0), hc),
        ('TEXTCOLOR',   (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 8),
        ('ALIGN',       (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 1), (-1, -1), 7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F8FAFC')]),
        ('GRID',        (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#CBD5E1')),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
    ])


@login_required
@superadmin_only
def resident_analysis(request):
    from apps.accounts.models import User as _User
    view_mode = request.GET.get('view', '')   # 'boys' | 'girls' | ''
    hostel_pk = request.GET.get('hostel', '')
    entry_type = request.GET.get('entry', '') # 'incampus' | 'outcampus' (for boys drill-down)
    selected  = Hostel.objects.filter(pk=hostel_pk).first() if hostel_pk else None

    all_active = Student.objects.filter(allocations__status='active').distinct()
    summary = {
        'total':         all_active.count(),
        'boys_total':    all_active.filter(gender='male').count(),
        'boys_incampus': all_active.filter(gender='male', allocations__room__designation='incampus', allocations__status='active').distinct().count(),
        'boys_outcampus':all_active.filter(gender='male', allocations__room__designation='outcampus', allocations__status='active').distinct().count(),
        'girls_total':   all_active.filter(gender='female').count(),
        'staff_total':   _User.objects.exclude(role__in=('superadmin','student')).filter(is_active=True).count(),
        'warden_count':  _User.objects.filter(role='warden', is_active=True).count(),
    }

    def _attach_rooms(students):
        alloc_map = {
            a.student_id: a
            for a in RoomAllocation.objects.filter(
                student__in=students, status='active'
            ).select_related('room')
        }
        for s in students:
            s.alloc_room = alloc_map.get(s.pk, None) and alloc_map[s.pk].room
        return students

    # ── Boys view ──────────────────────────────────────────────────────────────
    boys_incampus_rows  = []
    boys_outcampus_rows = []
    boys_incampus_total = boys_outcampus_total = 0
    incampus_students = outcampus_students = None

    if view_mode == 'boys':
        for h in Hostel.objects.order_by('name'):
            inc  = Student.objects.filter(allocations__room__hostel=h, allocations__status='active', allocations__room__designation='incampus', gender='male').distinct().count()
            outc = Student.objects.filter(allocations__room__hostel=h, allocations__status='active', allocations__room__designation='outcampus', gender='male').distinct().count()
            if inc:
                boys_incampus_rows.append({'hostel': h, 'count': inc})
                boys_incampus_total += inc
            if outc:
                boys_outcampus_rows.append({'hostel': h, 'count': outc})
                boys_outcampus_total += outc

        if selected and entry_type == 'incampus':
            incampus_students = list(Student.objects.filter(
                allocations__room__hostel=selected, allocations__status='active',
                allocations__room__designation='incampus', gender='male'
            ).distinct().order_by('roll_number'))
            _attach_rooms(incampus_students)

        elif selected and entry_type == 'outcampus':
            outcampus_students = list(Student.objects.filter(
                allocations__room__hostel=selected, allocations__status='active',
                allocations__room__designation='outcampus', gender='male'
            ).distinct().order_by('roll_number'))
            _attach_rooms(outcampus_students)

        elif selected:
            # no entry filter — show all boys in that hostel
            incampus_students = list(Student.objects.filter(
                allocations__room__hostel=selected, allocations__status='active',
                allocations__room__designation='incampus', gender='male'
            ).distinct().order_by('roll_number'))
            outcampus_students = list(Student.objects.filter(
                allocations__room__hostel=selected, allocations__status='active',
                allocations__room__designation='outcampus', gender='male'
            ).distinct().order_by('roll_number'))
            _attach_rooms(incampus_students)
            _attach_rooms(outcampus_students)

    # ── Girls view ─────────────────────────────────────────────────────────────
    girls_rows = []
    girls_students = None
    if view_mode == 'girls':
        for h in Hostel.objects.order_by('name'):
            cnt = Student.objects.filter(
                allocations__room__hostel=h, allocations__status='active', gender='female'
            ).distinct().count()
            girls_rows.append({'hostel': h, 'count': cnt})

        if selected:
            girls_students = list(Student.objects.filter(
                allocations__room__hostel=selected, allocations__status='active', gender='female'
            ).distinct().order_by('roll_number'))
            _attach_rooms(girls_students)

    # ── General hostel rows (overview page) ────────────────────────────────────
    hostel_rows = []
    hostel_students = None
    if not view_mode:
        for h in Hostel.objects.order_by('name'):
            qs = Student.objects.filter(allocations__room__hostel=h, allocations__status='active').distinct()
            boys   = qs.filter(gender='male').count()
            girls  = qs.filter(gender='female').count()
            inc    = qs.filter(gender='male', type_of_entry='incampus').count()
            outc   = qs.filter(gender='male', type_of_entry='outcampus').count()
            wardens = _User.objects.filter(role='warden', managed_hostels=h, is_active=True).count()
            hostel_rows.append({
                'hostel': h, 'total': boys + girls,
                'boys': boys, 'girls': girls,
                'boys_incampus': inc, 'boys_outcampus': outc,
                'wardens': wardens,
            })
        if selected:
            hostel_students = list(Student.objects.filter(
                allocations__room__hostel=selected, allocations__status='active'
            ).distinct().order_by('gender', 'roll_number'))
            _attach_rooms(hostel_students)

    return render(request, 'admin/resident_analysis.html', {
        'summary':              summary,
        'view_mode':            view_mode,
        'selected':             selected,
        'entry_type':           entry_type,
        # overview
        'hostel_rows':          hostel_rows,
        'hostel_students':      hostel_students,
        # boys
        'boys_incampus_rows':   boys_incampus_rows,
        'boys_outcampus_rows':  boys_outcampus_rows,
        'boys_incampus_total':  boys_incampus_total,
        'boys_outcampus_total': boys_outcampus_total,
        'incampus_students':    incampus_students,
        'outcampus_students':   outcampus_students,
        # girls
        'girls_rows':           girls_rows,
        'girls_students':       girls_students,
    })


@reports_download_access
def reports_download_center(request):
    hostels = Hostel.objects.all().order_by('name')
    return render(request, 'admin/reports_download.html', {'hostels': hostels})


# ── 1. Room Transfer Report ───────────────────────────────────────────────────

@reports_download_access
def report_room_transfer(request):
    fmt = request.GET.get('format', 'xlsx')
    qs = (RoomAllocation.objects
          .filter(status='checkout')
          .select_related('student', 'room__hostel', 'allocated_by')
          .order_by('-check_out'))

    headers = ['#', 'Roll No.', 'Student Name', 'Department', 'Hostel', 'Room No.',
               'Check-In', 'Check-Out', 'Allocated By', 'Notes']

    rows = []
    for i, a in enumerate(qs, 1):
        rows.append([
            i,
            a.student.roll_number,
            a.student.name,
            a.student.department,
            a.room.hostel.name,
            a.room.room_number,
            str(a.check_in),
            str(a.check_out) if a.check_out else '—',
            a.allocated_by.name if a.allocated_by else '—',
            a.notes or '—',
        ])

    today = dt_date.today()

    if fmt == 'pdf':
        resp = _pdf_response(f'room_transfer_report_{today}.pdf')
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm,
                                 bottomMargin=1*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles, ts, ss = _pdf_base_style()
        elems = [
            Paragraph('Room Transfer Report', ts),
            Paragraph(f'Generated: {today}  |  Total Records: {len(rows)}', ss),
            HRFlowable(color=rl_colors.HexColor('#1565C0'), thickness=1, spaceAfter=8),
        ]
        col_w = [0.5*cm, 2.2*cm, 3.5*cm, 3*cm, 3*cm, 2*cm, 2.2*cm, 2.2*cm, 3*cm, 3*cm]
        data = [headers] + rows
        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(_pdf_table_style())
        elems.append(t)
        doc.build(elems)
        resp.write(buf.getvalue())
        return resp

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Room Transfers'
    fill, font, align = _xlsx_header_style()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = fill; cell.font = font; cell.alignment = align
    for row in rows:
        ws.append(row)
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    buf = _io.BytesIO()
    wb.save(buf)
    resp = _xlsx_response(f'room_transfer_report_{today}.xlsx')
    resp.write(buf.getvalue())
    return resp


# ── 2. Daily Attendance Report ───────────────────────────────────────────────

@reports_download_access
def report_attendance(request):
    fmt = request.GET.get('format', 'xlsx')
    date_str = request.GET.get('date', str(dt_date.today()))
    try:
        report_date = dt_date.fromisoformat(date_str)
    except ValueError:
        report_date = dt_date.today()

    qs = (Attendance.objects
          .filter(date=report_date)
          .select_related('student', 'marked_by')
          .prefetch_related('student__allocations__room__hostel')
          .order_by('student__roll_number'))

    headers = ['#', 'Roll No.', 'Student Name', 'Department', 'Year', 'Hostel', 'Room', 'Status', 'Marked By', 'Remarks']

    rows = []
    for i, att in enumerate(qs, 1):
        alloc = att.student.allocations.filter(status='active').select_related('room__hostel').first()
        hostel = alloc.room.hostel.name if alloc else '—'
        room = alloc.room.room_number if alloc else '—'
        rows.append([
            i, att.student.roll_number, att.student.name,
            att.student.department, att.student.get_year_display(),
            hostel, room,
            att.get_status_display(),
            att.marked_by.name if att.marked_by else '—',
            att.remarks or '—',
        ])

    if fmt == 'pdf':
        resp = _pdf_response(f'attendance_report_{report_date}.pdf')
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm,
                                 bottomMargin=1*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles, ts, ss = _pdf_base_style()
        present = sum(1 for r in rows if r[7] == 'Present')
        absent  = sum(1 for r in rows if r[7] == 'Absent')
        on_leave = sum(1 for r in rows if r[7] == 'On Leave')
        elems = [
            Paragraph('Daily Attendance Report', ts),
            Paragraph(f'Date: {report_date}  |  Present: {present}  Absent: {absent}  On Leave: {on_leave}  Total: {len(rows)}', ss),
            HRFlowable(color=rl_colors.HexColor('#1565C0'), thickness=1, spaceAfter=8),
        ]
        col_w = [0.5*cm, 2.2*cm, 3.5*cm, 3*cm, 1.5*cm, 3*cm, 1.8*cm, 2*cm, 2.8*cm, 3.5*cm]
        data = [headers] + rows
        t = Table(data, colWidths=col_w, repeatRows=1)
        ts_style = _pdf_table_style()
        for i2, row in enumerate(rows, 1):
            if row[7] == 'Absent':
                ts_style.add('BACKGROUND', (0, i2), (-1, i2), rl_colors.HexColor('#FEF2F2'))
            elif row[7] == 'On Leave':
                ts_style.add('BACKGROUND', (0, i2), (-1, i2), rl_colors.HexColor('#FFFBEB'))
        t.setStyle(ts_style)
        elems.append(t)
        doc.build(elems)
        resp.write(buf.getvalue())
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Attendance {report_date}'
    fill, font, align = _xlsx_header_style()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = fill; cell.font = font; cell.alignment = align
    absent_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    leave_fill  = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    for ri, row in enumerate(rows, 2):
        ws.append(row)
        if row[7] == 'Absent':
            for ci in range(1, len(headers)+1):
                ws.cell(ri, ci).fill = absent_fill
        elif row[7] == 'On Leave':
            for ci in range(1, len(headers)+1):
                ws.cell(ri, ci).fill = leave_fill
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    buf = _io.BytesIO()
    wb.save(buf)
    resp = _xlsx_response(f'attendance_report_{report_date}.xlsx')
    resp.write(buf.getvalue())
    return resp


# ── 3. Fee / Fine Report ──────────────────────────────────────────────────────

@reports_download_access
def report_fees(request):
    fmt = request.GET.get('format', 'xlsx')
    qs = (FeeRecord.objects
          .select_related('student', 'fee_structure')
          .order_by('fee_structure__academic_year', 'fee_structure__semester', 'student__roll_number'))

    headers = ['#', 'Roll No.', 'Student', 'Department', 'Fee Structure',
               'Total (₹)', 'Paid (₹)', 'Balance (₹)', 'Status', 'Paid On', 'Mode', 'Ref No.']

    rows = []
    for i, rec in enumerate(qs, 1):
        rows.append([
            i, rec.student.roll_number, rec.student.name, rec.student.department,
            str(rec.fee_structure),
            float(rec.total_amount), float(rec.amount_paid), float(rec.balance),
            rec.get_status_display(),
            str(rec.paid_on) if rec.paid_on else '—',
            rec.payment_mode or '—', rec.transaction_ref or '—',
        ])

    today = dt_date.today()
    total_due = sum(r[7] for r in rows)
    total_collected = sum(r[6] for r in rows)

    if fmt == 'pdf':
        resp = _pdf_response(f'fee_report_{today}.pdf')
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm,
                                 bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm)
        styles, ts, ss = _pdf_base_style()
        elems = [
            Paragraph('Fee Report', ts),
            Paragraph(f'Generated: {today}  |  Total Collected: ₹{total_collected:,.2f}  |  Outstanding: ₹{total_due:,.2f}', ss),
            HRFlowable(color=rl_colors.HexColor('#1565C0'), thickness=1, spaceAfter=8),
        ]
        col_w = [0.5*cm, 2.2*cm, 3.5*cm, 2.8*cm, 3.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2.5*cm]
        data = [headers] + rows
        t = Table(data, colWidths=col_w, repeatRows=1)
        ts_style = _pdf_table_style()
        for i2, row in enumerate(rows, 1):
            if row[8] == 'Pending':
                ts_style.add('BACKGROUND', (0, i2), (-1, i2), rl_colors.HexColor('#FEF2F2'))
            elif row[8] == 'Partially Paid':
                ts_style.add('BACKGROUND', (0, i2), (-1, i2), rl_colors.HexColor('#FFFBEB'))
        t.setStyle(ts_style)
        elems.append(t)
        doc.build(elems)
        resp.write(buf.getvalue())
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fee Report'
    fill, font, align = _xlsx_header_style()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = fill; cell.font = font; cell.alignment = align
    for row in rows:
        ws.append(row)
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    buf = _io.BytesIO()
    wb.save(buf)
    resp = _xlsx_response(f'fee_report_{today}.xlsx')
    resp.write(buf.getvalue())
    return resp


# ── 4. Gate Pass Report ───────────────────────────────────────────────────────

@reports_download_access
def report_gatepass(request):
    fmt = request.GET.get('format', 'xlsx')
    date_from = request.GET.get('from', '')
    date_to   = request.GET.get('to', '')
    today = dt_date.today()

    qs = GatePass.objects.select_related('student', 'approved_by', 'category')\
        .prefetch_related('student__allocations__room__hostel')\
        .order_by('-departure_date')
    if date_from:
        try: qs = qs.filter(departure_date__gte=dt_date.fromisoformat(date_from))
        except ValueError: pass
    if date_to:
        try: qs = qs.filter(departure_date__lte=dt_date.fromisoformat(date_to))
        except ValueError: pass

    headers = ['#', 'Roll No.', 'Student', 'Hostel', 'Type', 'Category',
               'Destination', 'Departure', 'Exp. Return', 'Actual Return',
               'Status', 'Overstayed', 'Approved By']

    rows = []
    for i, gp in enumerate(qs, 1):
        alloc = gp.student.allocations.filter(status='active').select_related('room__hostel').first()
        hostel = alloc.room.hostel.name if alloc else '—'
        rows.append([
            i, gp.student.roll_number, gp.student.name, hostel,
            gp.get_pass_type_display(),
            gp.category.name if gp.category else '—',
            gp.destination,
            str(gp.departure_date), str(gp.expected_return_date),
            str(gp.actual_return_date) if gp.actual_return_date else '—',
            gp.get_status_display(),
            'Yes' if gp.is_overstayed else 'No',
            gp.approved_by.name if gp.approved_by else '—',
        ])

    suffix = f'{date_from}_to_{date_to}' if date_from or date_to else str(today)

    if fmt == 'pdf':
        resp = _pdf_response(f'gatepass_report_{suffix}.pdf')
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm,
                                 bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm)
        styles, ts, ss = _pdf_base_style()
        elems = [
            Paragraph('Gate Pass Report', ts),
            Paragraph(f'Period: {date_from or "All"} to {date_to or "All"}  |  Total: {len(rows)}', ss),
            HRFlowable(color=rl_colors.HexColor('#1565C0'), thickness=1, spaceAfter=8),
        ]
        col_w = [0.5*cm, 2*cm, 3.2*cm, 2.8*cm, 2*cm, 2.5*cm, 3*cm, 2*cm, 2*cm, 2.2*cm, 1.8*cm, 1.8*cm, 2.5*cm]
        data = [headers] + rows
        t = Table(data, colWidths=col_w, repeatRows=1)
        ts_style = _pdf_table_style()
        for i2, row in enumerate(rows, 1):
            if row[11] == 'Yes':
                ts_style.add('BACKGROUND', (0, i2), (-1, i2), rl_colors.HexColor('#FEF2F2'))
        t.setStyle(ts_style)
        elems.append(t)
        doc.build(elems)
        resp.write(buf.getvalue())
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gate Pass Report'
    fill, font, align = _xlsx_header_style()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = fill; cell.font = font; cell.alignment = align
    overstay_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    for ri, row in enumerate(rows, 2):
        ws.append(row)
        if row[11] == 'Yes':
            for ci in range(1, len(headers)+1):
                ws.cell(ri, ci).fill = overstay_fill
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    buf = _io.BytesIO()
    wb.save(buf)
    resp = _xlsx_response(f'gatepass_report_{suffix}.xlsx')
    resp.write(buf.getvalue())
    return resp


# ── 5. Mess Report (Wastage + Feedback) ───────────────────────────────────────

@reports_download_access
def report_mess(request):
    fmt = request.GET.get('format', 'xlsx')
    today = dt_date.today()

    # Wastage
    wastage_qs = (MessWastageRecord.objects
                  .select_related('hostel', 'filled_by')
                  .order_by('-date', 'hostel__name'))
    # Feedback
    feedback_qs = (MessFeedback.objects
                   .select_related('student', 'hostel')
                   .order_by('-date'))

    w_headers = ['#', 'Date', 'Hostel', 'Meal Type', 'Wastage (kg)', 'Filled By', 'Notes']
    f_headers = ['#', 'Date', 'Hostel', 'Meal Type', 'Roll No.', 'Student', 'Rating (1-5)', 'Comment']

    w_rows = [[i, str(w.date), w.hostel.name, w.meal_type.title(),
               float(w.wastage_kg), w.filled_by.name if w.filled_by else '—', w.notes or '—']
              for i, w in enumerate(wastage_qs, 1)]

    f_rows = [[i, str(fb.date), fb.hostel.name, fb.meal_type.title(),
               fb.student.roll_number, fb.student.name, fb.rating, fb.comment or '—']
              for i, fb in enumerate(feedback_qs, 1)]

    if fmt == 'pdf':
        resp = _pdf_response(f'mess_report_{today}.pdf')
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm,
                                 bottomMargin=1*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles, ts_sty, ss = _pdf_base_style()
        h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=12,
                             textColor=rl_colors.HexColor('#059669'), spaceBefore=14, spaceAfter=4)
        avg_rating = (sum(r[6] for r in f_rows) / len(f_rows)) if f_rows else 0
        total_waste = sum(r[4] for r in w_rows)
        elems = [
            Paragraph('Mess Report — Wastage & Feedback Analysis', ts_sty),
            Paragraph(f'Generated: {today}  |  Total Wastage: {total_waste:.2f} kg  |  Avg Feedback Rating: {avg_rating:.2f}/5', ss),
            HRFlowable(color=rl_colors.HexColor('#059669'), thickness=1, spaceAfter=10),
            Paragraph('Food Wastage Records', h2),
        ]
        if w_rows:
            col_w = [0.5*cm, 2*cm, 4*cm, 2.5*cm, 2.5*cm, 3.5*cm, 6.5*cm]
            t = Table([w_headers] + w_rows, colWidths=col_w, repeatRows=1)
            t.setStyle(_pdf_table_style('#059669'))
            elems.append(t)
        else:
            elems.append(Paragraph('No wastage records found.', styles['Normal']))

        elems += [Spacer(1, 0.5*cm), Paragraph('Student Feedback Records', h2)]
        if f_rows:
            col_w = [0.5*cm, 2*cm, 3.5*cm, 2.5*cm, 2.5*cm, 3.5*cm, 1.8*cm, 5.5*cm]
            t = Table([f_headers] + f_rows, colWidths=col_w, repeatRows=1)
            t.setStyle(_pdf_table_style('#059669'))
            elems.append(t)
        else:
            elems.append(Paragraph('No feedback records found.', styles['Normal']))

        doc.build(elems)
        resp.write(buf.getvalue())
        return resp

    wb = openpyxl.Workbook()
    # Sheet 1 — Wastage
    ws1 = wb.active
    ws1.title = 'Wastage'
    fill, font, align = _xlsx_header_style()
    green_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    for ci, h in enumerate(w_headers, 1):
        cell = ws1.cell(1, ci, h)
        cell.fill = green_fill; cell.font = font; cell.alignment = align
    for row in w_rows:
        ws1.append(row)
    ws1.freeze_panes = 'A2'
    # Sheet 2 — Feedback
    ws2 = wb.create_sheet('Feedback')
    for ci, h in enumerate(f_headers, 1):
        cell = ws2.cell(1, ci, h)
        cell.fill = green_fill; cell.font = font; cell.alignment = align
    for row in f_rows:
        ws2.append(row)
    ws2.freeze_panes = 'A2'
    for ws_item in [ws1, ws2]:
        for col in ws_item.columns:
            ws_item.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
    buf = _io.BytesIO()
    wb.save(buf)
    resp = _xlsx_response(f'mess_report_{today}.xlsx')
    resp.write(buf.getvalue())
    return resp


# ── 6. Daily Hostel Report (comprehensive per-hostel) ─────────────────────────

@reports_download_access
def report_hostel_daily(request, hostel_pk):
    fmt = request.GET.get('format', 'xlsx')
    date_str = request.GET.get('date', str(dt_date.today()))
    try:
        report_date = dt_date.fromisoformat(date_str)
    except ValueError:
        report_date = dt_date.today()

    hostel = get_object_or_404(Hostel, pk=hostel_pk)

    # ── data gathering ──
    rooms = Room.objects.filter(hostel=hostel)
    total_rooms    = rooms.count()
    total_capacity = sum(r.capacity for r in rooms)
    active_allocs  = RoomAllocation.objects.filter(room__hostel=hostel, status='active')\
                        .select_related('student', 'room')
    resident_count = active_allocs.count()
    vacant_capacity = total_capacity - resident_count

    # Attendance for date
    att_qs = (Attendance.objects
              .filter(date=report_date, student__allocations__room__hostel=hostel,
                      student__allocations__status='active')
              .select_related('student')
              .distinct())
    att_map = {a.student_id: a for a in att_qs}

    # Gate passes for date
    gp_qs = (GatePass.objects
             .filter(departure_date=report_date,
                     student__allocations__room__hostel=hostel,
                     student__allocations__status='active')
             .select_related('student')
             .distinct())

    # Complaints
    from apps.complaints.models import Complaint
    complaints_qs = (Complaint.objects
                     .filter(hostel=hostel, status__in=['open', 'in_progress'])
                     .select_related('raised_by'))

    # Fees — pending/partial in this hostel
    fee_qs = (FeeRecord.objects
              .filter(status__in=['pending', 'partial'],
                      student__allocations__room__hostel=hostel,
                      student__allocations__status='active')
              .select_related('student', 'fee_structure')
              .distinct())

    # Mess feedback
    feedback_qs = MessFeedback.objects.filter(hostel=hostel, date=report_date)
    avg_rating = feedback_qs.aggregate(avg=Avg('rating'))['avg'] or 0

    # Wastage
    wastage_qs = MessWastageRecord.objects.filter(hostel=hostel, date=report_date)
    total_wastage = wastage_qs.aggregate(t=Sum('wastage_kg'))['t'] or 0

    # Maintenance open
    maint_qs = MaintenanceLog.objects.filter(
        Q(room__hostel=hostel) | Q(hostel=hostel),
        status__in=['open', 'in_progress']
    ).select_related('room') if hasattr(MaintenanceLog, 'hostel') else \
    MaintenanceLog.objects.filter(room__hostel=hostel, status__in=['open', 'in_progress'])\
        .select_related('room')

    today = dt_date.today()

    if fmt == 'pdf':
        resp = _pdf_response(f'{hostel.name}_daily_report_{report_date}.pdf')
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm,
                                 bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        styles, ts, ss = _pdf_base_style()
        h2 = ParagraphStyle('H2r', parent=styles['Heading2'], fontSize=11,
                             textColor=rl_colors.HexColor('#1565C0'), spaceBefore=14, spaceAfter=4)
        normal = styles['Normal']
        normal.fontSize = 8

        def section(title):
            return [Paragraph(title, h2),
                    HRFlowable(color=rl_colors.HexColor('#CBD5E1'), thickness=0.5, spaceAfter=5)]

        att_present  = sum(1 for a in att_qs if a.status == 'present')
        att_absent   = sum(1 for a in att_qs if a.status == 'absent')
        att_leave    = sum(1 for a in att_qs if a.status == 'leave')
        att_unmarked = resident_count - len(att_map)

        elems = [
            Paragraph(f'{hostel.name} — Daily Report', ts),
            Paragraph(f'Date: {report_date}  |  Generated: {today}', ss),
            HRFlowable(color=rl_colors.HexColor('#1565C0'), thickness=1.5, spaceAfter=10),
        ]

        # Overview table
        overview = [
            ['Total Rooms', total_rooms, 'Total Capacity', total_capacity],
            ['Current Residents', resident_count, 'Vacant Beds', vacant_capacity],
            ['Present Today', att_present, 'Absent', att_absent],
            ['On Leave', att_leave, 'Unmarked', att_unmarked],
            ['Gate Passes Today', gp_qs.count(), 'Open Complaints', complaints_qs.count()],
            ['Pending Fees', fee_qs.count(), 'Open Maintenance', maint_qs.count()],
            ['Mess Avg Rating', f'{avg_rating:.2f}/5', 'Total Wastage', f'{float(total_wastage):.2f} kg'],
        ]
        ot = Table(overview, colWidths=[4.5*cm, 3*cm, 4.5*cm, 3*cm])
        ot.setStyle(TableStyle([
            ('BACKGROUND',  (0, 0), (0, -1), rl_colors.HexColor('#EFF6FF')),
            ('BACKGROUND',  (2, 0), (2, -1), rl_colors.HexColor('#EFF6FF')),
            ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME',    (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE',    (0, 0), (-1, -1), 8.5),
            ('GRID',        (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#CBD5E1')),
            ('TOPPADDING',  (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elems.append(ot)

        # Resident list
        elems += section('Resident List & Attendance')
        res_headers = ['#', 'Roll No.', 'Name', 'Room', 'Status']
        res_rows = []
        for i, alloc in enumerate(active_allocs.order_by('room__room_number', 'student__roll_number'), 1):
            att = att_map.get(alloc.student_id)
            status = att.get_status_display() if att else 'Not Marked'
            res_rows.append([i, alloc.student.roll_number, alloc.student.name,
                             alloc.room.room_number, status])
        if res_rows:
            rt = Table([res_headers] + res_rows, colWidths=[1*cm, 3*cm, 5.5*cm, 2.5*cm, 3*cm], repeatRows=1)
            rs_sty = _pdf_table_style()
            for i2, row in enumerate(res_rows, 1):
                if row[4] == 'Absent':
                    rs_sty.add('BACKGROUND', (0, i2), (-1, i2), rl_colors.HexColor('#FEF2F2'))
                elif row[4] == 'Not Marked':
                    rs_sty.add('BACKGROUND', (0, i2), (-1, i2), rl_colors.HexColor('#F8FAFC'))
            rt.setStyle(rs_sty)
            elems.append(rt)

        # Gate passes
        elems += section(f'Gate Passes — {report_date}')
        gp_headers = ['#', 'Roll No.', 'Student', 'Type', 'Destination', 'Exp. Return', 'Status']
        gp_rows = [[i, gp.student.roll_number, gp.student.name,
                    gp.get_pass_type_display(), gp.destination,
                    str(gp.expected_return_date), gp.get_status_display()]
                   for i, gp in enumerate(gp_qs, 1)]
        if gp_rows:
            gt = Table([gp_headers] + gp_rows, colWidths=[0.7*cm, 2.5*cm, 4*cm, 2.5*cm, 4*cm, 2.5*cm, 2*cm], repeatRows=1)
            gt.setStyle(_pdf_table_style())
            elems.append(gt)
        else:
            elems.append(Paragraph('No gate passes for this date.', normal))

        # Pending fees
        elems += section('Pending / Partial Fee Records')
        fee_headers = ['#', 'Roll No.', 'Student', 'Fee Structure', 'Total (₹)', 'Paid (₹)', 'Balance (₹)', 'Status']
        fee_rows = [[i, r.student.roll_number, r.student.name, str(r.fee_structure),
                     float(r.total_amount), float(r.amount_paid), float(r.balance), r.get_status_display()]
                    for i, r in enumerate(fee_qs, 1)]
        if fee_rows:
            ft = Table([fee_headers] + fee_rows, colWidths=[0.7*cm, 2.5*cm, 4*cm, 4*cm, 2*cm, 2*cm, 2*cm, 2*cm], repeatRows=1)
            ft.setStyle(_pdf_table_style())
            elems.append(ft)
        else:
            elems.append(Paragraph('No pending fee records.', normal))

        # Open complaints
        elems += section('Open Complaints')
        comp_headers = ['#', 'Title', 'Raised By', 'Category', 'Status', 'Date']
        comp_rows = [[i, c.title, c.raised_by.name if c.raised_by else '—',
                      c.category if hasattr(c, 'category') else '—',
                      c.get_status_display(), str(c.created_at.date())]
                     for i, c in enumerate(complaints_qs, 1)]
        if comp_rows:
            ct = Table([comp_headers] + comp_rows,
                       colWidths=[0.7*cm, 5*cm, 3.5*cm, 3*cm, 2.5*cm, 2.3*cm], repeatRows=1)
            ct.setStyle(_pdf_table_style())
            elems.append(ct)
        else:
            elems.append(Paragraph('No open complaints.', normal))

        doc.build(elems)
        resp.write(buf.getvalue())
        return resp

    # ── Excel version ──
    wb = openpyxl.Workbook()
    fill, hdr_font, hdr_align = _xlsx_header_style()

    def add_sheet(wb, title, headers, rows, hdr_color='1565C0'):
        ws = wb.create_sheet(title)
        hfill = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type='solid')
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.fill = hfill; cell.font = hdr_font; cell.alignment = hdr_align
        for row in rows:
            ws.append(row)
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4
        return ws

    # Remove default sheet
    del wb['Sheet']

    # Summary
    ws_sum = wb.create_sheet('Summary')
    summary_data = [
        ['GGI Hostel ERP — Daily Hostel Report'],
        [f'Hostel: {hostel.name}'],
        [f'Date: {report_date}'],
        [''],
        ['Metric', 'Value'],
        ['Total Rooms', total_rooms],
        ['Total Capacity', total_capacity],
        ['Current Residents', resident_count],
        ['Vacant Beds', vacant_capacity],
        ['Present Today', sum(1 for a in att_qs if a.status == 'present')],
        ['Absent Today', sum(1 for a in att_qs if a.status == 'absent')],
        ['On Leave Today', sum(1 for a in att_qs if a.status == 'leave')],
        ['Not Marked', resident_count - len(att_map)],
        ['Gate Passes Today', gp_qs.count()],
        ['Open Complaints', complaints_qs.count()],
        ['Pending Fee Records', fee_qs.count()],
        ['Open Maintenance', maint_qs.count()],
        ['Mess Avg Rating', round(avg_rating, 2)],
        ['Total Food Wastage (kg)', float(total_wastage)],
    ]
    for row in summary_data:
        ws_sum.append(row)
    ws_sum['A1'].font = Font(bold=True, size=14, color='1565C0')
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 20

    # Residents + Attendance
    res_headers = ['#', 'Roll No.', 'Name', 'Department', 'Room', 'Attendance Status', 'Attendance Remarks']
    res_rows = []
    for i, alloc in enumerate(active_allocs.order_by('room__room_number', 'student__roll_number'), 1):
        att = att_map.get(alloc.student_id)
        res_rows.append([i, alloc.student.roll_number, alloc.student.name,
                         alloc.student.department, alloc.room.room_number,
                         att.get_status_display() if att else 'Not Marked',
                         att.remarks if att else ''])
    add_sheet(wb, 'Residents & Attendance', res_headers, res_rows)

    # Gate Passes
    gp_headers = ['#', 'Roll No.', 'Student', 'Type', 'Category', 'Destination',
                  'Departure', 'Exp. Return', 'Actual Return', 'Status', 'Overstayed']
    gp_rows_xl = [[i, gp.student.roll_number, gp.student.name, gp.get_pass_type_display(),
                   gp.category.name if gp.category else '—', gp.destination,
                   str(gp.departure_date), str(gp.expected_return_date),
                   str(gp.actual_return_date) if gp.actual_return_date else '—',
                   gp.get_status_display(), 'Yes' if gp.is_overstayed else 'No']
                  for i, gp in enumerate(gp_qs, 1)]
    add_sheet(wb, 'Gate Passes', gp_headers, gp_rows_xl)

    # Fees
    fee_headers = ['#', 'Roll No.', 'Student', 'Fee Structure', 'Total (₹)', 'Paid (₹)', 'Balance (₹)', 'Status']
    fee_rows_xl = [[i, r.student.roll_number, r.student.name, str(r.fee_structure),
                    float(r.total_amount), float(r.amount_paid), float(r.balance), r.get_status_display()]
                   for i, r in enumerate(fee_qs, 1)]
    add_sheet(wb, 'Pending Fees', fee_headers, fee_rows_xl)

    # Complaints
    comp_headers = ['#', 'Title', 'Raised By', 'Status', 'Date']
    comp_rows_xl = [[i, c.title, c.raised_by.name if c.raised_by else '—',
                     c.get_status_display(), str(c.created_at.date())]
                    for i, c in enumerate(complaints_qs, 1)]
    add_sheet(wb, 'Open Complaints', comp_headers, comp_rows_xl)

    # Mess
    mess_headers = ['#', 'Meal Type', 'Avg Rating', 'Responses', 'Wastage (kg)']
    meal_types = ['breakfast', 'lunch', 'dinner', 'snacks']
    mess_rows_xl = []
    for mi, mt in enumerate(meal_types, 1):
        fb = feedback_qs.filter(meal_type=mt)
        avg = fb.aggregate(a=Avg('rating'))['a'] or 0
        wst = wastage_qs.filter(meal_type=mt).aggregate(t=Sum('wastage_kg'))['t'] or 0
        mess_rows_xl.append([mi, mt.title(), round(avg, 2), fb.count(), float(wst)])
    add_sheet(wb, 'Mess Summary', mess_headers, mess_rows_xl, '059669')

    buf = _io.BytesIO()
    wb.save(buf)
    resp = _xlsx_response(f'{hostel.name}_daily_report_{report_date}.xlsx')
    resp.write(buf.getvalue())
    return resp
